import json
import os
import re
import subprocess
import sys
import uuid
from copy import copy, deepcopy
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QEvent, QObject, QModelIndex, QPoint, QPointF, QRect, QRectF, Qt, QTimer
from PySide6.QtGui import (
    QBrush,
    QColor,
    QFont,
    QFontMetrics,
    QFontMetricsF,
    QGuiApplication,
    QPainter,
    QPainterPath,
    QPen,
    QPixmap,
    QPalette,
    QPolygon,
    QPolygonF,
)

from bill_constants import (
    ALLOW_PRINT_SCROLL_COL,
    ALLOW_PRINT_URL_DISPLAY_COL,
    CITIES,
    FROZEN_COLUMNS,
    DISPLAY_FIELDS,
    DURATIONS,
    FIELDS,
    HEADERS,
    HISTORY_SCROLL_COLUMNS,
    HISTORY_SCROLL_FIELDS,
    MAIN_SCROLL_COLUMNS,
    OP_MAP,
    PRINT_LOG_COL_COUNT,
    PRINT_LOG_DATA_FIELDS,
    PRINT_LOG_HEADERS,
    PROVINCE_TO_CITIES,
    PROVINCES,
    TYPE_MAP,
    coerce_allow_print_url,
    find_earliest_region_in_left,
    first_url,
    int_to_cn,
    region_in_task_name_requires_customer_prefix,
    sanitize_filename,
    split_multi,
    cities_under_provinces,
    normalize_city_name,
)
from bill_paths import (
    ACCESSORIES_FILE,
    APP_DIR,
    DATA_FILE,
    HISTORY_FILE,
    LICENSE_EXPIRED_MSG,
    PICKER_RECENT_FILE,
    PRINT_RECORDS_FILE,
    RECEIPT_DATA_FILE,
    SUM_TEMPLATE_FILE,
    TEMPLATE_FILE,
    THEME_FILE,
    is_license_valid,
)
from bill_theme import STYLESHEET_DARK, STYLESHEET_LIGHT
from bill_widgets import (
    AccessoryGraphView,
    AllowPrintUrlCellDelegate,
    BadgeCellDelegate,
    ColumnPickFilterPopup,
    HoverFilterHeaderView,
    MultiSelectDialog,
    UrlCellEditor,
    make_sidebar_logo_pixmap,
    style_combo_centered,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFormLayout,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMenu,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QProgressDialog,
    QSizePolicy,
    QStatusBar,
    QStackedWidget,
    QTreeWidget,
    QTreeWidgetItem,
    QTreeView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QListView,
)

# 主表行「格式 + 导出必填」校验涉及的字段（输入/离行仅标红；点击打印校验失败时弹窗）
MAIN_ROW_VALIDATE_FIELDS = (
    "task_name",
    "type_code",
    "operator_code",
    "industry_code",
    "url",
    "quantity",
    "duration",
    "age_max",
    "age_min",
    "pv",
    "province",
    "city",
)


class BillApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setObjectName("BillAppMain")
        self.setWindowTitle("高效办公")
        self.records: list[dict[str, Any]] = []
        self.current_business = "bill"
        self.history_records: list[dict[str, Any]] = []
        self.history_filtered_indices: list[int] = []
        self.filtered_indices: list[int] = []
        self.theme = self.load_theme()
        self.updating_table = False
        self.select_all_state = False
        self.field_errors: dict[tuple[int, str], bool] = {}
        self.field_error_msgs: dict[tuple[int, str], str] = {}
        self.picker_recent: dict[str, list[str]] = {}
        self.header_filters: dict[str, set[str]] = {}
        self.header_sort_field: str | None = None
        self.header_sort_order = Qt.SortOrder.AscendingOrder
        self._table_v_sync = False
        self._table_sel_sync = False
        self.history_header_filters: dict[str, set[str]] = {}
        self.history_sort_field: str | None = None
        self.history_sort_order = Qt.SortOrder.AscendingOrder
        self.print_records: list[dict[str, Any]] = []
        self.accessories_root: dict[str, Any] = {}
        self.print_log_filtered_indices: list[int] = []
        self.print_header_filters: dict[str, set[str]] = {}
        self.print_sort_field: str | None = "printed_at"
        self.print_sort_order = Qt.SortOrder.DescendingOrder
        self._hist_v_sync = False
        self._hist_sel_sync = False
        self._acc_v_sync = False
        self._acc_sel_sync = False
        self._acc_list_updating = False
        self._accessory_checked_node_ids: set[str] = set()
        self._accessory_draft_rows: list[dict[str, str]] = []
        self._accessory_selected_node_id = "root"
        self.accessory_header_filters: dict[str, set[str]] = {}
        self.accessory_sort_field: str | None = None
        self.accessory_sort_order = Qt.SortOrder.AscendingOrder
        self._filter_popup: ColumnPickFilterPopup | None = None
        self._license_timer = QTimer(self)
        self._license_timer.timeout.connect(self._check_license_and_exit_if_needed)
        self._license_timer.start(10_000)
        QTimer.singleShot(0, self._check_license_and_exit_if_needed)
        # 须在 _setup_ui 之前：建表时会触发列宽/URL 行高回调，依赖本属性
        _fm0 = QFontMetrics(self.font())
        self._data_row_height = max(44, int((_fm0.height() + max(_fm0.leading(), 0)) * 1.5) + 14)
        self._ensure_output_dirs()
        self._setup_ui()
        self.load_data()
        self.load_history_data()
        self.load_print_records()
        self.load_accessories()
        self.load_picker_recent()
        self.apply_theme()
        fm = QFontMetrics(self.font())
        self._data_row_height = max(44, int((fm.height() + max(fm.leading(), 0)) * 1.5) + 14)
        self.refresh_table()
        # 刷新后子控件尺寸提示已稳定，再套一次初始几何（避免宽表把窗口最小宽度撑满）
        self._apply_initial_window_geometry()
        cw = self.centralWidget()
        if cw is not None:
            cw.setMinimumWidth(0)

    def _check_license_and_exit_if_needed(self):
        if is_license_valid():
            return
        self._license_timer.stop()
        _show_license_expired_dialog(self)
        QApplication.quit()
        sys.exit(0)

    def _setup_ui(self):
        center = QWidget(self)
        center.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        self.setCentralWidget(center)
        root = QHBoxLayout(center)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setAutoFillBackground(True)
        self.sidebar.setFixedWidth(240)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(10, 16, 10, 20)
        sidebar_layout.setSpacing(8)
        self.sidebar_brand = QWidget()
        self.sidebar_brand.setObjectName("sidebarBrand")
        brand_row = QHBoxLayout(self.sidebar_brand)
        brand_row.setContentsMargins(0, 0, 0, 0)
        brand_row.setSpacing(12)
        self._sidebar_logo_icon = QLabel()
        self._sidebar_logo_icon.setFixedSize(44, 44)
        brand_titles = QVBoxLayout()
        brand_titles.setContentsMargins(0, 1, 0, 0)
        brand_titles.setSpacing(5)
        self._sidebar_logo_title = QLabel("高效办公")
        self._sidebar_logo_title.setObjectName("sidebarLogoTitle")
        self._sidebar_logo_sub = QLabel("＜初試人生＞")
        self._sidebar_logo_sub.setObjectName("sidebarLogoSub")
        brand_titles.addWidget(self._sidebar_logo_title)
        brand_titles.addWidget(self._sidebar_logo_sub)
        brand_row.addWidget(self._sidebar_logo_icon, 0, Qt.AlignmentFlag.AlignTop)
        brand_row.addLayout(brand_titles, 1)
        self.nav_bill = QPushButton("📝  提单表")
        self.nav_bill.setObjectName("navActive")
        self.nav_bill.clicked.connect(self.show_bill_page)
        self.nav_receipt = QPushButton("📬  回执管理")
        self.nav_receipt.setObjectName("navNormal")
        self.nav_receipt.clicked.connect(self.show_receipt_page)
        self.nav_accessories = QPushButton("📎  配件表")
        self.nav_accessories.setObjectName("navNormal")
        self.nav_accessories.clicked.connect(self.show_accessories_page)
        self.nav_history = QPushButton("🗂  历史任务")
        self.nav_history.setObjectName("navNormal")
        self.nav_history.clicked.connect(self.show_history_page)
        self.nav_print_records = QPushButton("🖨  历史打印记录")
        self.nav_print_records.setObjectName("navNormal")
        self.nav_print_records.clicked.connect(self.show_print_records_page)
        self.nav_excel_tools = QPushButton("🧰  Excel工具")
        self.nav_excel_tools.setObjectName("navNormal")
        self.nav_excel_tools.clicked.connect(self.show_excel_tools_page)
        self.nav_settings = QPushButton("⚙️  设置")
        self.nav_settings.setObjectName("navNormal")
        self.nav_settings.clicked.connect(self.show_settings_page)
        divider = QFrame()
        divider.setObjectName("sideDivider")
        divider.setFrameShape(QFrame.HLine)
        sidebar_layout.addWidget(self.sidebar_brand)
        sidebar_layout.addWidget(self.nav_bill)
        sidebar_layout.addWidget(self.nav_receipt)
        sidebar_layout.addWidget(self.nav_history)
        sidebar_layout.addWidget(self.nav_print_records)
        sidebar_layout.addWidget(self.nav_accessories)
        sidebar_layout.addWidget(self.nav_excel_tools)
        sidebar_layout.addWidget(divider)
        sidebar_layout.addWidget(self.nav_settings)
        sidebar_layout.addStretch(1)
        root.addWidget(self.sidebar)

        main_wrap = QWidget()
        main_wrap.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        main_layout = QVBoxLayout(main_wrap)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        self.content_stack = QStackedWidget()
        self.content_stack.setObjectName("contentStack")
        self.content_stack.setAutoFillBackground(True)
        # 水平 Ignored：避免 QStackedWidget 与各页「最宽子控件」累加成超大窗口最小宽度
        self.content_stack.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        main_layout.addWidget(self.content_stack, 1)

        bill_page = QWidget()
        bill_page.setObjectName("stackBillPage")
        bill_page.setAutoFillBackground(True)
        bill_page.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        bill_page_layout = QVBoxLayout(bill_page)
        bill_page_layout.setContentsMargins(0, 0, 0, 0)
        bill_page_layout.setSpacing(0)
        top = QHBoxLayout()
        top.setContentsMargins(16, 12, 16, 12)
        top.setSpacing(12)
        self.btn_toggle_sidebar = QPushButton("📂 菜单")
        self.btn_toggle_sidebar.setObjectName("btnGhost")
        self.btn_toggle_sidebar.clicked.connect(self.toggle_sidebar)
        self.search = QLineEdit()
        self.search.setPlaceholderText("搜索任务名...")
        self.search.setFixedWidth(220)
        self.search.textChanged.connect(self.refresh_table)
        btn_search = QPushButton("🔍 搜索")
        btn_search.setObjectName("btnGhost")
        btn_search.clicked.connect(self.on_bill_search_clicked)
        btn_clear = QPushButton("🔄 清空筛选")
        btn_clear.setObjectName("btnGhost")
        btn_clear.clicked.connect(self.clear_bill_search_and_filters)
        self.chk_print_url = QCheckBox("是否打印URL")
        self.lbl_hint = QLabel(
            "未勾选：导出文件中 URL 列为空。勾选：仅「允许」为是的行写入 URL。"
        )
        self.lbl_hint.setObjectName("hintLabel")
        # 说明在复选框下方一行区域展示；窄宽度时允许折行
        self.lbl_hint.setWordWrap(True)
        self.lbl_hint.setMinimumWidth(0)
        self.lbl_hint.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.btn_print = QPushButton("🖨 打印")
        self.btn_print.setObjectName("btnAccent")
        self.btn_print.clicked.connect(self.export_excel)
        self.btn_import_excel = QPushButton("📥 导入Excel")
        self.btn_import_excel.setObjectName("btnGhost")
        self.btn_import_excel.clicked.connect(self.import_bill_excel)
        self.btn_add = QPushButton("➕ 新增行")
        self.btn_add.setObjectName("btnSuccess")
        self.btn_add.clicked.connect(self.add_row)
        self.btn_bulk_fill = QPushButton("📋 批量填充")
        self.btn_bulk_fill.setObjectName("btnAccent")
        self.btn_bulk_fill.clicked.connect(self.bulk_fill_selected)
        self.btn_delete_sel = QPushButton("🗑 删除选中")
        self.btn_delete_sel.setObjectName("btnDanger")
        self.btn_delete_sel.clicked.connect(self.delete_selected)
        print_opt_layout = QVBoxLayout()
        print_opt_layout.setContentsMargins(0, 0, 0, 0)
        print_opt_layout.setSpacing(2)
        print_opt_layout.addWidget(
            self.chk_print_url, 0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        )
        print_opt_layout.addWidget(self.lbl_hint, 0)
        sep1 = QFrame()
        sep1.setObjectName("toolbarSep")
        sep1.setFrameShape(QFrame.VLine)
        sep2 = QFrame()
        sep2.setObjectName("toolbarSep")
        sep2.setFrameShape(QFrame.VLine)
        for w in [self.btn_toggle_sidebar, self.search, btn_search, btn_clear]:
            top.addWidget(w)
        top.addWidget(sep1)
        top.addLayout(print_opt_layout, 1)
        top.addWidget(sep2)
        for w in [self.btn_print, self.btn_import_excel, self.btn_add, self.btn_bulk_fill]:
            top.addWidget(w)
        top.addStretch(1)
        bill_page_layout.addLayout(top)

        table_split = QWidget()
        table_split.setObjectName("tableSplit")
        table_split.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        split_lo = QHBoxLayout(table_split)
        split_lo.setContentsMargins(0, 0, 0, 0)
        split_lo.setSpacing(0)
        self.table_frozen = QTableWidget(0, FROZEN_COLUMNS)
        self.table_frozen.setObjectName("tableFrozenCol")
        self.table_frozen.setHorizontalHeaderLabels([HEADERS["checked"], HEADERS["task_name"]])
        self.table_frozen.verticalHeader().setVisible(False)
        self.table_frozen.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_frozen.setAlternatingRowColors(True)
        self.table_frozen.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        # 与右侧主表同时保留横向滚动条占位，避免仅一侧出现横向条时纵向视口高度不一致导致滚动错位
        self.table_frozen.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table_frozen.horizontalScrollBar().setEnabled(False)
        self.table_frozen.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.table_frozen.setFrameShape(QFrame.NoFrame)
        self.table_frozen.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)
        self.table_frozen.cellClicked.connect(lambda r, c: self.on_cell_clicked(r, c))
        self.table_frozen.cellDoubleClicked.connect(lambda r, c: self.on_cell_double_click(r, c))
        self.table = QTableWidget(0, MAIN_SCROLL_COLUMNS)
        self.table.setObjectName("tableScrollPart")
        scroll_headers = [HEADERS[x] for x in DISPLAY_FIELDS[FROZEN_COLUMNS:]]
        self.table.setHorizontalHeaderLabels(scroll_headers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.itemChanged.connect(self.on_item_changed)
        self.table.cellDoubleClicked.connect(lambda r, c: self.on_cell_double_click(r, c + FROZEN_COLUMNS))
        self.table.cellClicked.connect(lambda r, c: self.on_cell_clicked(r, c + FROZEN_COLUMNS))
        self.table.currentCellChanged.connect(
            lambda cr, cc, pr, pc: self.on_current_cell_changed(
                cr, cc + FROZEN_COLUMNS if cc >= 0 else -1, pr, pc
            )
        )
        self.table.customContextMenuRequested.connect(self.on_table_context_menu)
        hdr_main_f = HoverFilterHeaderView(self.table_frozen, self, "main_frozen")
        hdr_main_s = HoverFilterHeaderView(self.table, self, "main_scroll")
        self.table_frozen.setHorizontalHeader(hdr_main_f)
        self.table.setHorizontalHeader(hdr_main_s)
        hdr_main_f.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr_main_f.setSectionResizeMode(1, QHeaderView.Interactive)
        hdr_main_s.setSectionResizeMode(QHeaderView.Interactive)
        self.table.setItemDelegateForColumn(ALLOW_PRINT_SCROLL_COL, AllowPrintUrlCellDelegate(self.table))
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        # 列总宽很大时，勿把整表宽度当作窗口最小宽度，否则初始宽度无效且无法横向缩小
        self.table.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        hdr_main_f.sectionClicked.connect(self._on_main_frozen_header_section_clicked)
        hdr_main_s.sectionClicked.connect(self._on_main_scroll_header_clicked)
        self.table.horizontalHeader().sectionResized.connect(self._on_main_scroll_column_resized_for_url)
        self.table_frozen.horizontalHeader().sectionResized.connect(self._on_main_frozen_section_resized)
        self.table.verticalScrollBar().valueChanged.connect(lambda v: self._sync_main_v_scroll(v, "scroll"))
        self.table_frozen.verticalScrollBar().valueChanged.connect(lambda v: self._sync_main_v_scroll(v, "frozen"))
        self.table.itemSelectionChanged.connect(self._sync_selection_main_to_frozen)
        self.table_frozen.itemSelectionChanged.connect(self._sync_selection_frozen_to_main)
        split_lo.addWidget(self.table_frozen, 0)
        split_lo.addWidget(self.table, 1)
        table_wrap = QWidget()
        table_wrap.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        table_layout = QVBoxLayout(table_wrap)
        table_layout.setContentsMargins(16, 10, 16, 16)
        table_layout.setSpacing(8)
        table_layout.addWidget(table_split, 1)
        bottom_actions = QHBoxLayout()
        bottom_actions.addWidget(self.btn_delete_sel)
        bottom_actions.addStretch(1)
        table_layout.addLayout(bottom_actions)
        bill_page_layout.addWidget(table_wrap, 1)
        self.content_stack.addWidget(bill_page)

        history_page = QWidget()
        history_page.setObjectName("stackHistoryPage")
        history_page.setAutoFillBackground(True)
        history_page.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        history_layout = QVBoxLayout(history_page)
        history_layout.setContentsMargins(16, 12, 16, 16)
        history_layout.setSpacing(10)
        history_top = QHBoxLayout()
        self.history_search = QLineEdit()
        self.history_search.setPlaceholderText("搜索任务名...")
        self.history_search.setFixedWidth(220)
        self.history_search.textChanged.connect(self.refresh_history_table)
        btn_history_search = QPushButton("🔍 搜索")
        btn_history_search.setObjectName("btnGhost")
        btn_history_search.clicked.connect(self.on_history_search_clicked)
        btn_history_clear = QPushButton("🔄 清空筛选")
        btn_history_clear.setObjectName("btnGhost")
        btn_history_clear.clicked.connect(self.clear_history_search_and_filters)
        history_top.addWidget(self.history_search)
        history_top.addWidget(btn_history_search)
        history_top.addWidget(btn_history_clear)
        history_top.addStretch(1)
        self.btn_restore_selected = QPushButton("↩ 恢复选中")
        self.btn_restore_selected.setObjectName("btnAccent")
        self.btn_restore_selected.clicked.connect(self.restore_selected_history)
        history_top.addWidget(self.btn_restore_selected)
        self.btn_history_delete_selected = QPushButton("🗑 删除选中")
        self.btn_history_delete_selected.setObjectName("btnDanger")
        self.btn_history_delete_selected.clicked.connect(self.delete_selected_history)
        history_top.addWidget(self.btn_history_delete_selected)
        history_layout.addLayout(history_top)
        hist_split = QWidget()
        hist_split.setObjectName("historyTableSplit")
        hist_split.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        hist_lo = QHBoxLayout(hist_split)
        hist_lo.setContentsMargins(0, 0, 0, 0)
        hist_lo.setSpacing(0)
        self.history_table_frozen = QTableWidget(0, FROZEN_COLUMNS)
        self.history_table_frozen.setObjectName("tableFrozenCol")
        self.history_table_frozen.setHorizontalHeaderLabels([HEADERS["checked"], HEADERS["task_name"]])
        self.history_table_frozen.verticalHeader().setVisible(False)
        self.history_table_frozen.setSelectionBehavior(QTableWidget.SelectRows)
        self.history_table_frozen.setAlternatingRowColors(True)
        self.history_table_frozen.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.history_table_frozen.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.history_table_frozen.horizontalScrollBar().setEnabled(False)
        self.history_table_frozen.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.history_table_frozen.setFrameShape(QFrame.NoFrame)
        self.history_table_frozen.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)
        self.history_table_frozen.cellClicked.connect(lambda r, c: self.on_history_cell_clicked(r, c))
        self.history_table = QTableWidget(0, HISTORY_SCROLL_COLUMNS)
        self.history_table.setObjectName("historyScrollPart")
        history_scroll_headers = [HEADERS[f] for f in HISTORY_SCROLL_FIELDS if f != "deleted_at"] + ["删除时间"]
        self.history_table.setHorizontalHeaderLabels(history_scroll_headers)
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.cellClicked.connect(lambda r, c: self.on_history_cell_clicked(r, c + FROZEN_COLUMNS))
        self.history_table.setEditTriggers(QTableWidget.NoEditTriggers)
        hdr_hist_f = HoverFilterHeaderView(self.history_table_frozen, self, "hist_frozen")
        hdr_hist_s = HoverFilterHeaderView(self.history_table, self, "hist_scroll")
        self.history_table_frozen.setHorizontalHeader(hdr_hist_f)
        self.history_table.setHorizontalHeader(hdr_hist_s)
        hdr_hist_f.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr_hist_f.setSectionResizeMode(1, QHeaderView.Interactive)
        hdr_hist_s.setSectionResizeMode(QHeaderView.Interactive)
        self.history_table.setItemDelegateForColumn(ALLOW_PRINT_SCROLL_COL, AllowPrintUrlCellDelegate(self.history_table))
        self.history_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.history_table.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        hdr_hist_f.sectionClicked.connect(self._on_hist_frozen_header_section_clicked)
        hdr_hist_s.sectionClicked.connect(self._on_history_scroll_header_clicked)
        self.history_table_frozen.horizontalHeader().sectionResized.connect(self._on_hist_frozen_section_resized)
        self.history_table.verticalScrollBar().valueChanged.connect(lambda v: self._sync_hist_v_scroll(v, "scroll"))
        self.history_table_frozen.verticalScrollBar().valueChanged.connect(lambda v: self._sync_hist_v_scroll(v, "frozen"))
        self.history_table.itemSelectionChanged.connect(self._sync_selection_hist_to_frozen)
        self.history_table_frozen.itemSelectionChanged.connect(self._sync_selection_hist_frozen_to_scroll)
        hist_lo.addWidget(self.history_table_frozen, 0)
        hist_lo.addWidget(self.history_table, 1)
        history_layout.addWidget(hist_split, 1)
        self.content_stack.addWidget(history_page)

        print_log_page = QWidget()
        print_log_page.setObjectName("stackPrintRecordsPage")
        print_log_page.setAutoFillBackground(True)
        print_log_page.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        pl_layout = QVBoxLayout(print_log_page)
        pl_layout.setContentsMargins(16, 12, 16, 16)
        pl_layout.setSpacing(10)
        pl_top = QHBoxLayout()
        self.print_log_search = QLineEdit()
        self.print_log_search.setPlaceholderText("搜索文件名或路径...")
        self.print_log_search.setFixedWidth(260)
        self.print_log_search.textChanged.connect(self.refresh_print_records_table)
        btn_pl_search = QPushButton("🔍 搜索")
        btn_pl_search.setObjectName("btnGhost")
        btn_pl_search.clicked.connect(self.on_print_log_search_clicked)
        btn_pl_clear = QPushButton("🔄 清空筛选")
        btn_pl_clear.setObjectName("btnGhost")
        btn_pl_clear.clicked.connect(self.clear_print_log_search_and_filters)
        self.btn_print_log_delete = QPushButton("🗑 删除选中")
        self.btn_print_log_delete.setObjectName("btnDanger")
        self.btn_print_log_delete.clicked.connect(self.delete_selected_print_records)
        pl_top.addWidget(self.print_log_search)
        pl_top.addWidget(btn_pl_search)
        pl_top.addWidget(btn_pl_clear)
        pl_top.addStretch(1)
        pl_top.addWidget(self.btn_print_log_delete)
        pl_layout.addLayout(pl_top)

        pl_headers = [HEADERS["checked"]] + [PRINT_LOG_HEADERS[k] for k in PRINT_LOG_DATA_FIELDS] + ["操作"]
        self.print_log_table = QTableWidget(0, PRINT_LOG_COL_COUNT)
        self.print_log_table.setObjectName("historyScrollPart")
        self.print_log_table.setHorizontalHeaderLabels(pl_headers)
        self.print_log_table.verticalHeader().setVisible(False)
        self.print_log_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.print_log_table.setAlternatingRowColors(True)
        self.print_log_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.print_log_table.cellClicked.connect(self.on_print_log_cell_clicked)
        hdr_print = HoverFilterHeaderView(self.print_log_table, self, "print_rec")
        self.print_log_table.setHorizontalHeader(hdr_print)
        hdr_print.setSectionResizeMode(QHeaderView.Interactive)
        # 历史打印记录：来源(印/合)与状态(有/无)使用与「允许」列一致的圆形徽标风格
        self.print_log_table.setItemDelegateForColumn(
            2,
            BadgeCellDelegate(
                {"印"},
                self.print_log_table,
                {
                    "印": (QColor(233, 220, 252), QColor(205, 180, 245), QColor(109, 64, 170)),
                    "合": (QColor(214, 234, 255), QColor(171, 209, 247), QColor(42, 106, 181)),
                },
            ),
        )
        self.print_log_table.setItemDelegateForColumn(6, BadgeCellDelegate({"有"}, self.print_log_table))
        hdr_print.sectionClicked.connect(self._on_print_rec_header_section_clicked)
        self.print_log_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.print_log_table.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        pl_layout.addWidget(self.print_log_table, 1)
        self.content_stack.addWidget(print_log_page)

        excel_tools_page = QWidget()
        excel_tools_page.setObjectName("stackExcelToolsPage")
        excel_tools_page.setAutoFillBackground(True)
        excel_tools_page.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        excel_tools_layout = QVBoxLayout(excel_tools_page)
        excel_tools_layout.setContentsMargins(24, 20, 24, 20)
        excel_tools_layout.setSpacing(12)
        excel_tools_title = QLabel("Excel工具")
        excel_tools_title.setObjectName("settingsTitle")
        excel_tools_layout.addWidget(excel_tools_title)
        excel_tools_hint = QLabel("可扩展的 Excel 功能入口（后续可继续新增卡片按钮）。")
        excel_tools_hint.setObjectName("hintLabel")
        excel_tools_layout.addWidget(excel_tools_hint)
        self.btn_excel_merge = QPushButton("🧩 合并Excel")
        self.btn_excel_merge.setObjectName("btnAccent")
        self.btn_excel_merge.setFixedHeight(44)
        self.btn_excel_merge.clicked.connect(self.open_merge_excel_dialog)
        excel_tools_layout.addWidget(self.btn_excel_merge, 0, Qt.AlignmentFlag.AlignLeft)
        excel_tools_layout.addStretch(1)
        self.content_stack.addWidget(excel_tools_page)

        accessory_page = QWidget()
        accessory_page.setObjectName("stackAccessoryPage")
        accessory_page.setAutoFillBackground(True)
        accessory_page.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        ac_layout = QVBoxLayout(accessory_page)
        ac_layout.setContentsMargins(16, 12, 16, 16)
        ac_layout.setSpacing(10)
        ac_top = QHBoxLayout()
        self.accessory_search = QLineEdit()
        self.accessory_search.setPlaceholderText("搜索节点：名称/描述/URL...")
        self.accessory_search.setFixedWidth(280)
        self.accessory_search.textChanged.connect(self.refresh_accessory_tree)
        btn_acc_clear = QPushButton("🔄 清空搜索")
        btn_acc_clear.setObjectName("btnGhost")
        btn_acc_clear.clicked.connect(self.clear_accessory_search_and_filters)
        self.btn_acc_add = QPushButton("➕ 新增节点")
        self.btn_acc_add.setObjectName("btnSuccess")
        self.btn_acc_add.clicked.connect(self.add_accessory)
        self.btn_acc_add_row = QPushButton("➕ 新增行")
        self.btn_acc_add_row.setObjectName("btnSuccess")
        self.btn_acc_add_row.clicked.connect(self.add_accessory_row)
        self.btn_acc_import = QPushButton("📥 导入Excel")
        self.btn_acc_import.setObjectName("btnGhost")
        self.btn_acc_import.clicked.connect(self.import_accessories_from_excel)
        self.btn_acc_copy_leaf = QPushButton("📄 复制叶子")
        self.btn_acc_copy_leaf.setObjectName("btnGhost")
        self.btn_acc_copy_leaf.clicked.connect(self.copy_accessory_leaf)
        self.btn_acc_copy_leaf.hide()
        self.btn_acc_delete = QPushButton("🗑 删除节点")
        self.btn_acc_delete.setObjectName("btnDanger")
        self.btn_acc_delete.clicked.connect(self.delete_accessory)
        self.btn_acc_delete_selected = QPushButton("🗑 删除选中")
        self.btn_acc_delete_selected.setObjectName("btnDanger")
        self.btn_acc_delete_selected.clicked.connect(self.delete_accessory_selected)
        self.accessory_mode_combo = QComboBox()
        self.accessory_mode_combo.addItem("树形结构", "tree")
        self.accessory_mode_combo.addItem("列表结构", "list")
        self.accessory_mode_combo.currentIndexChanged.connect(self.on_accessory_mode_changed)
        ac_top.addWidget(self.accessory_search)
        ac_top.addWidget(btn_acc_clear)
        ac_top.addWidget(self.btn_acc_add)
        ac_top.addWidget(self.btn_acc_add_row)
        ac_top.addWidget(self.btn_acc_import)
        ac_top.addWidget(self.btn_acc_copy_leaf)
        ac_top.addWidget(self.btn_acc_delete_selected)
        ac_top.addWidget(self.btn_acc_delete)
        ac_top.addStretch(1)
        ac_top.addWidget(self.accessory_mode_combo)
        ac_layout.addLayout(ac_top)
        self.accessory_view_stack = QStackedWidget()
        self.accessory_graph = AccessoryGraphView()
        self.accessory_graph.nodeSelected.connect(self._on_accessory_graph_selected)
        self.accessory_graph.nodeDoubleClicked.connect(lambda _nid: self.update_accessory())
        self.accessory_list_split = QWidget()
        self.accessory_list_split.setObjectName("historyTableSplit")
        acc_split_lo = QHBoxLayout(self.accessory_list_split)
        acc_split_lo.setContentsMargins(0, 0, 0, 0)
        acc_split_lo.setSpacing(0)
        self.accessory_list_frozen = QTableWidget(0, 2)
        self.accessory_list_frozen.setObjectName("tableFrozenCol")
        self.accessory_list_frozen.setHorizontalHeaderLabels(["☐", "类型"])
        self.accessory_list_frozen.verticalHeader().setVisible(False)
        self.accessory_list_frozen.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.accessory_list_frozen.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.accessory_list_frozen.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.SelectedClicked | QTableWidget.EditTrigger.EditKeyPressed)
        self.accessory_list_frozen.setAlternatingRowColors(True)
        self.accessory_list_frozen.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.accessory_list_frozen.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.accessory_list_frozen.horizontalScrollBar().setEnabled(False)
        self.accessory_list_frozen.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)
        self.accessory_list_table = QTableWidget(0, 7)
        self.accessory_list_table.setObjectName("historyScrollPart")
        self.accessory_list_table.setHorizontalHeaderLabels(["渠道", "名称", "描述", "URL", "备注", "创建时间", "操作"])
        self.accessory_list_table.verticalHeader().setVisible(False)
        self.accessory_list_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.accessory_list_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.accessory_list_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.SelectedClicked | QTableWidget.EditTrigger.EditKeyPressed)
        self.accessory_list_table.setAlternatingRowColors(True)
        fm_acc = QFontMetrics(self.font())
        w_type = max(100, fm_acc.horizontalAdvance("[类目一123]") + 24)
        self.accessory_list_frozen.setColumnWidth(0, 42)
        self.accessory_list_frozen.setColumnWidth(1, w_type)
        self.accessory_list_table.setColumnWidth(0, max(100, fm_acc.horizontalAdvance("渠道一123") + 24))
        self.accessory_list_table.setColumnWidth(1, max(150, fm_acc.horizontalAdvance("名称一123456") + 28))
        self.accessory_list_table.setColumnWidth(2, max(150, fm_acc.horizontalAdvance("描述一123456") + 28))
        self.accessory_list_table.setColumnWidth(3, max(200, fm_acc.horizontalAdvance("https://example.com/xxx") + 28))
        self.accessory_list_table.setColumnWidth(4, max(100, fm_acc.horizontalAdvance("备注一123") + 24))
        self.accessory_list_table.setColumnWidth(5, 150)
        self.accessory_list_table.setColumnWidth(6, 100)
        hdr_acc_f = HoverFilterHeaderView(self.accessory_list_frozen, self, "acc_frozen")
        hdr_acc_s = HoverFilterHeaderView(self.accessory_list_table, self, "acc_scroll")
        self.accessory_list_frozen.setHorizontalHeader(hdr_acc_f)
        self.accessory_list_table.setHorizontalHeader(hdr_acc_s)
        hdr_acc_f.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr_acc_f.setSectionResizeMode(1, QHeaderView.Interactive)
        hdr_acc_s.setSectionResizeMode(QHeaderView.Interactive)
        hdr_acc_f.sectionClicked.connect(self._on_accessory_frozen_header_clicked)
        hdr_acc_s.sectionClicked.connect(self._on_accessory_scroll_header_clicked)
        self.accessory_list_table.verticalScrollBar().valueChanged.connect(lambda v: self._sync_accessory_v_scroll(v, "scroll"))
        self.accessory_list_frozen.verticalScrollBar().valueChanged.connect(lambda v: self._sync_accessory_v_scroll(v, "frozen"))
        self.accessory_list_table.itemSelectionChanged.connect(self._sync_selection_accessory_to_frozen)
        self.accessory_list_frozen.itemSelectionChanged.connect(self._sync_selection_accessory_frozen_to_scroll)
        self.accessory_list_frozen.cellClicked.connect(self._on_accessory_list_frozen_cell_clicked)
        self.accessory_list_frozen.itemSelectionChanged.connect(self._on_accessory_list_selection_changed)
        self.accessory_list_table.itemSelectionChanged.connect(self._on_accessory_list_selection_changed)
        self.accessory_list_frozen.itemChanged.connect(self._on_accessory_list_item_changed)
        self.accessory_list_table.itemChanged.connect(self._on_accessory_list_item_changed)
        self.accessory_list_frozen.horizontalHeader().sectionResized.connect(self._on_accessory_frozen_section_resized)
        self.accessory_list_frozen.horizontalHeader().sectionResized.connect(self._save_accessory_list_col_widths)
        self.accessory_list_table.horizontalHeader().sectionResized.connect(self._save_accessory_list_col_widths)
        self.accessory_list_table.cellDoubleClicked.connect(self._on_accessory_list_cell_double_clicked)
        self.accessory_list_row_node_ids: list[str] = []
        self._load_accessory_list_col_widths()
        acc_split_lo.addWidget(self.accessory_list_frozen, 0)
        acc_split_lo.addWidget(self.accessory_list_table, 1)
        self.accessory_view_stack.addWidget(self.accessory_graph)
        self.accessory_view_stack.addWidget(self.accessory_list_split)
        ac_layout.addWidget(self.accessory_view_stack, 1)
        self.accessory_mode_combo.setCurrentIndex(1)
        self.content_stack.addWidget(accessory_page)

        settings_page = QWidget()
        settings_page.setObjectName("stackSettingsPage")
        settings_page.setAutoFillBackground(True)
        settings_layout = QVBoxLayout(settings_page)
        settings_layout.setContentsMargins(24, 24, 24, 24)
        settings_layout.setSpacing(10)
        settings_title = QLabel("设置")
        settings_title.setObjectName("settingsTitle")
        settings_layout.addWidget(settings_title)
        settings_layout.addWidget(QLabel("主题模式"))
        self.btn_theme = QPushButton("🎨 切换主题")
        self.btn_theme.setObjectName("btnGhost")
        self.btn_theme.clicked.connect(self.toggle_theme)
        settings_layout.addWidget(self.btn_theme)
        settings_layout.addStretch(1)
        self.content_stack.addWidget(settings_page)

        status = QStatusBar(self)
        self.setStatusBar(status)
        self.lbl_count = QLabel("任务数: 0")
        self.lbl_time = QLabel()
        self.lbl_save = QLabel("已自动保存")
        status.addPermanentWidget(self.lbl_count)
        status.addPermanentWidget(QLabel("|"))
        status.addPermanentWidget(self.lbl_time)
        status.addPermanentWidget(QLabel("|"))
        status.addPermanentWidget(self.lbl_save)
        status.addPermanentWidget(QLabel("|"))
        self.lbl_version = QLabel("Version:1.0.0 Copyright © 2026 Sprain")
        status.addPermanentWidget(self.lbl_version)
        timer = QTimer(self)
        timer.timeout.connect(self.update_time)
        timer.start(1000)
        self.update_time()
        root.addWidget(main_wrap, 1)
        self.show_bill_page()

        # 与滚动区列顺序一致：类型…行业、「允许」、URL…；长度须 >= MAIN_SCROLL_COLUMNS+1（含 col_widths[0] 给冻结任务名列）
        # 滚动区列序与 col_widths[1:] 对齐；末四列为 提单时间、打印次数、打印时间、操作（原打印相关宽顺序已随列序调整）
        col_widths = [260, 150, 150, 120, 72, 260, 90, 130, 90, 90, 90, 160, 160, 160, 160, 168, 88, 168, 128]
        self.table_frozen.setColumnWidth(0, 42)
        self.table_frozen.setColumnWidth(1, col_widths[0])
        for s in range(MAIN_SCROLL_COLUMNS):
            self.table.setColumnWidth(s, col_widths[s + 1])
        self.history_table_frozen.setColumnWidth(0, 42)
        self.history_table_frozen.setColumnWidth(1, col_widths[0])
        hist_widths = col_widths + [160]
        for s in range(HISTORY_SCROLL_COLUMNS):
            self.history_table.setColumnWidth(s, hist_widths[s + 1])
        self.print_log_table.setColumnWidth(0, 44)
        _pw = [280, 92, 168, 88, 100, 88, 380, 228]
        for s, w in enumerate(_pw, start=1):
            self.print_log_table.setColumnWidth(s, w)
        self.table_frozen.itemChanged.connect(self.on_item_changed)
        self._apply_initial_window_geometry()
        self._on_main_frozen_section_resized()
        self._on_hist_frozen_section_resized()
        self._on_accessory_frozen_section_resized()

    def _apply_initial_window_geometry(self):
        """初始窗口宽度 1500；高度为可用区估算值再加 50（无屏信息时 910）；窗口可自由拉伸。"""
        w, h = 1500, 910
        screen = QGuiApplication.primaryScreen()
        if screen is not None:
            ag = screen.availableGeometry()
            h = max(520, min(int(ag.height() * 0.88), ag.height() - 8)) + 50
            h = min(h, ag.height() - 8)
            x = ag.left() + max(0, (ag.width() - w) // 2)
            y = ag.top() + max(0, (ag.height() - h) // 2)
            self.setGeometry(x, y, w, h)
        # 统一再设客户区宽高，避免最小宽度约束导致实际宽度大于目标值
        self.resize(w, h)

    def showEvent(self, event):
        super().showEvent(event)
        if getattr(self, "_bill_post_show_geo", False):
            return
        self._bill_post_show_geo = True
        QTimer.singleShot(0, self._apply_initial_window_geometry)

    def _on_main_frozen_section_resized(self, *_args):
        """冻结区仅横向不滚动：勾选列固定宽，任务名列可调；总宽随两列之和更新。"""
        hf = self.table_frozen.horizontalHeader()
        hf.blockSignals(True)
        try:
            w0 = max(36, self.table_frozen.columnWidth(0))
            w1 = max(120, min(self.table_frozen.columnWidth(1), 2000))
            self.table_frozen.setColumnWidth(0, w0)
            self.table_frozen.setColumnWidth(1, w1)
            self.table_frozen.setFixedWidth(w0 + w1)
        finally:
            hf.blockSignals(False)

    def _on_hist_frozen_section_resized(self, *_args):
        hf = self.history_table_frozen.horizontalHeader()
        hf.blockSignals(True)
        try:
            w0 = max(36, self.history_table_frozen.columnWidth(0))
            w1 = max(120, min(self.history_table_frozen.columnWidth(1), 2000))
            self.history_table_frozen.setColumnWidth(0, w0)
            self.history_table_frozen.setColumnWidth(1, w1)
            self.history_table_frozen.setFixedWidth(w0 + w1)
        finally:
            hf.blockSignals(False)

    def _on_accessory_frozen_section_resized(self, *_args):
        if not hasattr(self, "accessory_list_frozen"):
            return
        hf = self.accessory_list_frozen.horizontalHeader()
        hf.blockSignals(True)
        try:
            w0 = 42
            w1 = max(120, min(self.accessory_list_frozen.columnWidth(1), 2000))
            self.accessory_list_frozen.setColumnWidth(0, w0)
            self.accessory_list_frozen.setColumnWidth(1, w1)
            self.accessory_list_frozen.setFixedWidth(w0 + w1)
        finally:
            hf.blockSignals(False)

    def load_theme(self):
        if THEME_FILE.exists():
            try:
                payload = json.loads(THEME_FILE.read_text(encoding="utf-8"))
                if isinstance(payload, dict):
                    self._ui_prefs = dict(payload)
                    return str(payload.get("theme", "light") or "light")
                self._ui_prefs = {}
                return "light"
            except Exception:
                self._ui_prefs = {}
                return "light"
        self._ui_prefs = {}
        return "light"

    def save_theme(self):
        payload = dict(getattr(self, "_ui_prefs", {}))
        payload["theme"] = self.theme
        THEME_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _load_accessory_list_col_widths(self):
        widths = (getattr(self, "_ui_prefs", {}) or {}).get("accessory_list_col_widths", [])
        if not isinstance(widths, list) or len(widths) not in (8, 9):
            return
        try:
            if len(widths) >= 9:
                self.accessory_list_frozen.setColumnWidth(0, int(widths[0]))
                self.accessory_list_frozen.setColumnWidth(1, int(widths[1]))
                start_idx = 2
                scroll_count = 7
            else:
                # 兼容旧配置：冻结区仅“类型”一列
                self.accessory_list_frozen.setColumnWidth(0, 42)
                self.accessory_list_frozen.setColumnWidth(1, int(widths[0]))
                start_idx = 1
                scroll_count = 7
            for i in range(scroll_count):
                w = int(widths[i + start_idx])
                # 末两列限制到更紧凑宽度，避免历史配置导致“创建时间/操作”过宽。
                if i == scroll_count - 2:
                    w = max(120, min(w, 160))
                elif i == scroll_count - 1:
                    w = max(90, min(w, 110))
                self.accessory_list_table.setColumnWidth(i, w)
            self._on_accessory_frozen_section_resized()
        except Exception:
            return

    def _save_accessory_list_col_widths(self, *_args):
        if not hasattr(self, "accessory_list_frozen"):
            return
        widths = [self.accessory_list_frozen.columnWidth(0), self.accessory_list_frozen.columnWidth(1)] + [
            self.accessory_list_table.columnWidth(i) for i in range(7)
        ]
        self._ui_prefs = dict(getattr(self, "_ui_prefs", {}))
        self._ui_prefs["accessory_list_col_widths"] = widths
        self.save_theme()

    def apply_theme(self):
        self.setStyleSheet(STYLESHEET_DARK if self.theme == "dark" else STYLESHEET_LIGHT)
        self._refresh_sidebar_logo()

    def _refresh_sidebar_logo(self):
        if not hasattr(self, "_sidebar_logo_icon"):
            return
        self._sidebar_logo_icon.setPixmap(make_sidebar_logo_pixmap(dark=self.theme == "dark", size=44))

    def toggle_theme(self):
        self.theme = "dark" if self.theme == "light" else "light"
        self.apply_theme()
        self.save_theme()

    def toggle_sidebar(self):
        visible = not self.sidebar.isVisible()
        self.sidebar.setVisible(visible)
        self.btn_toggle_sidebar.setText("📂 菜单" if visible else "📂 展开")

    def _restyle_sidebar_nav(self, page: str):
        styles = {
            "bill": ("navActive", "navNormal", "navNormal", "navNormal", "navNormal", "navNormal", "navNormal"),
            "receipt": ("navNormal", "navActive", "navNormal", "navNormal", "navNormal", "navNormal", "navNormal"),
            "history": ("navNormal", "navNormal", "navActive", "navNormal", "navNormal", "navNormal", "navNormal"),
            "print": ("navNormal", "navNormal", "navNormal", "navActive", "navNormal", "navNormal", "navNormal"),
            "excel_tools": ("navNormal", "navNormal", "navNormal", "navNormal", "navActive", "navNormal", "navNormal"),
            "accessory": ("navNormal", "navNormal", "navNormal", "navNormal", "navNormal", "navActive", "navNormal"),
            "settings": ("navNormal", "navNormal", "navNormal", "navNormal", "navNormal", "navNormal", "navActive"),
        }
        names = styles[page]
        for w, oname in zip(
            (
                self.nav_bill,
                self.nav_receipt,
                self.nav_history,
                self.nav_print_records,
                self.nav_excel_tools,
                self.nav_accessories,
                self.nav_settings,
            ),
            names
        ):
            w.setObjectName(oname)
            w.style().unpolish(w)
            w.style().polish(w)

    def show_bill_page(self):
        self._switch_business("bill")
        self.content_stack.setCurrentIndex(0)
        self._restyle_sidebar_nav("bill")
        self.apply_theme()

    def show_receipt_page(self):
        self._switch_business("receipt")
        self.content_stack.setCurrentIndex(0)
        self._restyle_sidebar_nav("receipt")
        self.apply_theme()

    def show_history_page(self):
        self.content_stack.setCurrentIndex(1)
        self._restyle_sidebar_nav("history")
        self.refresh_history_table()
        self.apply_theme()

    def show_print_records_page(self):
        self.content_stack.setCurrentIndex(2)
        self._restyle_sidebar_nav("print")
        self.refresh_print_records_table()
        self.apply_theme()

    def show_excel_tools_page(self):
        self.content_stack.setCurrentIndex(3)
        self._restyle_sidebar_nav("excel_tools")
        self.apply_theme()

    def show_settings_page(self):
        self.content_stack.setCurrentIndex(5)
        self._restyle_sidebar_nav("settings")
        self.apply_theme()

    def show_accessories_page(self):
        self.content_stack.setCurrentIndex(4)
        self._restyle_sidebar_nav("accessory")
        self.refresh_accessory_tree()
        self.apply_theme()

    def update_time(self):
        self.lbl_time.setText(datetime.now().strftime("%Y/%m/%d %H:%M:%S"))

    def default_record(self):
        d = {k: "" for k in FIELDS} | {"checked": False, "created_at": ""}
        d["allow_print_url"] = True
        d["print_count"] = 0
        d["last_printed_at"] = ""
        d["age_max"] = "55"
        d["pv"] = "1"
        return d

    def _current_business_noun(self) -> str:
        return "提单" if self.current_business == "bill" else "回执"

    def _current_business_table_name(self) -> str:
        return "提单表" if self.current_business == "bill" else "回执管理"

    def _current_data_file(self) -> Path:
        return DATA_FILE if self.current_business == "bill" else RECEIPT_DATA_FILE

    def _switch_business(self, business: str):
        if business not in ("bill", "receipt"):
            return
        if business == self.current_business:
            return
        # 切页前保存当前业务数据，保证提单与回执数据互不覆盖。
        self.save_data()
        self.current_business = business
        self.filtered_indices = []
        self.select_all_state = False
        self.field_errors.clear()
        self.field_error_msgs.clear()
        self.header_filters.clear()
        self.header_sort_field = None
        self.header_sort_order = Qt.SortOrder.AscendingOrder
        self.load_data()
        self.refresh_table()

    def load_data(self):
        data_file = self._current_data_file()
        if data_file.exists():
            try:
                self.records = json.loads(data_file.read_text(encoding="utf-8"))
            except Exception:
                self.records = []
        for rec in self.records:
            rec.setdefault("created_at", "")
            rec["allow_print_url"] = coerce_allow_print_url(rec.get("allow_print_url"))
            rec.setdefault("print_count", 0)
            try:
                rec["print_count"] = int(rec.get("print_count", 0) or 0)
            except (TypeError, ValueError):
                rec["print_count"] = 0
            rec.setdefault("last_printed_at", "")
            rec["last_printed_at"] = str(rec.get("last_printed_at", "") or "")
        if not self.records:
            self.records = [self.default_record()]

    def load_history_data(self):
        if HISTORY_FILE.exists():
            try:
                self.history_records = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
            except Exception:
                self.history_records = []
        for rec in self.history_records:
            rec.setdefault("checked", False)
            rec.setdefault("deleted_at", "")
            rec.setdefault("created_at", "")
            rec["allow_print_url"] = coerce_allow_print_url(rec.get("allow_print_url"))
            rec.setdefault("print_count", 0)
            try:
                rec["print_count"] = int(rec.get("print_count", 0) or 0)
            except (TypeError, ValueError):
                rec["print_count"] = 0
            rec.setdefault("last_printed_at", "")
            rec["last_printed_at"] = str(rec.get("last_printed_at", "") or "")

    def load_picker_recent(self):
        if PICKER_RECENT_FILE.exists():
            try:
                data = json.loads(PICKER_RECENT_FILE.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    self.picker_recent = {str(k): list(v) for k, v in data.items()}
            except Exception:
                self.picker_recent = {}

    def save_picker_recent(self):
        PICKER_RECENT_FILE.write_text(json.dumps(self.picker_recent, ensure_ascii=False, indent=2), encoding="utf-8")

    def update_picker_recent(self, field: str, values: list[str]):
        old = self.picker_recent.get(field, [])
        merged = values + [x for x in old if x not in values]
        self.picker_recent[field] = merged[:20]
        self.save_picker_recent()

    def save_data(self):
        stamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        for rec in self.records:
            if self.is_row_saved(rec) and not str(rec.get("created_at", "")).strip():
                rec["created_at"] = stamp
        self._current_data_file().write_text(json.dumps(self.records, ensure_ascii=False, indent=2), encoding="utf-8")
        self.lbl_save.setText("已自动保存")

    def save_history_data(self):
        HISTORY_FILE.write_text(json.dumps(self.history_records, ensure_ascii=False, indent=2), encoding="utf-8")
        self.lbl_save.setText("已自动保存")

    def load_print_records(self):
        if PRINT_RECORDS_FILE.exists():
            try:
                raw = json.loads(PRINT_RECORDS_FILE.read_text(encoding="utf-8"))
                rows = [x for x in raw if isinstance(x, dict)] if isinstance(raw, list) else []
            except Exception:
                rows = []
        else:
            rows = []
        for rec in rows:
            rec.setdefault("id", str(uuid.uuid4()))
            rec.setdefault("path", "")
            p = str(rec.get("path", "") or "").strip()
            if p:
                rec["path"] = p
            fn = str(rec.get("filename", "") or "").strip()
            if not fn and p:
                rec["filename"] = os.path.basename(p)
            elif not fn:
                rec["filename"] = "导出.xlsx"
            try:
                rec["row_count"] = int(rec.get("row_count", 0) or 0)
            except (TypeError, ValueError):
                rec["row_count"] = 0
            if "include_print_url" not in rec:
                rec["include_print_url"] = False
            else:
                rec["include_print_url"] = bool(rec.get("include_print_url"))
            src = str(rec.get("source", "") or "").strip()
            rec["source"] = src if src in ("打印", "合并") else "打印"
        self.print_records = [r for r in rows if str(r.get("path", "") or "").strip()]

    def save_print_records(self):
        clean = [{k: v for k, v in r.items() if k != "checked"} for r in self.print_records]
        PRINT_RECORDS_FILE.write_text(json.dumps(clean, ensure_ascii=False, indent=2), encoding="utf-8")
        self.lbl_save.setText("已自动保存")

    def load_accessories(self):
        def default_root() -> dict[str, Any]:
            return {"id": "root", "name": "配件表", "node_type": "root", "children": []}

        def normalize(node: dict[str, Any]) -> dict[str, Any]:
            ntype = str(node.get("node_type", "branch") or "branch")
            out = {
                "id": str(node.get("id") or uuid.uuid4()),
                "name": str(node.get("name", "") or "").strip(),
                "node_type": ntype,
                "desc": str(node.get("desc", "") or "").strip(),
                "url": str(node.get("url", "") or "").strip(),
                "remark": str(node.get("remark", "") or "").strip(),
                "created_at": str(node.get("created_at", "") or "").strip() or datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
                "children": [],
            }
            if ntype in ("root", "branch"):
                out["desc"] = ""
                out["url"] = ""
                out["remark"] = ""
                for ch in node.get("children", []) or []:
                    if isinstance(ch, dict):
                        out["children"].append(normalize(ch))
            else:
                out["children"] = []
            return out

        if ACCESSORIES_FILE.exists():
            try:
                raw = json.loads(ACCESSORIES_FILE.read_text(encoding="utf-8"))
                if isinstance(raw, dict) and raw.get("node_type") == "root":
                    self.accessories_root = normalize(raw)
                    self.accessories_root["id"] = "root"
                    self.accessories_root["name"] = "配件表"
                    return
                # 兼容旧版平铺数据，迁移为 root->branch(type)->branch(channel)->branch(name)->leaf(desc,url)
                rows = [x for x in raw if isinstance(x, dict)] if isinstance(raw, list) else []
            except Exception:
                rows = []
        else:
            rows = []
        root = default_root()
        type_map: dict[str, dict[str, dict[str, list[dict[str, Any]]]]] = {}
        for rec in rows:
            tp = str(rec.get("type", "") or "").strip() or "未分类"
            ch = str(rec.get("channel", "") or "").strip() or "未分渠道"
            nm = str(rec.get("name", "") or "").strip() or "未命名"
            type_map.setdefault(tp, {}).setdefault(ch, {}).setdefault(nm, []).append(
                {
                    "id": str(rec.get("id") or uuid.uuid4()),
                    "desc": str(rec.get("desc", "") or "").strip() or "（空描述）",
                    "url": str(rec.get("url", "") or "").strip(),
                    "created_at": str(rec.get("created_at", "") or "").strip() or datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
                }
            )
        for tp, chs in type_map.items():
            tp_node = {"id": str(uuid.uuid4()), "name": tp, "node_type": "branch", "children": []}
            root["children"].append(tp_node)
            for ch, nms in chs.items():
                ch_node = {"id": str(uuid.uuid4()), "name": ch, "node_type": "branch", "children": []}
                tp_node["children"].append(ch_node)
                for nm, leaves in nms.items():
                    nm_node = {"id": str(uuid.uuid4()), "name": nm, "node_type": "branch", "children": []}
                    ch_node["children"].append(nm_node)
                    for lf in leaves:
                        nm_node["children"].append(
                            {
                                "id": lf["id"],
                                "name": lf["desc"],
                                "node_type": "leaf",
                                "desc": lf["desc"],
                                "url": lf["url"],
                                "remark": str(lf.get("remark", "") or ""),
                                "created_at": lf["created_at"],
                                "children": [],
                            }
                        )
        self.accessories_root = normalize(root)
        self.accessories_root["id"] = "root"
        self.accessories_root["name"] = "配件表"

    def save_accessories(self):
        ACCESSORIES_FILE.write_text(json.dumps(self.accessories_root, ensure_ascii=False, indent=2), encoding="utf-8")
        self.lbl_save.setText("已自动保存")

    def _iter_accessory_nodes(self, node: dict[str, Any] | None = None, parent: dict[str, Any] | None = None):
        cur = self.accessories_root if node is None else node
        yield cur, parent
        for ch in cur.get("children", []) or []:
            if isinstance(ch, dict):
                yield from self._iter_accessory_nodes(ch, cur)

    def _find_accessory_node(self, node_id: str) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
        for node, parent in self._iter_accessory_nodes():
            if str(node.get("id", "")) == str(node_id):
                return node, parent
        return None, None

    def _find_node_path_names(self, target_id: str) -> list[str]:
        """返回从根到目标节点（不含根）的名称路径。"""

        def dfs(node: dict[str, Any], path: list[str]) -> list[str] | None:
            nid = str(node.get("id", ""))
            ntype = str(node.get("node_type", ""))
            next_path = path
            if ntype != "root":
                next_path = path + [str(node.get("name", "") or "").strip()]
            if nid == target_id:
                return next_path
            for ch in node.get("children", []) or []:
                if isinstance(ch, dict):
                    got = dfs(ch, next_path)
                    if got is not None:
                        return got
            return None

        if not isinstance(self.accessories_root, dict):
            return []
        return dfs(self.accessories_root, []) or []

    def _leaf_unique_key(self, parent: dict[str, Any] | None, desc: str) -> tuple[str, str, str, str]:
        """唯一键：类目+渠道+名称+描述。"""
        if parent is None:
            path = []
        else:
            path = self._find_node_path_names(str(parent.get("id", "")))
        cat = path[0] if len(path) > 0 else ""
        channel = path[1] if len(path) > 1 else ""
        name = path[2] if len(path) > 2 else (path[-1] if path else "")
        return (cat.strip(), channel.strip(), name.strip(), str(desc or "").strip())

    def _leaf_key_exists(self, key: tuple[str, str, str, str], *, exclude_id: str | None = None) -> bool:
        for node, parent in self._iter_accessory_nodes():
            if node.get("node_type") != "leaf":
                continue
            nid = str(node.get("id", ""))
            if exclude_id and nid == exclude_id:
                continue
            other = self._leaf_unique_key(parent, str(node.get("desc", "") or node.get("name", "")))
            if other == key:
                return True
        return False

    def _accessory_leaf_nodes(self) -> list[dict[str, Any]]:
        return [n for n, _ in self._iter_accessory_nodes() if n.get("node_type") == "leaf"]

    def _selected_url_descs_for_record(self, rec: dict[str, Any]) -> list[str]:
        text = str(rec.get("url", "") or "")
        parts = [x.strip() for x in text.replace("\r\n", "\n").replace("\r", "\n").replace("|", "\n").split("\n") if x.strip()]
        return parts

    def _accessory_desc_url_map(self) -> dict[str, str]:
        out: dict[str, str] = {}
        for lf in self._accessory_leaf_nodes():
            desc = str(lf.get("desc", "") or lf.get("name", "") or "").strip()
            url = str(lf.get("url", "") or "").strip()
            if desc and url and desc not in out:
                out[desc] = url
        return out

    def _resolve_urls_from_descs(self, raw_descs: str) -> list[str]:
        desc_map = self._accessory_desc_url_map()
        descs = [x.strip() for x in str(raw_descs or "").replace("\r\n", "\n").replace("\r", "\n").replace("|", "\n").split("\n") if x.strip()]
        out: list[str] = []
        for d in descs:
            u = desc_map.get(d, "")
            if u:
                out.append(u)
        return out

    def _resolve_urls_from_descs_with_missing(self, raw_descs: str) -> tuple[list[str], list[str]]:
        desc_map = self._accessory_desc_url_map()
        descs = [x.strip() for x in str(raw_descs or "").replace("\r\n", "\n").replace("\r", "\n").replace("|", "\n").split("\n") if x.strip()]
        out: list[str] = []
        missing: list[str] = []
        for d in descs:
            u = desc_map.get(d, "")
            if u:
                out.append(u)
            else:
                missing.append(d)
        return out, missing

    def refresh_accessory_tree(self):
        if not hasattr(self, "accessory_graph"):
            return
        kw = self.accessory_search.text().strip().lower()
        self.accessory_graph.set_tree_data(self.accessories_root, kw)
        if hasattr(self, "btn_acc_add"):
            btn_center_global = self.btn_acc_add.mapToGlobal(self.btn_acc_add.rect().center())
            scene_pt = self.accessory_graph.mapToScene(self.accessory_graph.mapFromGlobal(btn_center_global))
            self.accessory_graph.set_root_anchor_scene_x(scene_pt.x())
        self._refresh_accessory_list_table(kw)
        self._refresh_accessory_list_table(kw)

    def _on_accessory_graph_selected(self, node_id: str):
        self._accessory_selected_node_id = node_id

    def on_accessory_mode_changed(self, *_args):
        mode = str(self.accessory_mode_combo.currentData() or "tree")
        if mode == "list":
            self.accessory_view_stack.setCurrentWidget(self.accessory_list_split)
            self.btn_acc_add.hide()
            self.btn_acc_add_row.show()
            self.btn_acc_delete_selected.show()
            self.btn_acc_delete.hide()
        else:
            self.accessory_view_stack.setCurrentWidget(self.accessory_graph)
            self.btn_acc_add.show()
            self.btn_acc_add_row.hide()
            self.btn_acc_delete.show()
            self.btn_acc_delete_selected.hide()
        self.refresh_accessory_tree()

    def clear_accessory_search_and_filters(self):
        self.accessory_search.setText("")
        self.accessory_header_filters.clear()
        self.accessory_sort_field = None
        self.accessory_sort_order = Qt.SortOrder.AscendingOrder
        self.refresh_accessory_tree()

    def _refresh_accessory_list_table(self, kw: str):
        if not hasattr(self, "accessory_list_table"):
            return
        selected_id = str(getattr(self, "_accessory_selected_node_id", "root") or "root")
        rows: list[dict[str, str]] = []
        for d in list(self._accessory_draft_rows):
            rows.append(
                {
                    "id": str(d.get("id", "") or ""),
                    "type": str(d.get("type", "") or ""),
                    "channel": str(d.get("channel", "") or ""),
                    "name": str(d.get("name", "") or ""),
                    "desc": str(d.get("desc", "") or ""),
                    "url": str(d.get("url", "") or ""),
                    "remark": str(d.get("remark", "") or ""),
                    "created": str(d.get("created", "") or ""),
                }
            )
        for node, parent in self._iter_accessory_nodes():
            ntype = str(node.get("node_type", ""))
            if ntype != "leaf":
                continue
            nid = str(node.get("id", ""))
            path = self._find_node_path_names(nid)
            acc_type = path[0] if len(path) > 0 else ""
            channel = path[1] if len(path) > 1 else ""
            name = path[2] if len(path) > 2 else ""
            desc = str(node.get("desc", "") or "")
            url = str(node.get("url", "") or "")
            remark = str(node.get("remark", "") or "")
            created = str(node.get("created_at", "") or "")
            text_blob = f"{acc_type} {channel} {name} {desc} {url} {remark}".lower()
            if kw and kw not in text_blob:
                continue
            if not self._accessory_row_matches_header_filters_except(node, parent, None):
                continue
            rows.append(
                {
                    "id": nid,
                    "type": acc_type,
                    "channel": channel,
                    "name": name,
                    "desc": desc,
                    "url": url,
                    "remark": remark,
                    "created": created,
                }
            )
        draft_count = len(self._accessory_draft_rows)
        draft_rows = rows[:draft_count]
        data_rows = rows[draft_count:]
        if self.accessory_sort_field:
            rev = self.accessory_sort_order == Qt.SortOrder.DescendingOrder
            sf = self.accessory_sort_field
            key_map = {
                "acc_type": "type",
                "acc_channel": "channel",
                "acc_name": "name",
                "acc_desc": "desc",
                "acc_url": "url",
                "acc_remark": "remark",
                "acc_created": "created",
            }
            data_rows.sort(key=lambda r, f=sf: str(r.get(key_map.get(f, "name"), "") or "").lower(), reverse=rev)
        else:
            data_rows.sort(key=lambda r: (str(r.get("type", "")).lower(), str(r.get("channel", "")).lower(), str(r.get("name", "")).lower(), str(r.get("desc", "")).lower()))
        rows = draft_rows + data_rows
        self.accessory_list_row_node_ids = [r["id"] for r in rows]
        self.accessory_list_frozen.blockSignals(True)
        self.accessory_list_table.blockSignals(True)
        self._acc_list_updating = True
        try:
            self.accessory_list_frozen.setRowCount(len(rows))
            self.accessory_list_table.setRowCount(len(rows))
            for row, rec in enumerate(rows):
                ph0f = QTableWidgetItem("")
                ph0f.setFlags(
                    Qt.ItemFlag.ItemIsEnabled
                    | Qt.ItemFlag.ItemIsSelectable
                    | Qt.ItemFlag.ItemIsUserCheckable
                )
                if rec["id"] in self._accessory_checked_node_ids:
                    ph0f.setCheckState(Qt.CheckState.Checked)
                else:
                    ph0f.setCheckState(Qt.CheckState.Unchecked)
                ph0f.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.accessory_list_frozen.setItem(row, 0, ph0f)
                n_item = QTableWidgetItem(rec["type"])
                n_item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                n_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEditable)
                self.accessory_list_frozen.setItem(row, 1, n_item)
                vals = [rec["channel"], rec["name"], rec["desc"], first_url(rec["url"]), rec["remark"], rec["created"]]
                for col, val in enumerate(vals):
                    it = QTableWidgetItem(val)
                    table_col = col
                    if col in (0, 1, 2, 3, 4):
                        it.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                    else:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    flags = Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
                    if col <= 4 and col != 3:
                        flags |= Qt.ItemFlag.ItemIsEditable
                    it.setFlags(flags)
                    self.accessory_list_table.setItem(row, table_col, it)
                op_wrap = QWidget()
                op_lo = QHBoxLayout(op_wrap)
                op_lo.setContentsMargins(4, 2, 4, 2)
                op_lo.setSpacing(0)
                op_lo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                bdel = QPushButton("🗑 删除")
                bdel.setObjectName("btnDanger")
                bdel.setFixedSize(92, 30)
                nid = rec["id"]
                if nid.startswith("draft-"):
                    bdel.clicked.connect(lambda _=False, x=nid: self._remove_accessory_draft_row(x))
                else:
                    bdel.clicked.connect(lambda _=False, x=nid: self.delete_accessory_leaf_by_id(x))
                op_lo.addWidget(bdel)
                self.accessory_list_table.setCellWidget(row, 6, op_wrap)
                self.accessory_list_frozen.setRowHeight(row, self._data_row_height)
                self.accessory_list_table.setRowHeight(row, self._data_row_height)
        finally:
            self._acc_list_updating = False
            self.accessory_list_frozen.blockSignals(False)
            self.accessory_list_table.blockSignals(False)
        if not rows:
            self._update_accessory_header_check()
            self._refresh_accessory_header_labels()
            self._update_accessory_header_sort_indicator()
            self._update_accessory_header_tooltips()
            return
        target_row = 0
        for i, nid in enumerate(self.accessory_list_row_node_ids):
            if nid == selected_id:
                target_row = i
                break
        self.accessory_list_frozen.selectRow(target_row)
        self._refresh_accessory_header_labels()
        self._update_accessory_header_check()
        self._update_accessory_header_sort_indicator()
        self._update_accessory_header_tooltips()

    def _on_accessory_list_selection_changed(self):
        if not hasattr(self, "accessory_list_table"):
            return
        row = self.accessory_list_frozen.currentRow()
        if row < 0:
            row = self.accessory_list_table.currentRow()
        if row < 0 or row >= len(self.accessory_list_row_node_ids):
            return
        self._accessory_selected_node_id = self.accessory_list_row_node_ids[row]

    def _on_accessory_list_frozen_cell_clicked(self, row: int, col: int):
        if self._acc_list_updating:
            return
        if col != 0:
            return
        it = self.accessory_list_frozen.item(row, 0)
        if it is None:
            return
        nxt = Qt.CheckState.Unchecked if it.checkState() == Qt.CheckState.Checked else Qt.CheckState.Checked
        it.setCheckState(nxt)

    def _on_accessory_list_cell_double_clicked(self, row: int, col: int):
        # URL 列使用弹窗多行编辑，列表仅展示第一条。
        if col != 3:
            return
        if row < 0 or row >= len(self.accessory_list_row_node_ids):
            return
        node_id = self.accessory_list_row_node_ids[row]
        if str(node_id).startswith("draft-"):
            cur = str((self.accessory_list_table.item(row, 3).text() if self.accessory_list_table.item(row, 3) else "")).strip()
        else:
            node, _parent = self._find_accessory_node(node_id)
            if node is None or node.get("node_type") != "leaf":
                return
            cur = str(node.get("url", "") or "")
        dlg = QDialog(self)
        dlg.setWindowTitle("编辑URL（多行）")
        dlg.resize(620, 420)
        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(12, 12, 12, 12)
        lo.setSpacing(8)
        lo.addWidget(QLabel("每行一个URL："))
        ed = QPlainTextEdit(cur)
        ed.setPlaceholderText("可输入多行URL")
        lo.addWidget(ed, 1)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lo.addWidget(btns)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        idx = self.accessory_list_table.model().index(row, 3)
        rect = self.accessory_list_table.visualRect(idx)
        anchor = self.accessory_list_table.viewport().mapToGlobal(rect.bottomLeft())
        self._place_dialog_below_rect(dlg, anchor, rect.width())
        if dlg.exec() != QDialog.Accepted:
            return
        new_url = str(ed.toPlainText() or "").strip()
        if str(node_id).startswith("draft-"):
            self._acc_list_updating = True
            try:
                show_first = first_url(new_url)
                cur_item = self.accessory_list_table.item(row, 3)
                if cur_item is None:
                    cur_item = QTableWidgetItem(show_first)
                    self.accessory_list_table.setItem(row, 3, cur_item)
                else:
                    cur_item.setText(show_first)
            finally:
                self._acc_list_updating = False
            self._commit_or_update_accessory_draft_row(row, node_id, url_override=new_url)
            return
        node["url"] = new_url
        self.save_accessories()
        self.refresh_accessory_tree()

    def _on_accessory_list_item_changed(self, _item: QTableWidgetItem):
        if self._acc_list_updating:
            return
        if _item.tableWidget() is self.accessory_list_frozen and _item.column() == 0:
            self._on_accessory_list_checkbox_changed(
                _item.row(), _item.checkState() == Qt.CheckState.Checked
            )
            return
        row = self.accessory_list_frozen.currentRow()
        if row < 0:
            row = self.accessory_list_table.currentRow()
        if row < 0 or row >= len(self.accessory_list_row_node_ids):
            return
        node_id = self.accessory_list_row_node_ids[row]
        if str(node_id).startswith("draft-"):
            self._commit_or_update_accessory_draft_row(row, node_id)
            return
        node, old_parent = self._find_accessory_node(node_id)
        if node is None or old_parent is None or node.get("node_type") != "leaf":
            return
        acc_type = str((self.accessory_list_frozen.item(row, 1).text() if self.accessory_list_frozen.item(row, 1) else "")).strip()
        channel = str((self.accessory_list_table.item(row, 0).text() if self.accessory_list_table.item(row, 0) else "")).strip()
        name = str((self.accessory_list_table.item(row, 1).text() if self.accessory_list_table.item(row, 1) else "")).strip()
        desc = str((self.accessory_list_table.item(row, 2).text() if self.accessory_list_table.item(row, 2) else "")).strip()
        url = str((self.accessory_list_table.item(row, 3).text() if self.accessory_list_table.item(row, 3) else "")).strip()
        remark = str((self.accessory_list_table.item(row, 4).text() if self.accessory_list_table.item(row, 4) else "")).strip()
        if not all([acc_type, channel, name, desc, url]):
            QMessageBox.warning(self, "提示", "类型、渠道、名称、描述、URL 不能为空")
            self.refresh_accessory_tree()
            return
        target_branch = self._ensure_accessory_branch_path(acc_type, channel, name)
        key = self._leaf_unique_key(target_branch, desc)
        if self._leaf_key_exists(key, exclude_id=node_id):
            QMessageBox.warning(self, "提示", "同一“类目+渠道+名称+描述”不允许重复")
            self.refresh_accessory_tree()
            return
        if str(old_parent.get("id", "")) != str(target_branch.get("id", "")):
            old_parent["children"] = [x for x in (old_parent.get("children", []) or []) if str(x.get("id", "")) != node_id]
            target_branch.setdefault("children", []).append(node)
        node["name"] = desc
        node["desc"] = desc
        node["url"] = url
        node["remark"] = remark
        self.save_accessories()
        self.refresh_accessory_tree()

    def _remove_accessory_draft_row(self, draft_id: str):
        before = len(self._accessory_draft_rows)
        self._accessory_draft_rows = [r for r in self._accessory_draft_rows if str(r.get("id", "")) != draft_id]
        if len(self._accessory_draft_rows) == before:
            return
        self._accessory_checked_node_ids.discard(draft_id)
        self._accessory_selected_node_id = "root"
        self.refresh_accessory_tree()

    def _commit_or_update_accessory_draft_row(self, row: int, draft_id: str, url_override: str | None = None):
        acc_type = str((self.accessory_list_frozen.item(row, 1).text() if self.accessory_list_frozen.item(row, 1) else "")).strip()
        channel = str((self.accessory_list_table.item(row, 0).text() if self.accessory_list_table.item(row, 0) else "")).strip()
        name = str((self.accessory_list_table.item(row, 1).text() if self.accessory_list_table.item(row, 1) else "")).strip()
        desc = str((self.accessory_list_table.item(row, 2).text() if self.accessory_list_table.item(row, 2) else "")).strip()
        if url_override is None:
            url = str((self.accessory_list_table.item(row, 3).text() if self.accessory_list_table.item(row, 3) else "")).strip()
        else:
            url = str(url_override).strip()
        remark = str((self.accessory_list_table.item(row, 4).text() if self.accessory_list_table.item(row, 4) else "")).strip()
        created = str((self.accessory_list_table.item(row, 5).text() if self.accessory_list_table.item(row, 5) else "")).strip()
        for d in self._accessory_draft_rows:
            if str(d.get("id", "")) == draft_id:
                d["type"] = acc_type
                d["channel"] = channel
                d["name"] = name
                d["desc"] = desc
                d["url"] = url
                d["remark"] = remark
                d["created"] = created or str(d.get("created", "") or datetime.now().strftime("%Y/%m/%d %H:%M:%S"))
                break
        if not all([acc_type, channel, name, desc, url]):
            return
        target_branch = self._ensure_accessory_branch_path(acc_type, channel, name)
        key = self._leaf_unique_key(target_branch, desc)
        if self._leaf_key_exists(key):
            QMessageBox.warning(self, "提示", "同一“类目+渠道+名称+描述”不允许重复")
            return
        child = {
            "id": str(uuid.uuid4()),
            "name": desc,
            "node_type": "leaf",
            "desc": desc,
            "url": url,
            "remark": remark,
            "created_at": created or datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
            "children": [],
        }
        target_branch.setdefault("children", []).append(child)
        self._remove_accessory_draft_row(draft_id)
        self._accessory_selected_node_id = str(child["id"])
        self.save_accessories()
        self.refresh_accessory_tree()

    def _accessory_row_display_value(self, node: dict[str, Any], parent: dict[str, Any] | None, field: str) -> str:
        nid = str(node.get("id", ""))
        path = self._find_node_path_names(nid)
        if field == "acc_type":
            return path[0] if len(path) > 0 else ""
        if field == "acc_channel":
            return path[1] if len(path) > 1 else ""
        if field == "acc_name":
            return path[2] if len(path) > 2 else ""
        if field == "acc_desc":
            return str(node.get("desc", "") or "")
        if field == "acc_url":
            return str(node.get("url", "") or "")
        if field == "acc_remark":
            return str(node.get("remark", "") or "")
        if field == "acc_created":
            return str(node.get("created_at", "") or "")
        return ""

    def _accessory_row_matches_header_filters_except(self, node: dict[str, Any], parent: dict[str, Any] | None, skip_field: str | None) -> bool:
        for field, vals in self.accessory_header_filters.items():
            if skip_field and field == skip_field:
                continue
            if not vals:
                continue
            if self._accessory_row_display_value(node, parent, field) not in vals:
                return False
        return True

    def _unique_display_values_accessory(self, field: str) -> list[str]:
        kw = self.accessory_search.text().strip().lower()
        out: list[str] = []
        seen: set[str] = set()
        for node, parent in self._iter_accessory_nodes():
            ntype = str(node.get("node_type", ""))
            if ntype != "leaf":
                continue
            text_blob = " ".join(
                [
                    self._accessory_row_display_value(node, parent, "acc_type"),
                    self._accessory_row_display_value(node, parent, "acc_channel"),
                    self._accessory_row_display_value(node, parent, "acc_name"),
                    self._accessory_row_display_value(node, parent, "acc_desc"),
                    self._accessory_row_display_value(node, parent, "acc_url"),
                    self._accessory_row_display_value(node, parent, "acc_remark"),
                ]
            ).lower()
            if kw and kw not in text_blob:
                continue
            if not self._accessory_row_matches_header_filters_except(node, parent, field):
                continue
            dv = self._accessory_row_display_value(node, parent, field)
            if dv not in seen:
                seen.add(dv)
                out.append(dv)
        out.sort(key=lambda s: (s == "", s.casefold()))
        return out

    def _toggle_accessory_sort_for_field(self, field: str):
        if self.accessory_sort_field == field:
            if self.accessory_sort_order == Qt.SortOrder.AscendingOrder:
                self.accessory_sort_order = Qt.SortOrder.DescendingOrder
            else:
                self.accessory_sort_field = None
                self.accessory_sort_order = Qt.SortOrder.AscendingOrder
        else:
            self.accessory_sort_field = field
            self.accessory_sort_order = Qt.SortOrder.AscendingOrder
        self.refresh_accessory_tree()

    def _on_accessory_frozen_header_clicked(self, section: int):
        if section == 0:
            self._on_frozen_header_col0_clicked("acc_frozen")
            return
        field = self._field_for_header_section("acc_frozen", section)
        if not field:
            return
        self._toggle_accessory_sort_for_field(field)

    def _on_accessory_scroll_header_clicked(self, section: int):
        field = self._field_for_header_section("acc_scroll", section)
        if not field:
            return
        self._toggle_accessory_sort_for_field(field)

    def _sync_accessory_v_scroll(self, value: int, source: str):
        if self._acc_v_sync:
            return
        self._acc_v_sync = True
        try:
            if source == "scroll":
                self.accessory_list_frozen.verticalScrollBar().setValue(value)
            else:
                self.accessory_list_table.verticalScrollBar().setValue(value)
        finally:
            self._acc_v_sync = False

    def _sync_selection_accessory_to_frozen(self):
        if self._acc_sel_sync or self.accessory_list_table.selectionModel() is None:
            return
        self._acc_sel_sync = True
        try:
            self.accessory_list_frozen.clearSelection()
            for idx in self.accessory_list_table.selectionModel().selectedRows():
                self.accessory_list_frozen.selectRow(idx.row())
        finally:
            self._acc_sel_sync = False

    def _sync_selection_accessory_frozen_to_scroll(self):
        if self._acc_sel_sync or self.accessory_list_frozen.selectionModel() is None:
            return
        self._acc_sel_sync = True
        try:
            self.accessory_list_table.clearSelection()
            for idx in self.accessory_list_frozen.selectionModel().selectedRows():
                self.accessory_list_table.selectRow(idx.row())
        finally:
            self._acc_sel_sync = False

    def _update_accessory_header_sort_indicator(self):
        hf = self.accessory_list_frozen.horizontalHeader()
        hs = self.accessory_list_table.horizontalHeader()
        hf.setSortIndicatorShown(False)
        hs.setSortIndicatorShown(False)
        if not self.accessory_sort_field:
            return
        if self.accessory_sort_field == "acc_type":
            hf.setSortIndicatorShown(True)
            hf.setSortIndicator(1, self.accessory_sort_order)
            return
        fmap = ["acc_channel", "acc_name", "acc_desc", "acc_url", "acc_remark", "acc_created", None]
        if self.accessory_sort_field in fmap:
            sec = fmap.index(self.accessory_sort_field)
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(sec, self.accessory_sort_order)

    def _update_accessory_header_tooltips(self):
        frozen_title = self.accessory_list_frozen.horizontalHeaderItem(1)
        if frozen_title:
            tip = "类型\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.accessory_header_filters.get("acc_type", set())
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            frozen_title.setToolTip(tip)
        titles = ["渠道", "名称", "描述", "URL", "备注", "创建时间", "操作"]
        for sec, title in enumerate(titles):
            field = self._field_for_header_section("acc_scroll", sec)
            if not field:
                continue
            tip = title + "\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.accessory_header_filters.get(field, set())
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            it = self.accessory_list_table.horizontalHeaderItem(sec)
            if it:
                it.setToolTip(tip)

    def _refresh_accessory_header_labels(self):
        def filtered_text(base: str, field: str) -> str:
            vals = self.accessory_header_filters.get(field, set())
            return f"{base} (筛:{len(vals)})" if vals else base

        self.accessory_list_frozen.setHorizontalHeaderItem(0, QTableWidgetItem("☐"))
        self.accessory_list_frozen.setHorizontalHeaderItem(1, QTableWidgetItem(filtered_text("类型", "acc_type")))
        titles = [
            ("渠道", "acc_channel"),
            ("名称", "acc_name"),
            ("描述", "acc_desc"),
            ("URL", "acc_url"),
            ("备注", "acc_remark"),
            ("创建时间", "acc_created"),
            ("操作", ""),
        ]
        for sec, (base, field) in enumerate(titles):
            txt = filtered_text(base, field) if field else base
            self.accessory_list_table.setHorizontalHeaderItem(sec, QTableWidgetItem(txt))

    def _update_accessory_header_check(self):
        if not hasattr(self, "accessory_list_frozen"):
            return
        it = self.accessory_list_frozen.horizontalHeaderItem(0)
        if it is None:
            return
        n = len(self.accessory_list_row_node_ids)
        if n == 0:
            it.setText("☐")
            it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            return
        checked_count = sum(1 for nid in self.accessory_list_row_node_ids if nid in self._accessory_checked_node_ids)
        symbol = "☐" if checked_count == 0 else ("☑" if checked_count == n else "◩")
        it.setText(symbol)
        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def _selected_accessory_node(self) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
        rid = str(getattr(self, "_accessory_selected_node_id", "root") or "root")
        return self._find_accessory_node(rid)

    def _open_accessory_node_dialog(self, title: str, preset: dict[str, Any] | None = None, allow_type_change: bool = True):
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.resize(600, 400)
        form = QFormLayout(dlg)
        # Mac 下原生控件视觉更紧凑，统一放大输入控件与间距，减少跨平台观感差异。
        is_mac = sys.platform == "darwin"
        form.setContentsMargins(22, 20, 22, 18)
        form.setHorizontalSpacing(16)
        form.setVerticalSpacing(14 if is_mac else 12)
        form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)
        cmb_type = QComboBox()
        cmb_type.addItem("枝干", "branch")
        cmb_type.addItem("叶子", "leaf")
        if not allow_type_change:
            cmb_type.setEnabled(False)
        name = QLineEdit(str((preset or {}).get("name", "") or ""))
        desc = QLineEdit(str((preset or {}).get("desc", "") or ""))
        url = QLineEdit(str((preset or {}).get("url", "") or ""))
        ctrl_h = 38 if is_mac else 34
        cmb_type.setMinimumHeight(ctrl_h)
        cmb_type.setMinimumWidth(220)
        name.setMinimumHeight(ctrl_h)
        desc.setMinimumHeight(ctrl_h)
        url.setMinimumHeight(ctrl_h)
        btn_h = 40 if is_mac else 36
        row_type = "节点种类"
        row_name = "名称"
        row_desc = "描述"
        row_url = "URL"
        form.addRow(row_type, cmb_type)
        form.addRow(row_name, name)
        form.addRow(row_desc, desc)
        form.addRow(row_url, url)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        for b in btns.buttons():
            b.setMinimumHeight(btn_h)
            b.setMinimumWidth(96)
        form.addRow(btns)

        def update_visible():
            is_leaf = (cmb_type.currentData() == "leaf")
            # 叶子仅显示描述+URL；枝干仅显示名称。
            name.setVisible(not is_leaf)
            form.labelForField(name).setVisible(not is_leaf)
            desc.setVisible(is_leaf)
            form.labelForField(desc).setVisible(is_leaf)
            url.setVisible(is_leaf)
            form.labelForField(url).setVisible(is_leaf)
            if is_leaf:
                name.setText("")
            else:
                desc.setText("")
                url.setText("")

            if allow_type_change:
                cmb_type.setVisible(True)
                form.labelForField(cmb_type).setVisible(True)
            else:
                cmb_type.setVisible(False)
                form.labelForField(cmb_type).setVisible(False)

        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        cmb_type.currentIndexChanged.connect(update_visible)
        if preset and preset.get("node_type") == "leaf":
            cmb_type.setCurrentIndex(cmb_type.findData("leaf"))
        update_visible()
        if dlg.exec() != QDialog.Accepted:
            return None
        out = {
            "node_type": str(cmb_type.currentData() or "branch"),
            "name": name.text().strip(),
            "desc": desc.text().strip(),
            "url": url.text().strip(),
            "created_at": str((preset or {}).get("created_at", "") or datetime.now().strftime("%Y/%m/%d %H:%M:%S")),
        }
        if out["node_type"] == "branch" and not out["name"]:
            QMessageBox.warning(self, "提示", "枝干节点名称不能为空")
            return None
        if out["node_type"] == "leaf" and (not out["desc"] or not out["url"]):
            QMessageBox.warning(self, "提示", "叶子节点必须填写描述和URL")
            return None
        if out["node_type"] == "leaf":
            out["name"] = out["desc"]
        return out

    def add_accessory(self):
        node, _parent = self._selected_accessory_node()
        if node is None:
            return
        if node.get("node_type") == "leaf":
            QMessageBox.information(self, "提示", "叶子节点不允许新增子节点")
            return
        data = self._open_accessory_node_dialog("新增节点")
        if not data:
            return
        if data["node_type"] == "leaf":
            key = self._leaf_unique_key(node, data["desc"])
            if self._leaf_key_exists(key):
                QMessageBox.warning(self, "提示", "同一“类目+渠道+名称+描述”不允许重复")
                return
        child = {
            "id": str(uuid.uuid4()),
            "name": data["name"],
            "node_type": data["node_type"],
            "desc": data["desc"] if data["node_type"] == "leaf" else "",
            "url": data["url"] if data["node_type"] == "leaf" else "",
            "created_at": data["created_at"],
            "children": [],
        }
        node.setdefault("children", []).append(child)
        self.save_accessories()
        self.refresh_accessory_tree()

    def add_accessory_row(self):
        mode = str(self.accessory_mode_combo.currentData() or "tree")
        if mode != "list":
            return
        draft_id = f"draft-{uuid.uuid4()}"
        now_ts = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        self._accessory_draft_rows.insert(
            0,
            {
                "id": draft_id,
                "type": "",
                "channel": "",
                "name": "",
                "desc": "",
                "url": "",
                "remark": "",
                "created": now_ts,
            },
        )
        self._accessory_selected_node_id = draft_id
        self.refresh_accessory_tree()

    def update_accessory(self):
        node, _parent = self._selected_accessory_node()
        if node is None or node.get("node_type") == "root":
            return
        data = self._open_accessory_node_dialog("编辑节点", preset=node, allow_type_change=False)
        if not data:
            return
        if node.get("node_type") == "leaf":
            _cur, parent = self._find_accessory_node(str(node.get("id", "")))
            key = self._leaf_unique_key(parent, data["desc"])
            if self._leaf_key_exists(key, exclude_id=str(node.get("id", ""))):
                QMessageBox.warning(self, "提示", "同一“类目+渠道+名称+描述”不允许重复")
                return
        node["name"] = data["name"]
        if node.get("node_type") == "leaf":
            node["desc"] = data["desc"]
            node["url"] = data["url"]
        self.save_accessories()
        self.refresh_accessory_tree()

    def copy_accessory_leaf(self):
        node, _parent = self._selected_accessory_node()
        if node is None:
            return
        if node.get("node_type") != "leaf":
            QMessageBox.information(self, "提示", "请选择叶子节点后再复制")
            return
        base_desc = str(node.get("desc", "") or node.get("name", "") or "").strip()
        if not base_desc:
            QMessageBox.warning(self, "提示", "当前叶子节点描述为空，无法复制")
            return
        target_branch = self._pick_accessory_branch_for_copy()
        if target_branch is None:
            return
        if self._leaf_key_exists(self._leaf_unique_key(target_branch, base_desc)):
            QMessageBox.warning(self, "提示", "目标枝干下已存在同描述叶子，无法复制")
            return
        child = {
            "id": str(uuid.uuid4()),
            "name": base_desc,
            "node_type": "leaf",
            "desc": base_desc,
            "url": str(node.get("url", "") or ""),
            "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
            "children": [],
        }
        target_branch.setdefault("children", []).append(child)
        self.save_accessories()
        self.refresh_accessory_tree()
        target_name = str(target_branch.get("name", "") or "配件表")
        QMessageBox.information(self, "复制成功", f"已复制叶子节点：{base_desc}\n目标枝干：{target_name}")

    def _pick_accessory_branch_for_copy(self) -> dict[str, Any] | None:
        """弹窗选择复制目标枝干（树形，仅显示根/枝干）。"""
        if not isinstance(self.accessories_root, dict):
            QMessageBox.warning(self, "提示", "未找到可用枝干节点")
            return None

        dlg = QDialog(self)
        dlg.setWindowTitle("选择目标枝干")
        dlg.resize(520, 420)
        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(12, 12, 12, 12)
        lo.setSpacing(8)
        lo.addWidget(QLabel("请选择要复制到的枝干："))
        tree = QTreeWidget()
        tree.setHeaderLabels(["枝干"])
        tree.header().setStretchLastSection(True)
        tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        lo.addWidget(tree, 1)

        def add_branch_item(node: dict[str, Any], parent_item: QTreeWidgetItem | None):
            ntype = str(node.get("node_type", "") or "")
            if ntype not in ("root", "branch"):
                return
            txt = str(node.get("name", "") or "").strip() or "配件表"
            item = QTreeWidgetItem([txt])
            item.setData(0, Qt.ItemDataRole.UserRole, str(node.get("id", "")))
            if parent_item is None:
                tree.addTopLevelItem(item)
            else:
                parent_item.addChild(item)
            for ch in node.get("children", []) or []:
                if isinstance(ch, dict):
                    add_branch_item(ch, item)

        add_branch_item(self.accessories_root, None)
        tree.expandAll()
        if tree.topLevelItemCount() > 0:
            tree.setCurrentItem(tree.topLevelItem(0))
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lo.addWidget(btns)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        tree.itemDoubleClicked.connect(lambda _it, _col: dlg.accept())
        if dlg.exec() != QDialog.Accepted:
            return None
        it = tree.currentItem()
        if it is None:
            QMessageBox.information(self, "提示", "请先选择目标枝干")
            return None
        nid = str(it.data(0, Qt.ItemDataRole.UserRole) or "")
        node, _p = self._find_accessory_node(nid)
        if node is None or node.get("node_type") not in ("root", "branch"):
            QMessageBox.warning(self, "提示", "目标枝干无效，请重试")
            return None
        return node

    def _used_urls_in_bills(self) -> set[str]:
        used: set[str] = set()
        for rec in self.records:
            raw = str(rec.get("url", "") or "")
            for line in raw.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
                s = line.strip()
                if s:
                    used.add(s)
        return used

    def _ensure_accessory_branch_path(self, acc_type: str, channel: str, name: str) -> dict[str, Any]:
        def ensure_child(parent: dict[str, Any], child_name: str) -> dict[str, Any]:
            for ch in parent.get("children", []) or []:
                if isinstance(ch, dict) and ch.get("node_type") == "branch" and str(ch.get("name", "")) == child_name:
                    return ch
            new_node = {
                "id": str(uuid.uuid4()),
                "name": child_name,
                "node_type": "branch",
                "desc": "",
                "url": "",
                "remark": "",
                "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
                "children": [],
            }
            parent.setdefault("children", []).append(new_node)
            return new_node

        type_node = ensure_child(self.accessories_root, acc_type)
        channel_node = ensure_child(type_node, channel)
        name_node = ensure_child(channel_node, name)
        return name_node

    def _leaf_desc_used_in_bills(self, desc: str) -> bool:
        used = self._used_urls_in_bills()
        return str(desc or "").strip() in used

    def _on_accessory_list_checkbox_changed(self, row: int, checked: bool):
        if self._acc_list_updating:
            return
        if row < 0:
            return
        if row >= len(self.accessory_list_row_node_ids):
            return
        node_id = str(self.accessory_list_row_node_ids[row] or "")
        if not node_id:
            return
        if checked:
            self._accessory_checked_node_ids.add(node_id)
        else:
            self._accessory_checked_node_ids.discard(node_id)
        self._update_accessory_header_check()
        self.accessory_list_table.selectRow(row)
        self.accessory_list_frozen.selectRow(row)

    def _checked_accessory_leaf_node_ids(self) -> list[str]:
        return [nid for nid in self.accessory_list_row_node_ids if nid in self._accessory_checked_node_ids]

    def _delete_checked_accessory_leaves(self) -> bool:
        node_ids = self._checked_accessory_leaf_node_ids()
        if not node_ids:
            return False
        used = self._used_urls_in_bills()
        blocked: list[str] = []
        targets: list[tuple[str, dict[str, Any], dict[str, Any]]] = []
        for node_id in node_ids:
            node, parent = self._find_accessory_node(node_id)
            if node is None or parent is None or node.get("node_type") != "leaf":
                continue
            desc = str(node.get("desc", "") or node.get("name", "")).strip()
            url = str(node.get("url", "") or "").strip()
            if desc in used or url in used:
                blocked.append(desc or url or "（空描述）")
                continue
            targets.append((node_id, node, parent))
        if blocked:
            show = "、".join(blocked[:5])
            more = f" 等{len(blocked)}项" if len(blocked) > 5 else ""
            QMessageBox.warning(
                self,
                "提示",
                f"以下配件已被{self._current_business_table_name()}引用，不允许删除：\n{show}{more}",
            )
            return True
        if not targets:
            QMessageBox.information(self, "提示", "未找到可删除的勾选叶子")
            return True
        if QMessageBox.question(self, "确认删除", f"确认删除已勾选的 {len(targets)} 个叶子节点吗？此操作不可撤销。") != QMessageBox.Yes:
            return True
        by_parent: dict[str, tuple[dict[str, Any], set[str]]] = {}
        for node_id, _node, parent in targets:
            pid = str(parent.get("id", ""))
            if pid not in by_parent:
                by_parent[pid] = (parent, set())
            by_parent[pid][1].add(node_id)
        for parent, del_ids in by_parent.values():
            parent["children"] = [
                x for x in (parent.get("children", []) or []) if str(x.get("id", "")) not in del_ids
            ]
        self.save_accessories()
        self.refresh_accessory_tree()
        return True

    def _node_or_descendants_used(self, node: dict[str, Any], used_urls: set[str]) -> bool:
        if node.get("node_type") == "leaf":
            desc = str(node.get("desc", "") or node.get("name", "")).strip()
            url = str(node.get("url", "") or "").strip()
            if desc in used_urls or url in used_urls:
                return True
        for ch in node.get("children", []) or []:
            if isinstance(ch, dict) and self._node_or_descendants_used(ch, used_urls):
                return True
        return False

    def delete_accessory(self, _rid: str | None = None):
        node, parent = self._selected_accessory_node()
        if node is None or parent is None:
            return
        if node.get("node_type") == "root":
            QMessageBox.information(self, "提示", "根节点不可删除")
            return
        used_urls = self._used_urls_in_bills()
        if self._node_or_descendants_used(node, used_urls):
            QMessageBox.warning(self, "提示", f"该节点（或其子节点）已被{self._current_business_table_name()}使用，不可删除")
            return
        node_name = str(node.get("desc", "") or node.get("name", "") or "该节点")
        if QMessageBox.question(self, "确认删除", f"确认删除“{node_name}”节点吗？此操作不可撤销。") != QMessageBox.Yes:
            return
        parent["children"] = [x for x in (parent.get("children", []) or []) if x.get("id") != node.get("id")]
        self.save_accessories()
        self.refresh_accessory_tree()

    def delete_accessory_selected(self):
        if self._delete_checked_accessory_leaves():
            return
        QMessageBox.information(self, "提示", "请先勾选要删除的配件叶子")

    def delete_accessory_leaf_by_id(self, node_id: str):
        node, parent = self._find_accessory_node(node_id)
        if node is None or parent is None or node.get("node_type") != "leaf":
            return
        desc = str(node.get("desc", "") or node.get("name", "")).strip()
        url = str(node.get("url", "") or "").strip()
        used = self._used_urls_in_bills()
        if desc in used or url in used:
            QMessageBox.warning(self, "提示", f"该配件已被{self._current_business_table_name()}引用，不允许删除")
            return
        if QMessageBox.question(self, "确认删除", f"确认删除叶子“{desc}”吗？此操作不可撤销。") != QMessageBox.Yes:
            return
        parent["children"] = [x for x in (parent.get("children", []) or []) if str(x.get("id", "")) != node_id]
        self.save_accessories()
        self.refresh_accessory_tree()

    def import_accessories_from_excel(self):
        default_tpl = APP_DIR / "input-template.xlsx"
        start_dir = str(default_tpl.parent if default_tpl.exists() else APP_DIR)
        in_file, _ = QFileDialog.getOpenFileName(self, "选择配件导入文件", start_dir, "Excel (*.xlsx *.xlsm)")
        if not in_file:
            return
        try:
            wb = load_workbook(str(in_file), read_only=True, data_only=True)
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"读取文件失败：\n{e}")
            return
        added = 0
        updated = 0
        skipped = 0
        progress: QProgressDialog | None = None
        try:
            ws = wb[wb.sheetnames[0]]
            max_row = ws.max_row or 1
            if max_row < 2:
                QMessageBox.information(self, "导入提示", "文件为空")
                return
            now_ts = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            total_rows = max_row - 1
            progress = QProgressDialog("正在导入配件数据...", "取消", 0, max(1, total_rows), self)
            progress.setWindowTitle("导入中")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setAutoClose(True)
            progress.setAutoReset(True)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            for idx, row in enumerate(
                ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=6, values_only=True),
                start=1,
            ):
                if progress.wasCanceled():
                    break
                vals = [str(x or "").strip() for x in row]
                if len(vals) < 6:
                    vals.extend([""] * (6 - len(vals)))
                acc_type, channel, name, desc, url, remark = vals[:6]
                if not any([acc_type, channel, name, desc, url, remark]):
                    progress.setValue(idx)
                    if idx % 50 == 0 or idx == total_rows:
                        QApplication.processEvents()
                    continue
                if not all([acc_type, channel, name, desc, url]):
                    skipped += 1
                    progress.setValue(idx)
                    if idx % 50 == 0 or idx == total_rows:
                        QApplication.processEvents()
                    continue
                parent = self._ensure_accessory_branch_path(acc_type, channel, name)
                key = self._leaf_unique_key(parent, desc)
                existed = None
                for ch in parent.get("children", []) or []:
                    if isinstance(ch, dict) and ch.get("node_type") == "leaf":
                        other = self._leaf_unique_key(parent, str(ch.get("desc", "") or ch.get("name", "")))
                        if other == key:
                            existed = ch
                            break
                if existed is not None:
                    existed["name"] = desc
                    existed["desc"] = desc
                    existed["url"] = url
                    existed["remark"] = remark
                    existed["created_at"] = now_ts
                    updated += 1
                else:
                    parent.setdefault("children", []).append(
                        {
                            "id": str(uuid.uuid4()),
                            "name": desc,
                            "node_type": "leaf",
                            "desc": desc,
                            "url": url,
                            "remark": remark,
                            "created_at": now_ts,
                            "children": [],
                        }
                    )
                    added += 1
                progress.setValue(idx)
                if idx % 10 == 0 or idx == total_rows:
                    QApplication.processEvents()
        finally:
            if progress is not None:
                progress.setValue(progress.maximum())
                QApplication.processEvents()
                progress.close()
                progress.deleteLater()
                QApplication.processEvents()
            wb.close()
        self.save_accessories()
        self.refresh_accessory_tree()
        if progress is not None and progress.wasCanceled():
            self.statusBar().showMessage(
                f"导入已取消：新增 {added} 条，更新 {updated} 条，跳过 {skipped} 条",
                5000,
            )
            return
        self.statusBar().showMessage(
            f"导入完成：新增 {added} 条，更新 {updated} 条，跳过 {skipped} 条",
            5000,
        )

    @staticmethod
    def _print_record_file_ok(rec: dict) -> bool:
        p = str(rec.get("path", "") or "").strip()
        if not p:
            return False
        try:
            return Path(p).is_file()
        except OSError:
            return False

    def _print_cell_display_for_filter(self, rec: dict, field: str) -> str:
        if field == "row_count":
            return str(int(rec.get("row_count", 0) or 0))
        if field == "include_print_url":
            if str(rec.get("source", "") or "") == "合并":
                return "—"
            return "是" if rec.get("include_print_url") else "否"
        if field == "file_exists":
            return "有" if self._print_record_file_ok(rec) else "无"
        if field == "filename":
            return str(rec.get("filename", "") or "")
        if field == "source":
            return str(rec.get("source", "") or "")
        if field == "printed_at":
            return BillApp._created_at_filter_key(rec.get("printed_at"))
        if field == "path":
            return str(rec.get("path", "") or "")
        return ""

    def _print_row_matches_header_filters_except(self, rec: dict, skip_field: str | None) -> bool:
        for field, vals in self.print_header_filters.items():
            if skip_field and field == skip_field:
                continue
            if not vals:
                continue
            if field == "printed_at":
                dk = BillApp._created_at_filter_key(rec.get("printed_at"))
                allowed = {BillApp._created_at_filter_key(x) for x in vals}
                if dk not in allowed:
                    return False
            elif self._print_cell_display_for_filter(rec, field) not in vals:
                return False
        return True

    def _print_row_matches_header_filters(self, rec: dict) -> bool:
        return self._print_row_matches_header_filters_except(rec, None)

    def _unique_display_values_print(self, field: str) -> list[str]:
        query = self.print_log_search.text().strip().lower()
        out: list[str] = []
        seen: set[str] = set()
        for r in self.print_records:
            if query:
                fn = str(r.get("filename", "")).lower()
                pt = str(r.get("path", "")).lower()
                if query not in fn and query not in pt:
                    continue
            if not self._print_row_matches_header_filters_except(r, field):
                continue
            dv = self._print_cell_display_for_filter(r, field)
            if dv not in seen:
                seen.add(dv)
                out.append(dv)
        if field == "printed_at":
            out = BillApp._unique_created_at_filter_options(out)
        else:
            out.sort(key=lambda s: (s == "", s.casefold()))
        return out

    @staticmethod
    def _merge_size_display(n: int) -> str:
        kb = 1024
        mb = kb * 1024
        if n >= mb:
            return f"{n / mb:.2f} MB"
        if n >= kb:
            return f"{n / kb:.1f} KB"
        return f"{n} B"

    def _merge_cell_display_for_filter(self, rec: dict, field: str) -> str:
        if field == "merge_filename":
            return str(rec.get("name", "") or "")
        if field == "merge_size":
            return BillApp._merge_size_display(int(rec.get("size", 0) or 0))
        if field == "merge_mtime":
            return str(rec.get("mtime", "") or "")
        if field == "merge_path":
            return str(rec.get("path_display", "") or "")
        return ""

    def _merge_row_matches_header_filters_except(self, rec: dict, skip_field: str | None) -> bool:
        filters = getattr(self, "_merge_header_filters", {})
        for field, vals in filters.items():
            if skip_field and field == skip_field:
                continue
            if not vals:
                continue
            if self._merge_cell_display_for_filter(rec, field) not in vals:
                return False
        return True

    def _unique_display_values_merge(self, field: str) -> list[str]:
        rows = list(getattr(self, "_merge_rows_all", []))
        query = str(getattr(self, "_merge_active_query", "") or "").strip().lower()
        out: list[str] = []
        seen: set[str] = set()
        for r in rows:
            if query and query not in str(r.get("name", "")).lower():
                continue
            if not self._merge_row_matches_header_filters_except(r, field):
                continue
            dv = self._merge_cell_display_for_filter(r, field)
            if dv not in seen:
                seen.add(dv)
                out.append(dv)
        out.sort(key=lambda s: (s == "", s.casefold()))
        return out

    def _sort_key_print_field(self, rec: dict, field: str) -> Any:
        if field == "row_count":
            try:
                return int(rec.get("row_count", 0) or 0)
            except (TypeError, ValueError):
                return -1
        if field == "include_print_url":
            return 1 if rec.get("include_print_url") else 0
        if field == "file_exists":
            return 1 if self._print_record_file_ok(rec) else 0
        if field == "printed_at":
            return str(rec.get("printed_at", "") or "").strip()
        if field == "filename":
            return str(rec.get("filename", "") or "").lower()
        if field == "source":
            return str(rec.get("source", "") or "").lower()
        if field == "path":
            return str(rec.get("path", "") or "").lower()
        return ""

    def _toggle_print_sort_for_field(self, field: str):
        if self.print_sort_field == field:
            if self.print_sort_order == Qt.SortOrder.AscendingOrder:
                self.print_sort_order = Qt.SortOrder.DescendingOrder
            else:
                self.print_sort_field = None
        else:
            self.print_sort_field = field
            self.print_sort_order = Qt.SortOrder.AscendingOrder
        self.refresh_print_records_table()

    def _on_print_rec_header_section_clicked(self, section: int):
        if section < 1 or section > len(PRINT_LOG_DATA_FIELDS):
            return
        field = PRINT_LOG_DATA_FIELDS[section - 1]
        self._toggle_print_sort_for_field(field)

    def _update_print_log_header_sort_indicator(self):
        h = self.print_log_table.horizontalHeader()
        h.setSortIndicatorShown(False)
        if not self.print_sort_field or self.print_sort_field not in PRINT_LOG_DATA_FIELDS:
            return
        si = PRINT_LOG_DATA_FIELDS.index(self.print_sort_field) + 1
        h.setSortIndicatorShown(True)
        h.setSortIndicator(si, self.print_sort_order)

    def _update_print_log_header_tooltips(self):
        for s in range(self.print_log_table.columnCount()):
            field = self._field_for_header_section("print_rec", s)
            if not field:
                continue
            title = PRINT_LOG_HEADERS.get(field, field)
            tip = title + "\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.print_header_filters.get(field)
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            it = self.print_log_table.horizontalHeaderItem(s)
            if it:
                it.setToolTip(tip)

    def update_print_log_header_check(self):
        it = self.print_log_table.horizontalHeaderItem(0)
        if not it:
            return
        if not self.print_log_filtered_indices:
            it.setText("☐")
            return
        checked_count = sum(1 for i in self.print_log_filtered_indices if self.print_records[i].get("checked"))
        n = len(self.print_log_filtered_indices)
        symbol = "☐" if checked_count == 0 else ("☑" if checked_count == n else "◩")
        it.setText(symbol)

    def _on_print_log_header_col0_clicked(self):
        if not self.print_log_filtered_indices:
            return
        all_checked = all(self.print_records[i].get("checked") for i in self.print_log_filtered_indices)
        for i in self.print_log_filtered_indices:
            self.print_records[i]["checked"] = not all_checked
        self.refresh_print_records_table()

    def on_print_log_cell_clicked(self, row: int, col: int):
        if self.updating_table or col != 0:
            return
        if row < 0 or row >= len(self.print_log_filtered_indices):
            return
        rec_idx = self.print_log_filtered_indices[row]
        cur = bool(self.print_records[rec_idx].get("checked"))
        self.print_records[rec_idx]["checked"] = not cur
        self.refresh_print_records_table()

    def on_print_log_checkbox_changed(self, rec_idx: int, checked: bool):
        if self.updating_table:
            return
        if 0 <= rec_idx < len(self.print_records):
            self.print_records[rec_idx]["checked"] = bool(checked)
            self.refresh_print_records_table()

    def on_print_log_search_clicked(self):
        self.print_header_filters.clear()
        self.refresh_print_records_table()

    def clear_print_log_search_and_filters(self):
        self.print_log_search.setText("")
        self.print_header_filters.clear()
        self.print_sort_field = "printed_at"
        self.print_sort_order = Qt.SortOrder.DescendingOrder
        self.refresh_print_records_table()

    def refresh_print_records_table(self):
        self.updating_table = True
        query = self.print_log_search.text().strip().lower()
        self.print_log_filtered_indices = [
            i
            for i, r in enumerate(self.print_records)
            if (
                not query
                or query in str(r.get("filename", "")).lower()
                or query in str(r.get("source", "")).lower()
                or query in str(r.get("path", "")).lower()
            )
            and self._print_row_matches_header_filters(r)
        ]
        if not self.print_sort_field:
            self.print_log_filtered_indices.sort(
                key=lambda i: str(self.print_records[i].get("printed_at", "") or ""), reverse=True
            )
        else:
            f = self.print_sort_field
            rev = self.print_sort_order == Qt.SortOrder.DescendingOrder
            self.print_log_filtered_indices.sort(
                key=lambda i, ff=f: self._sort_key_print_field(self.print_records[i], ff), reverse=rev
            )
        n = len(self.print_log_filtered_indices)
        self.print_log_table.setRowCount(n)
        h = self._data_row_height
        for row, rec_idx in enumerate(self.print_log_filtered_indices):
            rec = self.print_records[rec_idx]
            box_wrap = QWidget()
            box_layout = QHBoxLayout(box_wrap)
            box_layout.setContentsMargins(0, 0, 0, 0)
            box_layout.setAlignment(Qt.AlignCenter)
            chk = QCheckBox()
            chk.setChecked(bool(rec.get("checked")))
            chk.stateChanged.connect(lambda _s, x=rec_idx, c=chk: self.on_print_log_checkbox_changed(x, c.isChecked()))
            box_layout.addWidget(chk)
            self.print_log_table.setCellWidget(row, 0, box_wrap)
            ph0 = QTableWidgetItem("")
            ph0.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.print_log_table.setItem(row, 0, ph0)

            fn = str(rec.get("filename", "") or "")
            it_fn = QTableWidgetItem(fn)
            it_fn.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            it_fn.setToolTip(fn)
            it_fn.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 1, it_fn)

            src = str(rec.get("source", "") or "")
            src_icon = "合" if src == "合并" else "印"
            it_src = QTableWidgetItem(src_icon)
            it_src.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_src.setToolTip(src)
            it_src.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 2, it_src)

            it_time = QTableWidgetItem(str(rec.get("printed_at", "") or ""))
            it_time.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_time.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 3, it_time)

            rc = int(rec.get("row_count", 0) or 0)
            it_rc = QTableWidgetItem(str(rc))
            it_rc.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_rc.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 4, it_rc)

            source = str(rec.get("source", "") or "")
            iurl_text = "—" if source == "合并" else ("是" if rec.get("include_print_url") else "否")
            iurl = QTableWidgetItem(iurl_text)
            iurl.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            iurl.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 5, iurl)

            ok = self._print_record_file_ok(rec)
            st = QTableWidgetItem("有" if ok else "无")
            st.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            st.setForeground(QColor("#2e8b57" if ok else "#c93042"))
            st.setToolTip("文件存在" if ok else "文件缺失")
            st.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 6, st)

            pth = str(rec.get("path", "") or "")
            disp = pth if len(pth) <= 64 else pth[:30] + "…" + pth[-28:]
            it_p = QTableWidgetItem(disp)
            it_p.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            it_p.setToolTip(pth)
            it_p.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.print_log_table.setItem(row, 7, it_p)

            act = QWidget()
            al = QHBoxLayout(act)
            al.setContentsMargins(4, 2, 4, 2)
            al.setSpacing(6)
            al.setAlignment(Qt.AlignmentFlag.AlignCenter)
            bp = QPushButton("👁 预览")
            bp.setObjectName("btnGhost")
            bp.setFixedSize(84, 30)
            bp.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            bp.clicked.connect(lambda _=False, p=pth: self.preview_print_record_file(p))
            bd = QPushButton("📂 打开")
            bd.setObjectName("btnGhost")
            bd.setFixedSize(84, 30)
            bd.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            bd.clicked.connect(lambda _=False, p=pth: self.open_print_record_folder(p))
            al.addWidget(bp)
            al.addWidget(bd)
            self.print_log_table.setCellWidget(row, 8, act)

        for r in range(n):
            self.print_log_table.setRowHeight(r, h)
        self._update_print_log_header_sort_indicator()
        self._update_print_log_header_tooltips()
        self.update_print_log_header_check()
        self.updating_table = False

    def preview_print_record_file(self, path: str):
        p = Path(path)
        if not p.is_file():
            QMessageBox.warning(self, "预览", "文件不存在或已被移动，可勾选后批量删除无效记录。")
            return
        if p.suffix.lower() != ".xlsx":
            QMessageBox.information(self, "预览", "暂仅支持预览 .xlsx 表格内容。")
            return
        hint = ""
        try:
            wb = load_workbook(str(p), read_only=True, data_only=True)
            try:
                ws = wb[wb.sheetnames[0]]
                max_rows = min(int(ws.max_row or 1), 200)
                max_cols = min(int(ws.max_column or 1), 64)
                lines: list[str] = []
                for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                    cells: list[str] = []
                    for v in row:
                        if v is None:
                            cells.append("")
                        else:
                            cells.append(str(v).replace("\t", " ").replace("\n", " "))
                    lines.append("\t".join(cells))
                text = "\n".join(lines)
                if (ws.max_row or 0) > 200:
                    hint = f"\n\n…（仅显示前 200 行，共 {ws.max_row} 行）"
            finally:
                wb.close()
        except Exception as e:
            QMessageBox.warning(self, "预览失败", str(e))
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"预览 — {p.name}")
        dlg.resize(960, 720)
        dlg.setMinimumSize(640, 400)
        te = QPlainTextEdit()
        te.setReadOnly(True)
        te.setPlainText(text + hint)
        te.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        te.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        te.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(8, 8, 8, 8)
        lo.addWidget(te, 1)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        bb.rejected.connect(dlg.reject)
        lo.addWidget(bb)
        dlg.exec()

    def open_print_record_folder(self, path: str):
        p = Path(path)
        if not p.is_file():
            QMessageBox.warning(self, "打开文件夹", "文件不存在或已被移动。")
            return
        try:
            if sys.platform.startswith("win"):
                # Windows 下优先异步打开目录；失败时异步定位文件。
                try:
                    os.startfile(str(p.parent))  # type: ignore[attr-defined]
                except OSError:
                    subprocess.Popen(["explorer", "/select,", str(p)])
            elif sys.platform == "darwin":
                subprocess.Popen(["open", "-R", str(p)])
            else:
                subprocess.Popen(["xdg-open", str(p.parent)])
        except Exception as e:
            QMessageBox.warning(self, "打开文件夹失败", str(e))
            return

    @staticmethod
    def _collect_excel_files_under(folder: Path) -> list[Path]:
        exts = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
        files = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in exts]
        files.sort(key=lambda x: str(x).lower())
        return files

    def _pick_merge_roots(self) -> list[Path]:
        """选择多个待合并文件夹；不支持多选时回退为单选。"""
        dlg = QFileDialog(self, "选择待合并文件夹", str(APP_DIR))
        dlg.setFileMode(QFileDialog.FileMode.Directory)
        dlg.setOption(QFileDialog.Option.ShowDirsOnly, True)
        dlg.setOption(QFileDialog.Option.DontUseNativeDialog, True)
        views = list(dlg.findChildren(QListView)) + list(dlg.findChildren(QTreeView))
        for view in views:
            view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        picked: list[Path] = []
        if dlg.exec():
            seen: set[str] = set()
            for p in dlg.selectedFiles():
                sp = str(Path(p))
                if sp in seen:
                    continue
                seen.add(sp)
                picked.append(Path(p))
        if picked:
            return picked
        single = QFileDialog.getExistingDirectory(self, "选择待合并文件夹", str(APP_DIR))
        return [Path(single)] if single else []

    def open_merge_excel_dialog(self):
        merge_roots = self._pick_merge_roots()
        if not merge_roots:
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("合并Excel")
        dlg.resize(980, 640)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)

        root_tip = QLabel("")
        root_tip.setObjectName("hintLabel")
        lay.addWidget(root_tip)

        folder_bar = QHBoxLayout()
        btn_add_folder = QPushButton("➕ 添加文件夹")
        btn_add_folder.setObjectName("btnGhost")
        folder_bar.addWidget(btn_add_folder)
        folder_bar.addStretch(1)
        lay.addLayout(folder_bar)

        top = QHBoxLayout()
        name_search = QLineEdit()
        name_search.setPlaceholderText("检索文件名（支持关键字）...")
        name_search.setFixedWidth(200)
        btn_search = QPushButton("🔍 搜索")
        btn_search.setObjectName("btnGhost")
        btn_clear = QPushButton("🔄 清空筛选")
        btn_clear.setObjectName("btnGhost")
        top.addWidget(name_search)
        top.addWidget(btn_search)
        top.addWidget(btn_clear)
        top.addStretch(1)
        lay.addLayout(top)

        file_table = QTableWidget(0, 6)
        file_table.setHorizontalHeaderLabels(["☐", "文件名", "大小", "日期", "路径名", "操作"])
        h0 = file_table.horizontalHeaderItem(0)
        if h0:
            h0.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        file_table.verticalHeader().setVisible(False)
        file_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        file_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        file_table.setAlternatingRowColors(True)
        file_table.setColumnWidth(0, 44)
        hdr = HoverFilterHeaderView(file_table, self, "merge_excel")
        file_table.setHorizontalHeader(hdr)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        file_table.setColumnWidth(1, 360)
        file_table.setColumnWidth(2, 110)
        file_table.setColumnWidth(3, 170)
        file_table.setColumnWidth(4, 280)
        file_table.setColumnWidth(5, 188)
        hdr.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        lay.addWidget(file_table, 1)

        foot = QHBoxLayout()
        lbl_stat = QLabel("共 0 个文件，已勾选 0 个")
        foot.addWidget(lbl_stat)
        foot.addStretch(1)
        btn_merge = QPushButton("🧩 合并选中")
        btn_merge.setObjectName("btnAccent")
        btn_cancel = QPushButton("取消")
        btn_cancel.setObjectName("btnGhost")
        foot.addWidget(btn_merge)
        foot.addWidget(btn_cancel)
        lay.addLayout(foot)

        all_files: list[dict[str, Any]] = []
        checked_map: dict[str, bool] = {}
        row_paths: list[str] = []
        sort_state = {"field": None, "asc": True}
        active_query = {"text": ""}
        self._merge_header_filters = {}
        self._merge_rows_all = []
        self._merge_active_query = ""
        self._merge_render_cb = None
        self._merge_filter_parent = dlg
        self._merge_on_header_click_cb = None
        self._merge_file_table = file_table

        def update_header_check():
            if not row_paths:
                file_table.horizontalHeaderItem(0).setText("☐")
                file_table.horizontalHeaderItem(0).setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                return
            checked_cnt = sum(1 for p in row_paths if checked_map.get(p, True))
            symbol = "☐" if checked_cnt == 0 else ("☑" if checked_cnt == len(row_paths) else "◩")
            file_table.horizontalHeaderItem(0).setText(symbol)
            file_table.horizontalHeaderItem(0).setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        def update_stat():
            total = len(row_paths)
            checked = sum(1 for p in row_paths if checked_map.get(p, True))
            lbl_stat.setText(f"共 {total} 个文件，已勾选 {checked} 个")
            update_header_check()

        def sort_key(rec: dict[str, Any]):
            f = sort_state["field"]
            if f == 1:
                return str(rec.get("name", "")).lower()
            if f == 2:
                return int(rec.get("size", 0))
            if f == 3:
                return float(rec.get("mtime_ts", 0))
            if f == 4:
                return str(rec.get("path_display", "")).lower()
            return str(rec.get("name", "")).lower()

        def pass_filters(rec: dict[str, Any]) -> bool:
            kw_name = active_query["text"]
            if kw_name and kw_name not in str(rec.get("name", "")).lower():
                return False
            if not self._merge_row_matches_header_filters_except(rec, None):
                return False
            return True

        def filtered_header_text(base: str, section: int) -> str:
            fmap = {1: "merge_filename", 2: "merge_size", 3: "merge_mtime", 4: "merge_path"}
            field = fmap.get(section, "")
            kw = ""
            if field:
                vals = self._merge_header_filters.get(field, set())
                if vals:
                    kw = str(len(vals))
            return f"{base} (筛:{kw})" if kw else base

        def refresh_header_labels():
            file_table.setHorizontalHeaderItem(1, QTableWidgetItem(filtered_header_text("文件名", 1)))
            file_table.setHorizontalHeaderItem(2, QTableWidgetItem(filtered_header_text("大小", 2)))
            file_table.setHorizontalHeaderItem(3, QTableWidgetItem(filtered_header_text("日期", 3)))
            file_table.setHorizontalHeaderItem(4, QTableWidgetItem(filtered_header_text("路径名", 4)))
            file_table.setHorizontalHeaderItem(5, QTableWidgetItem("操作"))

        def update_merge_header_tooltips():
            tips = {1: "文件名", 2: "大小", 3: "日期", 4: "路径名"}
            for col, title in tips.items():
                it = file_table.horizontalHeaderItem(col)
                if not it:
                    continue
                tip = title + "\n单击：排序；悬停右侧 ▼ 筛选"
                fmap = {1: "merge_filename", 2: "merge_size", 3: "merge_mtime", 4: "merge_path"}
                f = fmap.get(col, "")
                vals = self._merge_header_filters.get(f, set()) if f else set()
                if vals:
                    tip += f"\n已选 {len(vals)} 项"
                it.setToolTip(tip)

        def refresh_root_tip():
            if not merge_roots:
                root_tip.setText("当前路径：未选择")
                return
            names = "；".join(str(p) for p in merge_roots)
            root_tip.setText(f"当前路径（{len(merge_roots)}个）：{names}")

        def render_table():
            nonlocal row_paths
            row_paths = []
            rows = [r for r in all_files if pass_filters(r)]
            self._merge_rows_all = list(all_files)
            self._merge_active_query = active_query["text"]
            if sort_state["field"] in (1, 2, 3):
                rows.sort(key=sort_key, reverse=not sort_state["asc"])
            file_table.blockSignals(True)
            try:
                file_table.setRowCount(len(rows))
                for row, rec in enumerate(rows):
                    p = str(rec["path"])
                    row_paths.append(p)
                    ph = QTableWidgetItem("")
                    ph.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                    file_table.setItem(row, 0, ph)

                    wrap = QWidget()
                    wl = QHBoxLayout(wrap)
                    wl.setContentsMargins(0, 0, 0, 0)
                    wl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    chk = QCheckBox()
                    chk.setChecked(checked_map.get(p, True))
                    chk.stateChanged.connect(
                        lambda _s=0, fp=p, c=chk: (
                            checked_map.__setitem__(fp, c.isChecked()),
                            update_stat(),
                        )
                    )
                    wl.addWidget(chk)
                    file_table.setCellWidget(row, 0, wrap)

                    it_name = QTableWidgetItem(str(rec["name"]))
                    it_name.setToolTip(str(rec["path"]))
                    file_table.setItem(row, 1, it_name)
                    it_size = QTableWidgetItem(BillApp._merge_size_display(int(rec["size"])))
                    it_size.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    file_table.setItem(row, 2, it_size)
                    it_time = QTableWidgetItem(str(rec["mtime"]))
                    it_time.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    file_table.setItem(row, 3, it_time)
                    it_path = QTableWidgetItem(str(rec.get("path_display", "")))
                    it_path.setToolTip(str(rec["path"]))
                    it_path.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                    file_table.setItem(row, 4, it_path)

                    op_wrap = QWidget()
                    op_lo = QHBoxLayout(op_wrap)
                    op_lo.setContentsMargins(4, 2, 4, 2)
                    op_lo.setSpacing(6)
                    op_lo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    bp = QPushButton("👁 预览")
                    bp.setObjectName("btnGhost")
                    bp.setFixedSize(84, 26)
                    bp.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                    bd = QPushButton("📂 打开")
                    bd.setObjectName("btnGhost")
                    bd.setFixedSize(84, 26)
                    bd.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                    fp = str(rec["path"])
                    bp.clicked.connect(lambda _=False, p=fp: self.preview_print_record_file(p))
                    bd.clicked.connect(lambda _=False, p=fp: self.open_print_record_folder(p))
                    op_lo.addWidget(bp)
                    op_lo.addWidget(bd)
                    file_table.setCellWidget(row, 5, op_wrap)
                    file_table.setRowHeight(row, 42)
            finally:
                file_table.blockSignals(False)
            refresh_header_labels()
            update_merge_header_tooltips()
            update_stat()
            self._merge_render_cb = render_table

        def load_files():
            nonlocal all_files
            files: list[Path] = []
            seen_paths: set[str] = set()
            for root in merge_roots:
                for p in self._collect_excel_files_under(root):
                    sp = str(p)
                    if sp in seen_paths:
                        continue
                    seen_paths.add(sp)
                    files.append(p)
            all_files = []
            for p in files:
                try:
                    st = p.stat()
                    mtime_dt = datetime.fromtimestamp(st.st_mtime)
                    mtime_s = mtime_dt.strftime("%Y/%m/%d %H:%M:%S")
                    size = int(st.st_size or 0)
                    mts = float(st.st_mtime)
                except OSError:
                    mtime_s = ""
                    size = 0
                    mts = 0.0
                all_files.append(
                    {
                        "path": p,
                        "name": p.name,
                        "size": size,
                        "mtime": mtime_s,
                        "mtime_ts": mts,
                        "path_display": self._merge_path_display_for_roots(p, merge_roots),
                    }
                )
                checked_map.setdefault(str(p), False)
            render_table()

        def add_merge_folder():
            cur = str(merge_roots[-1] if merge_roots else APP_DIR)
            picked = QFileDialog.getExistingDirectory(dlg, "添加待合并文件夹", cur)
            if not picked:
                return
            new_root = Path(picked)
            if any(str(x) == str(new_root) for x in merge_roots):
                QMessageBox.information(dlg, "提示", "该文件夹已添加")
                return
            merge_roots.append(new_root)
            refresh_root_tip()
            load_files()

        def on_header_clicked(section: int):
            if section == 0:
                if not row_paths:
                    return
                all_checked = all(checked_map.get(p, True) for p in row_paths)
                for p in row_paths:
                    checked_map[p] = not all_checked
                render_table()
                return
            if section not in (1, 2, 3, 4):
                return
            if sort_state["field"] == section:
                if sort_state["asc"]:
                    sort_state["asc"] = False
                else:
                    sort_state["field"] = None
                    sort_state["asc"] = True
            else:
                sort_state["field"] = section
                sort_state["asc"] = True
            render_table()
        self._merge_on_header_click_cb = on_header_clicked

        def on_cell_clicked(row: int, _col: int):
            if row < 0 or row >= len(row_paths):
                return
            p = row_paths[row]
            checked_map[p] = not checked_map.get(p, True)
            render_table()

        def on_header_menu(pos):
            section = hdr.logicalIndexAt(pos)
            menu = QMenu(file_table)
            if section in (1, 2, 3, 4):
                action_set = menu.addAction("打开本列筛选")
                action_clear = menu.addAction("清除本列过滤")
                menu.addSeparator()
                action_clear_all = menu.addAction("清除全部过滤")
                chosen = menu.exec(hdr.mapToGlobal(pos))
                if chosen is action_set:
                    self._open_header_filter_from_header("merge_excel", section)
                elif chosen is action_clear:
                    f = self._field_for_header_section("merge_excel", section)
                    if f:
                        self._merge_header_filters.pop(f, None)
                    render_table()
                elif chosen is action_clear_all:
                    self._merge_header_filters.clear()
                    render_table()
            else:
                action_clear_all = menu.addAction("清除全部过滤")
                chosen = menu.exec(hdr.mapToGlobal(pos))
                if chosen is action_clear_all:
                    self._merge_header_filters.clear()
                    render_table()

        def merge_selected():
            selected_paths: list[Path] = []
            for rec in all_files:
                p = str(rec["path"])
                if checked_map.get(p, True):
                    selected_paths.append(Path(p))
            if not selected_paths:
                QMessageBox.information(dlg, "提示", "请先勾选至少一个 Excel 文件")
                return
            confirm_box = QMessageBox(dlg)
            confirm_box.setIcon(QMessageBox.Icon.Question)
            confirm_box.setWindowTitle("确认合并")
            confirm_box.setText(f"确定要把选中的这 {len(selected_paths)} 个 Excel 文件合并吗？")
            btn_ok = confirm_box.addButton("OK", QMessageBox.ButtonRole.AcceptRole)
            confirm_box.addButton("取消", QMessageBox.ButtonRole.RejectRole)
            confirm_box.exec()
            if confirm_box.clickedButton() is not btn_ok:
                return
            out_wb = None
            out_ws = None
            header_written = False
            header_cols = 0
            merged_file_count = 0
            merged_data_rows = 0
            failed_files: list[str] = []
            header_ref_canon: tuple[str, ...] = tuple()
            merged_rows: list[list[Any]] = []

            def _canon(vals: list[Any], width: int) -> tuple[str, ...]:
                arr = list(vals)
                if width > 0:
                    if len(arr) < width:
                        arr.extend([None] * (width - len(arr)))
                    elif len(arr) > width:
                        arr = arr[:width]
                return tuple(str(v).strip().lower() if v is not None else "" for v in arr)

            def _normalize(vals: list[Any], width: int) -> list[Any]:
                arr = list(vals)
                if width > 0:
                    if len(arr) < width:
                        arr.extend([None] * (width - len(arr)))
                    elif len(arr) > width:
                        arr = arr[:width]
                return arr

            def _has_effective_data(vals: list[Any]) -> bool:
                return any(str(v).strip() for v in vals if v is not None)
            template_path: Path | None = None
            for p in selected_paths:
                if p.is_file() and p.suffix.lower() != ".xls":
                    template_path = p
                    break
            if template_path is None:
                QMessageBox.warning(dlg, "合并失败", "未找到可作为模板的文件（.xls 暂不支持）。")
                return
            try:
                out_wb = load_workbook(str(template_path))
                out_ws = out_wb[out_wb.sheetnames[0]]
            except Exception as e:
                QMessageBox.warning(dlg, "合并失败", f"模板文件读取失败：\n{template_path}\n\n{e}")
                return
            template_rows = list(out_ws.iter_rows(values_only=True))
            header_idx = None
            for i, row in enumerate(template_rows):
                if row is not None and any(str(v).strip() for v in row if v is not None):
                    header_idx = i
                    break
            if header_idx is None:
                QMessageBox.warning(dlg, "合并失败", f"模板文件无有效表头：\n{template_path}")
                return
            header_written = True
            header_cols = len(list(template_rows[header_idx]))
            header_ref_canon = _canon(list(template_rows[header_idx]), header_cols)
            for row in template_rows[header_idx + 1 :]:
                if row is None:
                    continue
                vals = _normalize(list(row), header_cols)
                if not _has_effective_data(vals):
                    continue
                if _canon(vals, header_cols) == header_ref_canon:
                    continue
                merged_rows.append(vals)
                merged_data_rows += 1
            merged_file_count = 1
            for fp in selected_paths:
                if fp == template_path:
                    continue
                if not fp.is_file():
                    failed_files.append(f"{fp.name}（文件不存在）")
                    continue
                if fp.suffix.lower() == ".xls":
                    failed_files.append(f"{fp.name}（暂不支持 .xls，请先另存为 .xlsx）")
                    continue
                try:
                    wb = load_workbook(str(fp), read_only=True, data_only=True)
                    try:
                        ws = wb[wb.sheetnames[0]]
                        rows_all = list(ws.iter_rows(values_only=True))
                        if not rows_all:
                            continue
                        header_idx = None
                        for i, row in enumerate(rows_all):
                            if row is None:
                                continue
                            if any(str(v).strip() for v in row if v is not None):
                                header_idx = i
                                break
                        if header_idx is None:
                            continue
                        for row in rows_all[header_idx + 1 :]:
                            if row is None:
                                continue
                            vals = _normalize(list(row), header_cols)
                            if not _has_effective_data(vals):
                                continue
                            if _canon(vals, header_cols) == header_ref_canon:
                                continue
                            merged_rows.append(vals)
                            merged_data_rows += 1
                        merged_file_count += 1
                    finally:
                        wb.close()
                except Exception as e:
                    failed_files.append(f"{fp.name}（{e}）")
            base_dir = self._default_output_dir_for("merge")
            try:
                base_dir.mkdir(parents=True, exist_ok=True)
            except OSError as e:
                QMessageBox.warning(dlg, "合并失败", f"创建目录失败：\n{base_dir}\n\n{e}")
                return
            suggest_name = self._build_merge_default_filename(merged_rows, merged_data_rows)
            out_default = base_dir / suggest_name
            out_file, _ = QFileDialog.getSaveFileName(
                dlg,
                "选择保存位置并命名文件",
                str(out_default),
                "Excel (*.xlsx)",
            )
            if not out_file:
                return
            out_path = Path(out_file)
            if out_path.suffix.lower() != ".xlsx":
                out_path = out_path.with_suffix(".xlsx")
            try:
                out_path.parent.mkdir(parents=True, exist_ok=True)
            except OSError as e:
                QMessageBox.warning(dlg, "合并失败", f"创建目录失败：\n{out_path.parent}\n\n{e}")
                return
            try:
                assert out_wb is not None
                assert out_ws is not None
                # 固定从模板表头下一行开始连续写入，避免 append 因模板历史样式把数据追加到很靠后位置。
                write_row = header_idx + 2
                for vals in merged_rows:
                    for ci in range(header_cols):
                        out_ws.cell(row=write_row, column=ci + 1, value=vals[ci] if ci < len(vals) else None)
                    write_row += 1
                self._write_merge_stats_sheet(out_wb, merged_rows)
                out_wb.save(str(out_path))
            except OSError as e:
                QMessageBox.warning(dlg, "合并失败", f"保存文件失败：\n{out_path}\n\n{e}")
                return
            msg = f"已基于模板「{template_path.name}」合并 {merged_file_count} 个文件，输出文件：\n{out_path}"
            if failed_files:
                shown = "\n".join(failed_files[:6])
                more = f"\n... 其余 {len(failed_files) - 6} 个失败文件未展开" if len(failed_files) > 6 else ""
                msg += f"\n\n以下文件合并失败（最多显示6条）：\n{shown}{more}"
            stamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            self.print_records.insert(
                0,
                {
                    "id": str(uuid.uuid4()),
                    "path": str(out_path),
                    "filename": out_path.name,
                    "source": "合并",
                    "printed_at": stamp,
                    "row_count": max(0, merged_data_rows),
                    "include_print_url": False,
                },
            )
            self.save_print_records()
            if self.content_stack.currentIndex() == 2:
                self.refresh_print_records_table()
            box = QMessageBox(dlg)
            box.setIcon(QMessageBox.Icon.Information)
            box.setWindowTitle("合并成功")
            box.setText(msg)
            btn_open_dir = box.addButton("打开文件夹", QMessageBox.ButtonRole.ActionRole)
            box.addButton(QMessageBox.StandardButton.Ok)
            box.exec()
            if box.clickedButton() is btn_open_dir:
                self.open_print_record_folder(str(out_path))
            dlg.accept()

        def on_search_clicked():
            active_query["text"] = name_search.text().strip().lower()
            self._merge_header_filters.clear()
            render_table()

        def on_clear_clicked():
            name_search.setText("")
            active_query["text"] = ""
            self._merge_header_filters.clear()
            sort_state["field"] = None
            sort_state["asc"] = True
            render_table()

        hdr.sectionClicked.connect(on_header_clicked)
        hdr.customContextMenuRequested.connect(on_header_menu)
        file_table.cellClicked.connect(on_cell_clicked)
        btn_search.clicked.connect(on_search_clicked)
        btn_clear.clicked.connect(on_clear_clicked)
        btn_add_folder.clicked.connect(add_merge_folder)
        btn_merge.clicked.connect(merge_selected)
        btn_cancel.clicked.connect(dlg.reject)
        refresh_root_tip()
        load_files()
        dlg.exec()
        self._merge_header_filters = {}
        self._merge_rows_all = []
        self._merge_active_query = ""
        self._merge_render_cb = None
        self._merge_filter_parent = None
        self._merge_on_header_click_cb = None
        self._merge_file_table = None

    @staticmethod
    def _merge_path_display_for_roots(path: Path, roots: list[Path]) -> str:
        for root in roots:
            try:
                rel = path.parent.relative_to(root)
                return f"{root.name}/{rel}" if str(rel) != "." else root.name
            except ValueError:
                continue
        return str(path.parent)

    def delete_selected_print_records(self):
        to_del = [i for i, r in enumerate(self.print_records) if r.get("checked")]
        if not to_del:
            QMessageBox.information(self, "提示", "请先勾选要删除的打印记录")
            return
        if (
            QMessageBox.question(
                self,
                "确认删除",
                f"将删除 {len(to_del)} 条打印记录。仅从列表移除记录，不会删除磁盘上的 Excel 文件。",
            )
            != QMessageBox.Yes
        ):
            return
        for idx in sorted(to_del, reverse=True):
            self.print_records.pop(idx)
        self.save_print_records()
        self.refresh_print_records_table()

    def add_row(self):
        insert_at = self.filtered_indices[0] if self.filtered_indices else 0
        self.records.insert(insert_at, self.default_record())
        self._shift_field_errors_after_insert(insert_at)
        self.save_data()
        self.refresh_table()

    def import_bill_excel(self):
        in_file, _ = QFileDialog.getOpenFileName(
            self,
            f"选择{self._current_business_noun()}导入文件",
            str(APP_DIR),
            "Excel (*.xlsx *.xlsm)",
        )
        if not in_file:
            return
        try:
            wb = load_workbook(in_file, read_only=True, data_only=True)
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"读取Excel失败：\n{e}")
            return

        # 第1行表头；第2行起为数据。从 A 列起依次为：
        # 任务名、类型、运营商、行业编码、url、数量、时长、年龄上限、年龄下限、pv、省份、排除省份、地市、排除地市
        col_map = {
            "task_name": 1,
            "type_code": 2,
            "operator_code": 3,
            "industry_code": 4,
            "url": 5,
            "quantity": 6,
            "duration": 7,
            "age_max": 8,
            "age_min": 9,
            "pv": 10,
            "province": 11,
            "exclude_province": 12,
            "city": 13,
            "exclude_city": 14,
        }
        type_name_to_code = {str(v): str(k) for k, v in TYPE_MAP.items()}
        op_name_to_code = {str(v): str(k) for k, v in OP_MAP.items()}

        def to_text(v: Any) -> str:
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        added = 0
        skipped = 0
        progress: QProgressDialog | None = None
        try:
            ws = wb[wb.sheetnames[0]]
            start_row = 2
            max_row = ws.max_row or 1
            total_rows = max(0, max_row - start_row + 1)
            progress = QProgressDialog(f"正在导入{self._current_business_noun()}数据...", "取消", 0, max(1, total_rows), self)
            progress.setWindowTitle("导入中")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setAutoClose(True)
            progress.setAutoReset(True)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            fields_in_order = list(col_map.keys())
            for idx, row_vals in enumerate(
                ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=14, values_only=True),
                start=1,
            ):
                if progress.wasCanceled():
                    break
                vals: dict[str, str] = {}
                for i, f in enumerate(fields_in_order):
                    vals[f] = to_text(row_vals[i] if i < len(row_vals) else "")
                if not any(vals.values()):
                    progress.setValue(idx)
                    if idx % 10 == 0 or idx == total_rows:
                        QApplication.processEvents()
                    continue
                rec = self.default_record()
                rec.update(vals)
                rec["url"] = rec["url"].replace("\r\n", "\n").replace("\r", "\n")
                if not str(rec.get("age_max", "")).strip():
                    rec["age_max"] = "55"
                if not str(rec.get("pv", "")).strip():
                    rec["pv"] = "1"
                tp = rec.get("type_code", "")
                op = rec.get("operator_code", "")
                rec["type_code"] = type_name_to_code.get(tp, tp)
                rec["operator_code"] = op_name_to_code.get(op, op)
                if str(rec.get("task_name", "")).strip() and (
                    not str(rec.get("province", "")).strip() and not str(rec.get("city", "")).strip()
                ):
                    # 导入时若省/市为空，则按任务名自动解析地域并回填。
                    self.after_field_change(rec, "task_name")
                self.records.append(rec)
                added += 1
                progress.setValue(idx)
                if idx % 10 == 0 or idx == total_rows:
                    QApplication.processEvents()
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"读取数据失败：\n{e}")
            return
        finally:
            if progress is not None:
                progress.setValue(progress.maximum())
                QApplication.processEvents()
                progress.close()
                progress.deleteLater()
                QApplication.processEvents()
            wb.close()

        if progress is not None and progress.wasCanceled():
            self.save_data()
            self.refresh_table()
            self.statusBar().showMessage(f"导入已取消：已导入 {added} 条数据", 5000)
            return
        if added <= 0:
            self.statusBar().showMessage("导入结果：未导入任何数据（仅识别第2行起 A~N 列非空数据）", 5000)
            return
        self.save_data()
        self.refresh_table()
        self.statusBar().showMessage(f"导入成功：已导入 {added} 条数据", 5000)

    def is_row_saved(self, rec):
        return any(str(rec.get(k, "")).strip() for k in FIELDS if k != "allow_print_url")

    def delete_selected(self):
        to_del = [i for i in range(len(self.records)) if self.records[i].get("checked")]
        if not to_del:
            QMessageBox.information(self, "提示", "请先勾选要删除的数据")
            return
        has_saved = any(self.is_row_saved(self.records[i]) for i in to_del)
        msg = "选中的记录中包含已保存数据，确认删除吗？此操作不可撤销。" if has_saved else "确认删除选中的记录吗？"
        if QMessageBox.question(self, "确认删除", msg) != QMessageBox.Yes:
            return
        for idx in sorted(to_del, reverse=True):
            rec = self.records[idx]
            if self.is_row_saved(rec):
                self.archive_record(rec)
            self.records.pop(idx)
            self._shift_field_errors_after_delete(idx)
        self.save_data()
        self.save_history_data()
        self.refresh_table()

    def delete_row(self, idx):
        if idx < 0 or idx >= len(self.records):
            return
        has_saved = self.is_row_saved(self.records[idx])
        msg = "该行包含已保存数据，确认删除吗？此操作不可撤销。" if has_saved else "确认删除该行吗？"
        if QMessageBox.question(self, "确认删除", msg) != QMessageBox.Yes:
            return
        rec = self.records[idx]
        if self.is_row_saved(rec):
            self.archive_record(rec)
        self.records.pop(idx)
        self._shift_field_errors_after_delete(idx)
        self.save_data()
        self.save_history_data()
        self.refresh_table()

    def archive_record(self, rec):
        hist = dict(rec)
        hist["checked"] = False
        hist["deleted_at"] = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        self.history_records.insert(0, hist)

    def clear_history_search_and_filters(self):
        self.history_search.setText("")
        self.history_header_filters.clear()
        self.history_sort_field = None
        self.refresh_history_table()

    def refresh_history_table(self):
        self.updating_table = True
        query = self.history_search.text().strip().lower()
        self.history_filtered_indices = [
            i
            for i, r in enumerate(self.history_records)
            if (not query or query in str(r.get("task_name", "")).lower()) and self._history_row_matches_header_filters(r)
        ]
        if not self.history_sort_field:
            self.history_filtered_indices.sort(
                key=lambda i: self._created_at_sort_key(self.history_records[i]), reverse=True
            )
        else:
            f = self.history_sort_field
            rev = self.history_sort_order == Qt.SortOrder.DescendingOrder
            self.history_filtered_indices.sort(
                key=lambda i, ff=f: self._sort_key_history_field(self.history_records[i], ff),
                reverse=rev,
            )
        n = len(self.history_filtered_indices)
        self.history_table.setRowCount(n)
        self.history_table_frozen.setRowCount(n)
        for row, rec_idx in enumerate(self.history_filtered_indices):
            rec = self.history_records[rec_idx]
            box_wrap = QWidget()
            box_layout = QHBoxLayout(box_wrap)
            box_layout.setContentsMargins(0, 0, 0, 0)
            box_layout.setAlignment(Qt.AlignCenter)
            chk = QCheckBox()
            chk.setChecked(bool(rec.get("checked")))
            chk.stateChanged.connect(lambda _=0, x=rec_idx, c=chk: self.on_history_checkbox_changed(x, c.isChecked()))
            box_layout.addWidget(chk)
            self.history_table_frozen.setCellWidget(row, 0, box_wrap)
            ph0 = QTableWidgetItem("")
            ph0.setFlags(Qt.ItemIsEnabled)
            self.history_table_frozen.setItem(row, 0, ph0)
            tn_item = QTableWidgetItem(self.display_val(rec, "task_name"))
            tn_item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            tn_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.history_table_frozen.setItem(row, 1, tn_item)
            for sc, f in enumerate(FIELDS[1:]):
                if f == "allow_print_url":
                    yes = coerce_allow_print_url(rec.get("allow_print_url"))
                    aw_item = QTableWidgetItem("是" if yes else "否")
                    aw_item.setFont(QFont(self.history_table_frozen.font()))
                    aw_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    aw_item.setToolTip("允许打印URL：是" if yes else "允许打印URL：否")
                    aw_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                    self.history_table.setItem(row, sc, aw_item)
                    continue
                item = QTableWidgetItem(self.display_val(rec, f))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.history_table.setItem(row, sc, item)
            ca_hist = QTableWidgetItem(str(rec.get("created_at", "")))
            ca_hist.setTextAlignment(Qt.AlignCenter)
            ca_hist.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.history_table.setItem(row, len(FIELDS) - 1, ca_hist)
            pc_hist = QTableWidgetItem(str(int(rec.get("print_count", 0) or 0)))
            pc_hist.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pc_hist.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.history_table.setItem(row, len(FIELDS), pc_hist)
            pt_hist = QTableWidgetItem(str(rec.get("last_printed_at", "") or ""))
            pt_hist.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pt_hist.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.history_table.setItem(row, len(FIELDS) + 1, pt_hist)
            deleted_item = QTableWidgetItem(str(rec.get("deleted_at", "")))
            deleted_item.setTextAlignment(Qt.AlignCenter)
            deleted_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.history_table.setItem(row, len(FIELDS) + 2, deleted_item)
        h = self._data_row_height
        for r in range(n):
            self.history_table.setRowHeight(r, h)
            self.history_table_frozen.setRowHeight(r, h)
        self._update_history_header_sort_indicator()
        self._update_history_header_tooltips()
        self.update_history_header_check()
        self.updating_table = False

    def update_history_header_check(self):
        it = self.history_table_frozen.horizontalHeaderItem(0)
        if not it:
            return
        if not self.history_filtered_indices:
            it.setText("☐")
            return
        checked_count = sum(1 for i in self.history_filtered_indices if self.history_records[i].get("checked"))
        symbol = "☐" if checked_count == 0 else ("☑" if checked_count == len(self.history_filtered_indices) else "◩")
        it.setText(symbol)

    def on_history_cell_clicked(self, row, col):
        if self.updating_table or col != 0:
            return
        if row < 0 or row >= len(self.history_filtered_indices):
            return
        rec_idx = self.history_filtered_indices[row]
        self.on_history_checkbox_changed(rec_idx, not self.history_records[rec_idx].get("checked", False))

    def on_history_checkbox_changed(self, rec_idx: int, checked: bool):
        if self.updating_table:
            return
        if rec_idx < 0 or rec_idx >= len(self.history_records):
            return
        self.history_records[rec_idx]["checked"] = bool(checked)
        self.save_history_data()
        self.refresh_history_table()

    def restore_selected_history(self):
        restore_idx = [i for i, rec in enumerate(self.history_records) if rec.get("checked")]
        if not restore_idx:
            QMessageBox.information(self, "提示", "请先勾选要恢复的数据")
            return
        if (
            QMessageBox.question(
                self,
                "确认恢复",
                f"确认恢复 {len(restore_idx)} 条历史任务到{self._current_business_table_name()}吗？",
            )
            != QMessageBox.Yes
        ):
            return
        for idx in sorted(restore_idx):
            src = self.history_records[idx]
            rec = {k: src.get(k, "") for k in FIELDS}
            rec["allow_print_url"] = coerce_allow_print_url(src.get("allow_print_url"))
            try:
                rec["print_count"] = int(src.get("print_count", 0) or 0)
            except (TypeError, ValueError):
                rec["print_count"] = 0
            rec["last_printed_at"] = str(src.get("last_printed_at", "") or "")
            rec["checked"] = False
            rec["created_at"] = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            self.records.append(rec)
        for idx in sorted(restore_idx, reverse=True):
            self.history_records.pop(idx)
        self.save_data()
        self.save_history_data()
        self.refresh_table()
        self.refresh_history_table()
        QMessageBox.information(
            self,
            "恢复成功",
            f"已恢复 {len(restore_idx)} 条数据到{self._current_business_table_name()}",
        )

    def delete_selected_history(self):
        to_del = [i for i, rec in enumerate(self.history_records) if rec.get("checked")]
        if not to_del:
            QMessageBox.information(self, "提示", "请先勾选要删除的历史任务")
            return
        if QMessageBox.question(self, "确认删除", f"确认删除 {len(to_del)} 条历史任务吗？此操作不可撤销。") != QMessageBox.Yes:
            return
        for idx in sorted(to_del, reverse=True):
            self.history_records.pop(idx)
        self.save_history_data()
        self.refresh_history_table()

    def display_val(self, rec, field):
        if field == "type_code":
            return TYPE_MAP.get(rec.get(field, ""), "")
        if field == "operator_code":
            return OP_MAP.get(rec.get(field, ""), "")
        if field == "allow_print_url":
            return "是" if coerce_allow_print_url(rec.get("allow_print_url")) else "否"
        if field == "url":
            descs = self._selected_url_descs_for_record(rec)
            return "\n".join(descs)
        if field == "print_count":
            return str(int(rec.get("print_count", 0) or 0))
        if field == "last_printed_at":
            return str(rec.get("last_printed_at", "") or "")
        return str(rec.get(field, ""))

    @staticmethod
    def _geo_tokens_invalid(field: str, parts: list[str], rec: dict | None = None) -> list[str]:
        if field in ("province", "exclude_province"):
            allowed = set(PROVINCES)
            return [p for p in parts if p not in allowed]
        allow_cities: set[str] | None = None
        if field == "exclude_city" and rec is not None:
            provs = split_multi(rec.get("province", ""))
            if provs:
                allow_cities = set(cities_under_provinces(provs))
        if allow_cities is not None:
            return [p for p in parts if p not in allow_cities]
        return [p for p in parts if p not in set(CITIES)]

    @staticmethod
    def _normalize_geo_field_value(field: str, raw: str, rec: dict | None = None) -> str | None:
        text = str(raw or "").replace("\n", " ").replace("\r", " ").strip()
        if not text:
            return ""
        parts = split_multi(text.replace("，", "|").replace(",", "|"))
        if field in ("city", "exclude_city"):
            allow_set: set[str] | None = None
            if field == "exclude_city" and rec is not None:
                provs = split_multi(rec.get("province", ""))
                if provs:
                    allow_set = set(cities_under_provinces(provs))
            norm_parts: list[str] = []
            for p in parts:
                n = normalize_city_name(p, allow_set)
                if n is None:
                    return None
                if n and n not in norm_parts:
                    norm_parts.append(n)
            return "|".join(norm_parts)
        if BillApp._geo_tokens_invalid(field, parts, rec):
            return None
        return "|".join(parts)

    @staticmethod
    def _city_picker_source(field: str, rec: dict) -> list[str]:
        """地市类弹窗候选项：排除地市在已选省份时仅显示这些省下的地市。"""
        if field == "exclude_city" and split_multi(rec.get("province", "")):
            return cities_under_provinces(split_multi(rec.get("province", "")))
        if field in ("city", "exclude_city"):
            return list(CITIES)
        raise ValueError(field)

    @staticmethod
    def _created_at_sort_key(rec: dict) -> str:
        return str(rec.get("created_at", "") or "").strip()

    @staticmethod
    def _created_at_filter_key(raw: Any) -> str:
        """表头筛选用：只保留 YYYY-MM-DD，同一日期去重；选项中不显示时分秒。
        亦用于「打印时间」last_printed_at / 打印记录 printed_at（支持 YYYY/MM/DD …）。"""
        t = str(raw or "").strip()
        if not t:
            return ""
        # 常见：YYYY-MM-DD 后接空格/T 及时间
        if re.match(r"^\d{4}-\d{2}-\d{2}", t):
            head = t[:10]
            if len(t) == 10 or (len(t) > 10 and t[10] in " Tt"):
                return head
        if "T" in t[:32]:
            left = t.split("T", 1)[0].strip()
            if len(left) >= 10 and re.match(r"^\d{4}-\d{2}-\d{2}$", left[:10]):
                return left[:10]
        m = re.match(r"^(\d{4})[./-](\d{1,2})[./-](\d{1,2})", t)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return datetime(y, mo, d).strftime("%Y-%m-%d")
            except ValueError:
                pass
        try:
            if "T" in t:
                s = t.replace("Z", "").split("+", 1)[0].split(".", 1)[0][:19]
                return datetime.fromisoformat(s).strftime("%Y-%m-%d")
            if re.search(r"\d{1,2}:\d{2}", t):
                return datetime.strptime(t[:19], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        except Exception:
            pass
        try:
            return datetime.strptime(t[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
        except Exception:
            pass
        m2 = re.match(r"^(\d{4}-\d{2}-\d{2})", t)
        if m2:
            return m2.group(1)
        return ""

    def _main_table_sort_key(self, rec_idx: int) -> tuple:
        """未保存（无提单时间）的草稿排在最前；同组内按下标排序（新增行插在首条可见记录之前）。"""
        ca = self._created_at_sort_key(self.records[rec_idx])
        if ca:
            return (0, ca)
        return (1, -rec_idx)

    def on_bill_search_clicked(self):
        """搜索按钮：刷新列表并清空所有表头列筛选。"""
        self.header_filters.clear()
        self.refresh_table()

    def on_history_search_clicked(self):
        """历史页搜索按钮：刷新并清空表头列筛选。"""
        self.history_header_filters.clear()
        self.refresh_history_table()

    def clear_bill_search_and_filters(self):
        self.search.setText("")
        self.header_filters.clear()
        self.header_sort_field = None
        self.refresh_table()

    def _sync_main_v_scroll(self, value: int, source: str):
        if self._table_v_sync:
            return
        self._table_v_sync = True
        try:
            if source == "scroll":
                self.table_frozen.verticalScrollBar().setValue(value)
            else:
                self.table.verticalScrollBar().setValue(value)
        finally:
            self._table_v_sync = False

    def _sync_selection_main_to_frozen(self):
        if self._table_sel_sync or self.table.selectionModel() is None:
            return
        self._table_sel_sync = True
        try:
            self.table_frozen.clearSelection()
            for idx in self.table.selectionModel().selectedRows():
                self.table_frozen.selectRow(idx.row())
        finally:
            self._table_sel_sync = False

    def _sync_selection_frozen_to_main(self):
        if self._table_sel_sync or self.table_frozen.selectionModel() is None:
            return
        self._table_sel_sync = True
        try:
            self.table.clearSelection()
            for idx in self.table_frozen.selectionModel().selectedRows():
                self.table.selectRow(idx.row())
        finally:
            self._table_sel_sync = False

    def _cell_display_for_filter_main(self, rec: dict, field: str) -> str:
        if field == "created_at":
            return BillApp._created_at_filter_key(rec.get("created_at"))
        if field == "print_count":
            return str(int(rec.get("print_count", 0) or 0))
        if field == "last_printed_at":
            return BillApp._created_at_filter_key(rec.get("last_printed_at"))
        if field in ("type_code", "operator_code"):
            return self.display_val(rec, field) or str(rec.get(field, "") or "")
        if field == "allow_print_url":
            return self.display_val(rec, "allow_print_url")
        if field == "url":
            return first_url(rec.get(field, ""))
        return str(rec.get(field, "") or "")

    def _main_row_matches_header_filters_except(self, rec_idx: int, skip_field: str | None) -> bool:
        rec = self.records[rec_idx]
        for field, vals in self.header_filters.items():
            if skip_field and field == skip_field:
                continue
            if not vals:
                continue
            if field == "created_at":
                dk = BillApp._created_at_filter_key(rec.get("created_at"))
                allowed = {BillApp._created_at_filter_key(x) for x in vals}
                if dk not in allowed:
                    return False
            elif field == "last_printed_at":
                dk = BillApp._created_at_filter_key(rec.get("last_printed_at"))
                allowed = {BillApp._created_at_filter_key(x) for x in vals}
                if dk not in allowed:
                    return False
            elif self._cell_display_for_filter_main(rec, field) not in vals:
                return False
        return True

    def _main_row_matches_header_filters(self, rec_idx: int) -> bool:
        return self._main_row_matches_header_filters_except(rec_idx, None)

    def _unique_display_values_main(self, field: str) -> list[str]:
        query = self.search.text().strip().lower()
        out: list[str] = []
        seen: set[str] = set()
        for i, r in enumerate(self.records):
            if query and query not in str(r.get("task_name", "")).lower():
                continue
            if not self._main_row_matches_header_filters_except(i, field):
                continue
            dv = self._cell_display_for_filter_main(r, field)
            if dv not in seen:
                seen.add(dv)
                out.append(dv)
        if field == "created_at":
            out = BillApp._unique_created_at_filter_options(out)
        elif field == "last_printed_at":
            out = BillApp._unique_created_at_filter_options(out)
        else:
            out.sort(key=lambda s: (s == "", s.casefold()))
        return out

    @staticmethod
    def _unique_created_at_filter_options(values: list[str]) -> list[str]:
        """筛选项：仅 YYYY-MM-DD、按日期去重；空日期保留一条；日期新在前。"""
        seen: set[str] = set()
        norm: list[str] = []
        for x in values:
            k = BillApp._created_at_filter_key(x)
            if k in seen:
                continue
            seen.add(k)
            norm.append(k)

        def sort_key(d: str):
            if not d:
                return (2, "")
            try:
                return (0, -datetime.strptime(d, "%Y-%m-%d").timestamp())
            except Exception:
                return (1, d)

        norm.sort(key=sort_key)
        return norm

    def _field_for_header_section(self, mode: str, section: int) -> str | None:
        if mode in ("main_frozen", "hist_frozen"):
            return None if section == 0 else "task_name"
        if mode == "print_rec":
            if section == 0 or section > len(PRINT_LOG_DATA_FIELDS):
                return None
            return PRINT_LOG_DATA_FIELDS[section - 1]
        if mode == "merge_excel":
            if section == 1:
                return "merge_filename"
            if section == 2:
                return "merge_size"
            if section == 3:
                return "merge_mtime"
            if section == 4:
                return "merge_path"
            return None
        if mode == "acc_frozen":
            return "acc_type" if section == 1 else None
        if mode == "acc_scroll":
            fmap = ["acc_channel", "acc_name", "acc_desc", "acc_url", "acc_remark", "acc_created", None]
            if 0 <= section < len(fmap):
                return fmap[section]
            return None
        if mode == "main_scroll":
            return self._main_scroll_field_for_section(section)
        if mode == "hist_scroll":
            return self._history_scroll_field_for_section(section)
        return None

    def _header_show_filter_btn(self, mode: str, section: int) -> bool:
        return self._field_for_header_section(mode, section) is not None

    def _on_merge_header_section_clicked(self, section: int):
        cb = getattr(self, "_merge_on_header_click_cb", None)
        if callable(cb):
            cb(section)

    def _default_output_base_dir(self) -> Path:
        """输出目录：优先可执行文件目录（便于 U 盘携带），不可写时回退 APP_DIR。"""
        if getattr(sys, "frozen", False):
            exe_dir = Path(sys.executable).resolve().parent
            probe = exe_dir / ".tidanmgr_write_probe"
            try:
                probe.mkdir(parents=False, exist_ok=True)
                probe.rmdir()
                return exe_dir
            except OSError:
                return APP_DIR
        return APP_DIR

    def _default_output_dir_for(self, kind: str) -> Path:
        """按类型返回输出目录（print/merge），位于启动程序同级。"""
        root = self._default_output_base_dir()
        safe = "merge" if str(kind).strip().lower() == "merge" else "print"
        return root / safe

    def _ensure_output_dirs(self):
        """启动时预创建输出目录（print / merge）。"""
        for k in ("print", "merge"):
            try:
                self._default_output_dir_for(k).mkdir(parents=True, exist_ok=True)
            except OSError:
                # 不阻塞启动；实际导出时会再提示具体错误
                pass

    def _close_column_filter_popup(self):
        w = self._filter_popup
        if w is None:
            return
        w.close()
        w.deleteLater()
        self._filter_popup = None

    def _on_column_filter_popup_closed(self, dlg: "ColumnPickFilterPopup"):
        if self._filter_popup is dlg:
            self._filter_popup = None

    def _filter_button_global_bottom_right(self, mode: str, section: int) -> QPoint | None:
        if mode == "main_frozen":
            h = self.table_frozen.horizontalHeader()
        elif mode == "main_scroll":
            h = self.table.horizontalHeader()
        elif mode == "hist_frozen":
            h = self.history_table_frozen.horizontalHeader()
        elif mode == "hist_scroll":
            h = self.history_table.horizontalHeader()
        elif mode == "print_rec":
            h = self.print_log_table.horizontalHeader()
        elif mode == "merge_excel":
            mt = getattr(self, "_merge_file_table", None)
            if mt is None:
                return None
            h = mt.horizontalHeader()
        elif mode == "acc_frozen":
            h = self.accessory_list_frozen.horizontalHeader()
        elif mode == "acc_scroll":
            h = self.accessory_list_table.horizontalHeader()
        else:
            return None
        if not isinstance(h, HoverFilterHeaderView):
            return None
        r = h._filter_btn_rect(section)
        return h.viewport().mapToGlobal(r.bottomRight())

    def _open_header_filter_from_header(self, mode: str, section: int):
        field = self._field_for_header_section(mode, section)
        if not field:
            return
        if mode.startswith("main"):
            opts = self._unique_display_values_main(field)
            cur = set(self.header_filters.get(field, ()))
        elif mode.startswith("print"):
            opts = self._unique_display_values_print(field)
            cur = set(self.print_header_filters.get(field, ()))
        elif mode.startswith("merge"):
            opts = self._unique_display_values_merge(field)
            cur = set(getattr(self, "_merge_header_filters", {}).get(field, ()))
        elif mode.startswith("acc"):
            opts = self._unique_display_values_accessory(field)
            cur = set(self.accessory_header_filters.get(field, ()))
        else:
            opts = self._unique_display_values_hist(field)
            cur = set(self.history_header_filters.get(field, ()))
        if field in ("created_at", "last_printed_at", "printed_at"):
            cur = {BillApp._created_at_filter_key(x) for x in cur}
        acc_titles = {
            "acc_name": "名称",
            "acc_type": "类型",
            "acc_channel": "渠道",
            "acc_desc": "描述",
            "acc_url": "URL",
            "acc_remark": "备注",
            "acc_created": "创建时间",
        }
        title = HEADERS.get(field, acc_titles.get(field, field))
        self._close_column_filter_popup()
        anchor = self._filter_button_global_bottom_right(mode, section)
        if anchor is None:
            anchor = self.mapToGlobal(self.rect().topRight())
        popup_parent = getattr(self, "_merge_filter_parent", None) if mode.startswith("merge") else self
        dlg = ColumnPickFilterPopup(
            self,
            mode,
            field,
            f"筛选：{title}",
            opts,
            cur,
            anchor,
            parent=popup_parent or self,
        )
        dlg.destroyed.connect(lambda *a, d=dlg: self._on_column_filter_popup_closed(d))
        self._filter_popup = dlg
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _toggle_main_sort_for_field(self, field: str):
        if self.header_sort_field == field:
            if self.header_sort_order == Qt.SortOrder.AscendingOrder:
                self.header_sort_order = Qt.SortOrder.DescendingOrder
            else:
                self.header_sort_field = None
        else:
            self.header_sort_field = field
            self.header_sort_order = Qt.SortOrder.AscendingOrder
        self.refresh_table()

    def _on_main_frozen_header_section_clicked(self, section: int):
        if section == 0:
            # 由 HoverFilterHeaderView.mouseReleaseEvent 统一处理，避免与 sectionClicked 重复触发
            return
        self._toggle_main_sort_for_field("task_name")

    def _main_scroll_field_for_section(self, section: int) -> str | None:
        if 0 <= section < len(FIELDS) - 1:
            return FIELDS[section + 1]
        if section == len(FIELDS) - 1:
            return "created_at"
        if section == len(FIELDS):
            return "print_count"
        if section == len(FIELDS) + 1:
            return "last_printed_at"
        return None

    def _on_main_scroll_header_clicked(self, section: int):
        field = self._main_scroll_field_for_section(section)
        if not field:
            return
        self._toggle_main_sort_for_field(field)

    def _sort_key_for_main_field(self, rec_idx: int, field: str):
        rec = self.records[rec_idx]
        if field == "created_at":
            return self._created_at_sort_key(rec)
        if field == "print_count":
            try:
                return int(rec.get("print_count", 0) or 0)
            except (TypeError, ValueError):
                return -1
        if field == "last_printed_at":
            return str(rec.get("last_printed_at", "") or "").strip()
        if field in ("quantity", "age_max", "age_min", "pv"):
            raw = str(rec.get(field, "") or "").strip()
            if not raw:
                return float("-inf")
            try:
                return float(raw)
            except ValueError:
                return raw.lower()
        return self._cell_display_for_filter_main(rec, field).lower()

    def _apply_main_sort_to_filtered(self):
        if not self.header_sort_field:
            self.filtered_indices.sort(key=self._main_table_sort_key, reverse=True)
        else:
            field = self.header_sort_field
            rev = self.header_sort_order == Qt.SortOrder.DescendingOrder
            self.filtered_indices.sort(key=lambda i, f=field: self._sort_key_for_main_field(i, f), reverse=rev)

    def _update_main_header_sort_indicator(self):
        hf = self.table_frozen.horizontalHeader()
        hs = self.table.horizontalHeader()
        hf.setSortIndicatorShown(False)
        hs.setSortIndicatorShown(False)
        if not self.header_sort_field:
            return
        f = self.header_sort_field
        if f == "task_name":
            hf.setSortIndicatorShown(True)
            hf.setSortIndicator(1, self.header_sort_order)
            return
        if f in FIELDS:
            si = FIELDS.index(f) - 1
            if si >= 0:
                hs.setSortIndicatorShown(True)
                hs.setSortIndicator(si, self.header_sort_order)
            return
        if f == "created_at":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS) - 1, self.header_sort_order)
            return
        if f == "print_count":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS), self.header_sort_order)
            return
        if f == "last_printed_at":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS) + 1, self.header_sort_order)
            return

    def _update_main_header_tooltips(self):
        for s in range(self.table_frozen.columnCount()):
            field = self._field_for_header_section("main_frozen", s)
            if not field:
                continue
            tip = HEADERS.get(field, field) + "\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.header_filters.get(field)
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            it = self.table_frozen.horizontalHeaderItem(s)
            if it:
                it.setToolTip(tip)
        for s in range(self.table.columnCount()):
            field = self._main_scroll_field_for_section(s)
            if not field:
                continue
            tip = HEADERS.get(field, field) + "\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.header_filters.get(field)
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            it = self.table.horizontalHeaderItem(s)
            if it:
                it.setToolTip(tip)

    def _sync_hist_v_scroll(self, value: int, source: str):
        if self._hist_v_sync:
            return
        self._hist_v_sync = True
        try:
            if source == "scroll":
                self.history_table_frozen.verticalScrollBar().setValue(value)
            else:
                self.history_table.verticalScrollBar().setValue(value)
        finally:
            self._hist_v_sync = False

    def _sync_selection_hist_to_frozen(self):
        if self._hist_sel_sync or self.history_table.selectionModel() is None:
            return
        self._hist_sel_sync = True
        try:
            self.history_table_frozen.clearSelection()
            for idx in self.history_table.selectionModel().selectedRows():
                self.history_table_frozen.selectRow(idx.row())
        finally:
            self._hist_sel_sync = False

    def _sync_selection_hist_frozen_to_scroll(self):
        if self._hist_sel_sync or self.history_table_frozen.selectionModel() is None:
            return
        self._hist_sel_sync = True
        try:
            self.history_table.clearSelection()
            for idx in self.history_table_frozen.selectionModel().selectedRows():
                self.history_table.selectRow(idx.row())
        finally:
            self._hist_sel_sync = False

    def _cell_display_for_filter_hist(self, rec: dict, field: str) -> str:
        if field == "deleted_at":
            return str(rec.get("deleted_at", "") or "")
        if field == "print_count":
            return str(int(rec.get("print_count", 0) or 0))
        if field == "last_printed_at":
            return BillApp._created_at_filter_key(rec.get("last_printed_at"))
        if field == "created_at":
            return BillApp._created_at_filter_key(rec.get("created_at"))
        if field in ("type_code", "operator_code"):
            return self.display_val(rec, field) or str(rec.get(field, "") or "")
        if field == "allow_print_url":
            return self.display_val(rec, "allow_print_url")
        if field == "url":
            return first_url(rec.get(field, ""))
        return str(rec.get(field, "") or "")

    def _history_row_matches_header_filters_except(self, rec: dict, skip_field: str | None) -> bool:
        for field, vals in self.history_header_filters.items():
            if skip_field and field == skip_field:
                continue
            if not vals:
                continue
            if field == "created_at":
                dk = BillApp._created_at_filter_key(rec.get("created_at"))
                allowed = {BillApp._created_at_filter_key(x) for x in vals}
                if dk not in allowed:
                    return False
            elif field == "last_printed_at":
                dk = BillApp._created_at_filter_key(rec.get("last_printed_at"))
                allowed = {BillApp._created_at_filter_key(x) for x in vals}
                if dk not in allowed:
                    return False
            elif self._cell_display_for_filter_hist(rec, field) not in vals:
                return False
        return True

    def _history_row_matches_header_filters(self, rec: dict) -> bool:
        return self._history_row_matches_header_filters_except(rec, None)

    def _unique_display_values_hist(self, field: str) -> list[str]:
        query = self.history_search.text().strip().lower()
        out: list[str] = []
        seen: set[str] = set()
        for r in self.history_records:
            if query and query not in str(r.get("task_name", "")).lower():
                continue
            if not self._history_row_matches_header_filters_except(r, field):
                continue
            dv = self._cell_display_for_filter_hist(r, field)
            if dv not in seen:
                seen.add(dv)
                out.append(dv)
        if field == "created_at":
            out = BillApp._unique_created_at_filter_options(out)
        elif field == "last_printed_at":
            out = BillApp._unique_created_at_filter_options(out)
        else:
            out.sort(key=lambda s: (s == "", s.casefold()))
        return out

    def _history_scroll_field_for_section(self, section: int) -> str | None:
        if 0 <= section < HISTORY_SCROLL_COLUMNS:
            return HISTORY_SCROLL_FIELDS[section]
        return None

    def _sort_key_history_field(self, rec: dict, field: str | None):
        if not field:
            return self._created_at_sort_key(rec)
        if field == "deleted_at":
            return str(rec.get("deleted_at", "") or "")
        if field == "created_at":
            return self._created_at_sort_key(rec)
        if field == "print_count":
            try:
                return int(rec.get("print_count", 0) or 0)
            except (TypeError, ValueError):
                return -1
        if field == "last_printed_at":
            return str(rec.get("last_printed_at", "") or "").strip()
        if field in ("quantity", "age_max", "age_min", "pv"):
            raw = str(rec.get(field, "") or "").strip()
            if not raw:
                return float("-inf")
            try:
                return float(raw)
            except ValueError:
                return raw.lower()
        return self._cell_display_for_filter_hist(rec, field).lower()

    def _toggle_history_sort_for_field(self, field: str):
        if self.history_sort_field == field:
            if self.history_sort_order == Qt.SortOrder.AscendingOrder:
                self.history_sort_order = Qt.SortOrder.DescendingOrder
            else:
                self.history_sort_field = None
        else:
            self.history_sort_field = field
            self.history_sort_order = Qt.SortOrder.AscendingOrder
        self.refresh_history_table()

    def _on_hist_frozen_header_section_clicked(self, section: int):
        if section == 0:
            return
        self._toggle_history_sort_for_field("task_name")

    def _on_history_scroll_header_clicked(self, section: int):
        field = self._history_scroll_field_for_section(section)
        if not field:
            return
        self._toggle_history_sort_for_field(field)

    def _update_history_header_sort_indicator(self):
        hf = self.history_table_frozen.horizontalHeader()
        hs = self.history_table.horizontalHeader()
        hf.setSortIndicatorShown(False)
        hs.setSortIndicatorShown(False)
        if not self.history_sort_field:
            return
        f = self.history_sort_field
        if f == "task_name":
            hf.setSortIndicatorShown(True)
            hf.setSortIndicator(1, self.history_sort_order)
            return
        if f in FIELDS:
            si = FIELDS.index(f) - 1
            if si >= 0:
                hs.setSortIndicatorShown(True)
                hs.setSortIndicator(si, self.history_sort_order)
            return
        if f == "created_at":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS) - 1, self.history_sort_order)
            return
        if f == "print_count":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS), self.history_sort_order)
            return
        if f == "last_printed_at":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS) + 1, self.history_sort_order)
            return
        if f == "deleted_at":
            hs.setSortIndicatorShown(True)
            hs.setSortIndicator(len(FIELDS) + 2, self.history_sort_order)

    def _update_history_header_tooltips(self):
        for s in range(self.history_table_frozen.columnCount()):
            field = self._field_for_header_section("hist_frozen", s)
            if not field:
                continue
            tip = HEADERS.get(field, field) + "\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.history_header_filters.get(field)
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            it = self.history_table_frozen.horizontalHeaderItem(s)
            if it:
                it.setToolTip(tip)
        for s in range(self.history_table.columnCount()):
            field = self._history_scroll_field_for_section(s)
            if not field:
                continue
            title = HEADERS.get(field, field) if field in HEADERS else field
            tip = title + "\n单击：排序；悬停右侧 ▼ 筛选"
            vals = self.history_header_filters.get(field)
            if vals:
                tip += f"\n已选 {len(vals)} 项"
            it = self.history_table.horizontalHeaderItem(s)
            if it:
                it.setToolTip(tip)

    def _url_cell_editor_display_height(self) -> int:
        """URL 编辑框固定单行高度；不换行、无滚动条，用左右方向键移动光标查看超出部分。"""
        fm = QFontMetrics(self.table.font())
        inner = fm.height() + 8
        return max(26, min(self._data_row_height - 4, inner))

    def _url_plain_editor_at(self, row: int, url_col: int) -> QPlainTextEdit | None:
        """URL 列可能为外层垂直居中容器 + UrlCellEditor。"""
        w = self.table.cellWidget(row, url_col)
        if w is None:
            return None
        if isinstance(w, QPlainTextEdit) and w.objectName() == "urlCellEditor":
            return w
        ed = w.findChild(QPlainTextEdit, "urlCellEditor")
        return ed if isinstance(ed, QPlainTextEdit) else None

    def _on_main_scroll_column_resized_for_url(self, logical_index: int, _old_size: int, _new_size: int):
        url_sc = FIELDS.index("url") - 1
        if logical_index != url_sc:
            return
        self._reflow_all_url_row_heights()

    def _reflow_all_url_row_heights(self):
        """URL 列宽变化后，统一为单行框高（不随内容增高）。"""
        if self.updating_table:
            return
        url_sc = FIELDS.index("url") - 1
        nh = self._url_cell_editor_display_height()
        rh = self._data_row_height
        for row in range(self.table.rowCount()):
            wgt = self._url_plain_editor_at(row, url_sc)
            if wgt is None:
                continue
            wgt.blockSignals(True)
            wgt.setFixedHeight(nh)
            wgt.blockSignals(False)
            self.table.setRowHeight(row, rh)
            self.table_frozen.setRowHeight(row, rh)

    @staticmethod
    def _url_editor_stylesheet(has_error: bool) -> str:
        border = "#e05263" if has_error else "rgba(148,163,184,0.35)"
        return (
            f"QPlainTextEdit#urlCellEditor {{ border: 1px solid {border}; border-radius: 4px; padding: 2px; }}"
        )

    def _on_url_cell_text_changed(self, rec_idx: int, ed: QPlainTextEdit):
        if self.updating_table:
            return
        if rec_idx < 0 or rec_idx >= len(self.records):
            return
        txt = ed.toPlainText()
        if self.records[rec_idx].get("url", "") == txt:
            return
        self.records[rec_idx]["url"] = txt
        self.clear_error(rec_idx, "url")
        ok, msg = self.validate_urls(txt)
        if not ok:
            self.mark_error(rec_idx, "url", msg or "URL 格式不正确")
            ed.setStyleSheet(self._url_editor_stylesheet(True))
        else:
            ed.setStyleSheet(self._url_editor_stylesheet(False))
        self.save_data()
        self._refresh_main_row_validation_marks(rec_idx)

    def _on_url_cell_focus_out(self, rec_idx: int, ed: QPlainTextEdit):
        """URL 失焦仅同步样式；整行标红由输入/离行刷新，弹窗仅在点击打印时。"""
        if self.updating_table:
            return
        if rec_idx < 0 or rec_idx >= len(self.records):
            return
        if str(self.records[rec_idx].get("url", "")) != ed.toPlainText():
            return
        self.clear_error(rec_idx, "url")
        ok, msg = self.validate_urls(ed.toPlainText())
        if not ok:
            self.mark_error(rec_idx, "url", msg or "URL 格式不正确")
            ed.setStyleSheet(self._url_editor_stylesheet(True))
        else:
            ed.setStyleSheet(self._url_editor_stylesheet(False))
        self._refresh_main_row_validation_marks(rec_idx)

    def refresh_table(self):
        self.updating_table = True
        query = self.search.text().strip().lower()
        self.filtered_indices = [
            i
            for i, r in enumerate(self.records)
            if (not query or query in str(r.get("task_name", "")).lower()) and self._main_row_matches_header_filters(i)
        ]
        self._apply_main_sort_to_filtered()
        n = len(self.filtered_indices)
        self.table.setRowCount(n)
        self.table_frozen.setRowCount(n)
        per_row_heights: list[int] = []
        for row, rec_idx in enumerate(self.filtered_indices):
            rec = self.records[rec_idx]
            row_max_h = self._data_row_height
            box_wrap = QWidget()
            box_layout = QHBoxLayout(box_wrap)
            box_layout.setContentsMargins(0, 0, 0, 0)
            box_layout.setAlignment(Qt.AlignCenter)
            chk = QCheckBox()
            chk.setChecked(bool(rec.get("checked")))
            chk.stateChanged.connect(lambda _=0, x=rec_idx, c=chk: self.on_row_checkbox_changed(x, c.isChecked()))
            box_layout.addWidget(chk)
            self.table_frozen.setCellWidget(row, 0, box_wrap)
            ph0 = QTableWidgetItem("")
            ph0.setFlags(Qt.ItemIsEnabled)
            self.table_frozen.setItem(row, 0, ph0)
            tn_item = QTableWidgetItem(self.display_val(rec, "task_name"))
            tn_item.setFlags(
                Qt.ItemFlag.ItemIsSelectable
                | Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsEditable
            )
            tn_item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            if not str(rec.get("task_name", "")).strip():
                tn_item.setToolTip("示例：客户全国######-产品")
            if self.field_errors.get((rec_idx, "task_name")):
                tn_item.setBackground(QColor("#e05263"))
                tn_item.setToolTip(self.field_error_msgs.get((rec_idx, "task_name"), "字段校验失败"))
            self.table_frozen.setItem(row, 1, tn_item)
            for sc, f in enumerate(FIELDS[1:]):
                if f == "allow_print_url":
                    yes = coerce_allow_print_url(rec.get("allow_print_url"))
                    aw_item = QTableWidgetItem("是" if yes else "否")
                    aw_item.setFont(QFont(self.table_frozen.font()))
                    aw_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    aw_item.setToolTip(
                        "允许打印URL：是（单击切换为否）" if yes else "允许打印URL：否（单击切换为是）"
                    )
                    aw_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                    self.table.setItem(row, sc, aw_item)
                    continue
                if f == "url":
                    url_text = self.display_val(rec, f)
                    item = QTableWidgetItem(url_text)
                    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                    if url_text.strip():
                        item.setToolTip(url_text)
                    if self.field_errors.get((rec_idx, f)):
                        item.setBackground(QColor("#e05263"))
                        item.setToolTip(self.field_error_msgs.get((rec_idx, f), "字段校验失败"))
                    self.table.setItem(row, sc, item)
                    continue
                item = QTableWidgetItem(self.display_val(rec, f))
                if f in ("province", "exclude_province", "city", "exclude_city"):
                    item.setFlags(
                        Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable
                    )
                item.setTextAlignment(Qt.AlignCenter)
                if self.field_errors.get((rec_idx, f)):
                    item.setBackground(QColor("#e05263"))
                    item.setToolTip(self.field_error_msgs.get((rec_idx, f), "字段校验失败"))
                self.table.setItem(row, sc, item)
                if f in ("type_code", "operator_code", "duration"):
                    combo = QComboBox()
                    combo.addItem("-- 请选择 --", "")
                    if f == "type_code":
                        for code, name in TYPE_MAP.items():
                            combo.addItem(name, code)
                    elif f == "operator_code":
                        for code, name in OP_MAP.items():
                            combo.addItem(name, code)
                    else:
                        for val in DURATIONS:
                            combo.addItem(val, val)
                    combo.setCurrentIndex(max(0, combo.findData(rec.get(f, ""))))
                    combo.currentIndexChanged.connect(
                        lambda _=0, x=rec_idx, ff=f, cb=combo: self.on_inline_combo_changed(x, ff, cb.currentData() or "")
                    )
                    style_combo_centered(combo)
                    self.table.setCellWidget(row, sc, combo)
            ca_item = QTableWidgetItem(str(rec.get("created_at", "")))
            ca_item.setTextAlignment(Qt.AlignCenter)
            ca_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(row, len(FIELDS) - 1, ca_item)
            pc_item = QTableWidgetItem(str(int(rec.get("print_count", 0) or 0)))
            pc_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pc_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(row, len(FIELDS), pc_item)
            pt_item = QTableWidgetItem(str(rec.get("last_printed_at", "") or ""))
            pt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pt_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(row, len(FIELDS) + 1, pt_item)
            act_del = QWidget()
            del_lo = QHBoxLayout(act_del)
            del_lo.setContentsMargins(4, 2, 4, 2)
            del_lo.setSpacing(0)
            del_lo.setAlignment(Qt.AlignmentFlag.AlignCenter)
            btn = QPushButton("🗑 删除")
            btn.setObjectName("btnDanger")
            btn.setFixedSize(92, 30)
            btn.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            btn.clicked.connect(lambda _=False, x=rec_idx: self.delete_row(x))
            del_lo.addWidget(btn)
            self.table.setCellWidget(row, MAIN_SCROLL_COLUMNS - 1, act_del)
            per_row_heights.append(row_max_h)
        for r, rh in enumerate(per_row_heights):
            self.table.setRowHeight(r, rh)
            self.table_frozen.setRowHeight(r, rh)
        self.lbl_count.setText(f"{self._current_business_noun()}数: {len(self.filtered_indices)}")
        self.update_header_check()
        self._update_main_header_sort_indicator()
        self._update_main_header_tooltips()
        self.updating_table = False

    def update_header_check(self):
        if not self.filtered_indices:
            self.table_frozen.horizontalHeaderItem(0).setText("☐")
            return
        checked_count = sum(1 for i in self.filtered_indices if self.records[i].get("checked"))
        symbol = "☐" if checked_count == 0 else ("☑" if checked_count == len(self.filtered_indices) else "◩")
        self.table_frozen.horizontalHeaderItem(0).setText(symbol)

    def on_header_clicked(self, section):
        if section != 0:
            return
        if not self.filtered_indices:
            return
        all_checked = all(self.records[i].get("checked") for i in self.filtered_indices)
        for i in self.filtered_indices:
            self.records[i]["checked"] = not all_checked
        self.save_data()
        self.refresh_table()

    def _on_frozen_header_col0_clicked(self, mode: str):
        if mode == "main_frozen":
            self.on_header_clicked(0)
            return
        if mode == "hist_frozen":
            if not self.history_filtered_indices:
                return
            all_checked = all(self.history_records[i].get("checked") for i in self.history_filtered_indices)
            for i in self.history_filtered_indices:
                self.history_records[i]["checked"] = not all_checked
            self.save_history_data()
            self.refresh_history_table()
            return
        if mode == "acc_frozen":
            if not self.accessory_list_row_node_ids:
                return
            all_checked = all(nid in self._accessory_checked_node_ids for nid in self.accessory_list_row_node_ids)
            if all_checked:
                for nid in self.accessory_list_row_node_ids:
                    self._accessory_checked_node_ids.discard(nid)
            else:
                self._accessory_checked_node_ids.update(self.accessory_list_row_node_ids)
            self.refresh_accessory_tree()

    def on_cell_clicked(self, row, col):
        if self.updating_table:
            return
        if row < 0 or row >= len(self.filtered_indices):
            return
        if col == 0:
            rec_idx = self.filtered_indices[row]
            new_state = not self.records[rec_idx].get("checked", False)
            self.on_row_checkbox_changed(rec_idx, new_state)
            return
        if col == ALLOW_PRINT_URL_DISPLAY_COL:
            rec_idx = self.filtered_indices[row]
            cur = coerce_allow_print_url(self.records[rec_idx].get("allow_print_url"))
            self.records[rec_idx]["allow_print_url"] = not cur
            self.save_data()
            self.refresh_table()

    def on_row_checkbox_changed(self, rec_idx: int, checked: bool):
        if self.updating_table:
            return
        if rec_idx < 0 or rec_idx >= len(self.records):
            return
        self.records[rec_idx]["checked"] = bool(checked)
        self.save_data()
        self.refresh_table()

    def on_cell_double_click(self, row, col):
        if col <= 0 or col >= len(DISPLAY_FIELDS) - 1:
            return
        name = DISPLAY_FIELDS[col]
        if name in ("print_count", "last_printed_at", "created_at", "action"):
            return
        rec_idx = self.filtered_indices[row]
        field = name
        rec = self.records[rec_idx]
        if field == "allow_print_url":
            cur = coerce_allow_print_url(rec.get("allow_print_url"))
            rec["allow_print_url"] = not cur
            self.save_data()
            self.refresh_table()
            return
        if field == "url":
            sc = FIELDS.index("url") - 1
            idx0 = self.table.model().index(row, sc)
            rect = self.table.visualRect(idx0)
            anchor = self.table.viewport()
            anchor.setProperty("_url_anchor_global", self.table.viewport().mapToGlobal(rect.bottomLeft()))
            anchor.setProperty("_url_anchor_width", rect.width())
            self._open_url_picker_dialog(rec_idx, anchor)
            return
        if field in ("province", "exclude_province", "city", "exclude_city"):
            if field == "city" and rec.get("province", "").strip():
                QMessageBox.warning(self, "提示", "已选择省份后，地市不可再选择；请先清空省份")
                return
            if field == "province" and rec.get("city", "").strip():
                QMessageBox.warning(self, "提示", "已选择地市后，省份不可再选择；请先清空地市")
                return
            if field == "province" and split_multi(rec.get("exclude_province", "")):
                QMessageBox.warning(self, "提示", "已填写排除省份，省份不可再编辑；请先清空排除省份")
                return
            if field == "city" and (
                split_multi(rec.get("exclude_province", "")) or split_multi(rec.get("exclude_city", ""))
            ):
                QMessageBox.warning(self, "提示", "已填写排除省份或排除地市，地市不可再编辑；请先清空对应排除项")
                return
            if field == "exclude_province" and (
                rec.get("province", "").strip() or rec.get("city", "").strip()
            ):
                QMessageBox.warning(self, "提示", "已选择省份或地市后，排除省份不可再选择；请先清空省份与地市")
                return
            if field == "exclude_city" and rec.get("city", "").strip():
                QMessageBox.warning(self, "提示", "已选择地市后，排除地市不可再选择；请先清空地市")
                return
            cur_parts = split_multi(rec.get(field, ""))
            if self._geo_tokens_invalid(field, cur_parts, rec):
                QMessageBox.warning(self, "提示", "输入错误，请重新输入或者双击选择")
                return
            items = PROVINCES if "province" in field else self._city_picker_source(field, rec)
            if field == "exclude_city" and split_multi(rec.get("province", "")) and not items:
                QMessageBox.warning(self, "提示", "所选省份暂无地市级数据，无法选择排除地市。")
                return
            recent = self.picker_recent.get(field, [])
            ordered = sorted(
                items,
                key=lambda x: (
                    0 if x in split_multi(rec.get(field, "")) else 1,
                    0 if x in recent else 1,
                    recent.index(x) if x in recent else 999,
                    x,
                ),
            )
            dlg = MultiSelectDialog("选择", ordered, split_multi(rec.get(field, "")), self)
            cell_rect = self.table.visualRect(self.table.model().index(row, col - FROZEN_COLUMNS))
            anchor = self.table.viewport().mapToGlobal(cell_rect.bottomLeft())
            self._place_dialog_below_rect(dlg, anchor, cell_rect.width())
            if dlg.exec() == QDialog.Accepted:
                values = dlg.values()
                self.update_picker_recent(field, values)
                self.apply_field_update(rec_idx, field, "|".join(values))
            return

    def on_current_cell_changed(self, current_row, current_col, _previous_row, _previous_col):
        if self.updating_table:
            return
        should_refresh = False
        if 0 <= _previous_row < len(self.filtered_indices):
            prev_rec_idx = self.filtered_indices[_previous_row]
            should_refresh = self._validate_row_on_leave(prev_rec_idx)
        if should_refresh:
            self.refresh_table()
            if 0 <= current_row < len(self.filtered_indices) and 0 <= current_col < self.table.columnCount():
                self.table.setCurrentCell(current_row, current_col)
        if current_row < 0 or current_row >= len(self.filtered_indices):
            return
        if current_col <= 0 or current_col >= len(DISPLAY_FIELDS) - 1:
            return
        name = DISPLAY_FIELDS[current_col]
        if name in ("print_count", "last_printed_at", "created_at", "action"):
            return
        field = name
        if field == "task_name":
            self.statusBar().showMessage("任务名格式：客户名+省份/地市/全国/几省/几市+运营商码值+类型码值+行业编码-产品名", 5000)

    def on_table_context_menu(self, pos):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        row = idx.row()
        col = idx.column() + FROZEN_COLUMNS
        if row < 0 or row >= len(self.filtered_indices):
            return
        if col <= 0 or col >= len(DISPLAY_FIELDS) - 1:
            return
        name = DISPLAY_FIELDS[col]
        if name in ("print_count", "last_printed_at", "created_at", "action"):
            return
        field = name
        if field not in ("task_name", "operator_code", "type_code", "duration"):
            return
        rec_idx = self.filtered_indices[row]
        rec = self.records[rec_idx]
        menu = self.table.createStandardContextMenu()
        action_map: dict[object, tuple[str, str]] = {}
        if field == "task_name":
            sample = (
                f"客户{self.region_part(rec)}{rec.get('operator_code') or '##'}"
                f"{rec.get('type_code') or '##'}{rec.get('industry_code') or '##'}-产品"
            )
            action_fill = menu.addAction("套用任务名示例")
            action_map[action_fill] = ("task_name", sample)
        elif field == "operator_code":
            menu.addSeparator()
            for code, name in OP_MAP.items():
                act = menu.addAction(f"设置运营商：{name}（{code}）")
                action_map[act] = ("operator_code", code)
        elif field == "type_code":
            menu.addSeparator()
            for code, name in TYPE_MAP.items():
                act = menu.addAction(f"设置类型：{name}（{code}）")
                action_map[act] = ("type_code", code)
        elif field == "duration":
            menu.addSeparator()
            for val in DURATIONS:
                act = menu.addAction(f"设置时长：{val}")
                action_map[act] = ("duration", val)
        chosen = menu.exec(self.table.viewport().mapToGlobal(pos))
        if chosen in action_map:
            target_field, target_value = action_map[chosen]
            self.apply_field_update(rec_idx, target_field, target_value)

    def bulk_fill_selected(self):
        selected_idx = [idx for idx, rec in enumerate(self.records) if rec.get("checked")]
        if not selected_idx:
            QMessageBox.information(self, "提示", "请先勾选要批量填充的记录")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("批量填充")
        scope_combo = QComboBox(dlg)
        scope_combo.addItem("全部已勾选记录", "all_checked")
        scope_combo.addItem("仅当前筛选结果中已勾选", "filtered_checked")
        field_combo = QComboBox(dlg)
        field_combo.addItem("运营商", "operator_code")
        field_combo.addItem("类型", "type_code")
        field_combo.addItem("时长", "duration")
        field_combo.addItem("行业编码", "industry_code")
        field_combo.addItem("数量", "quantity")
        field_combo.addItem("年龄上限", "age_max")
        field_combo.addItem("年龄下限", "age_min")
        field_combo.addItem("pv", "pv")
        mode_combo = QComboBox(dlg)
        mode_combo.addItem("批量填充", "fill")
        mode_combo.addItem("批量清空", "clear")
        value_combo = QComboBox(dlg)
        value_combo.setMinimumWidth(260)
        value_edit = QLineEdit(dlg)
        value_edit.setMinimumWidth(260)

        def render_values():
            field = field_combo.currentData()
            value_combo.clear()
            value_edit.clear()
            value_edit.hide()
            value_combo.show()
            if mode_combo.currentData() == "clear":
                value_combo.setEnabled(False)
                value_edit.setEnabled(False)
                value_combo.addItem("（将清空所选字段）", "")
                return
            value_combo.setEnabled(True)
            value_edit.setEnabled(True)
            if field == "operator_code":
                for code, name in OP_MAP.items():
                    value_combo.addItem(f"{name}（{code}）", code)
            elif field == "type_code":
                for code, name in TYPE_MAP.items():
                    value_combo.addItem(f"{name}（{code}）", code)
            elif field == "duration":
                for val in DURATIONS:
                    value_combo.addItem(val, val)
            elif field == "industry_code":
                value_combo.hide()
                value_edit.show()
                value_edit.setPlaceholderText("请输入行业编码（字母或数字）")
            else:
                value_combo.hide()
                value_edit.show()
                value_edit.setPlaceholderText("请输入数字")

        field_combo.currentIndexChanged.connect(render_values)
        mode_combo.currentIndexChanged.connect(render_values)
        render_values()
        btns = QDialogButtonBox(QDialogButtonBox.Cancel | QDialogButtonBox.Ok)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("作用范围"))
        lay.addWidget(scope_combo)
        lay.addWidget(QLabel("执行方式"))
        lay.addWidget(mode_combo)
        lay.addWidget(QLabel("填充字段"))
        lay.addWidget(field_combo)
        lay.addWidget(QLabel("填充值"))
        lay.addWidget(value_combo)
        lay.addWidget(value_edit)
        lay.addWidget(btns)
        if dlg.exec() != QDialog.Accepted:
            return
        mode = mode_combo.currentData()
        field = field_combo.currentData()
        scope = scope_combo.currentData()
        if scope == "filtered_checked":
            visible_set = set(self.filtered_indices)
            selected_idx = [idx for idx in selected_idx if idx in visible_set]
            if not selected_idx:
                QMessageBox.information(self, "提示", "当前筛选结果中没有已勾选记录")
                return
        if mode == "clear":
            value = ""
        elif field in ("industry_code", "quantity", "age_max", "age_min", "pv"):
            value = value_edit.text().strip()
        else:
            value = value_combo.currentData() or ""
        if mode == "fill" and field in ("quantity", "age_max", "age_min", "pv"):
            if value and not value.isdigit():
                QMessageBox.warning(self, "批量操作", "数量、年龄上下限、pv 仅支持数字")
                return
            if field == "quantity" and value == "0":
                QMessageBox.warning(self, "批量操作", "数量必须大于0")
                return
        field_label = field_combo.currentText()
        value_label = "清空" if mode == "clear" else (value_combo.currentText() if value_combo.isVisible() else value)
        confirm = QMessageBox.question(
            self,
            "确认批量操作",
            f"即将对 {len(selected_idx)} 条记录执行：{field_label} -> {value_label}\n确认继续吗？",
        )
        if confirm != QMessageBox.Yes:
            return
        failed: list[int] = []
        fail_details: list[tuple[int, str, str]] = []
        for idx in selected_idx:
            rec = self.records[idx]
            old_rec = dict(rec)
            rec[field] = value
            self.clear_error(idx, field)
            if field in ("age_min", "age_max"):
                self.clear_error(idx, "age_min")
                self.clear_error(idx, "age_max")
            self.after_field_change(rec, field, idx)
            self._refresh_main_row_validation_marks(idx)
            ok, msg = self.validate_record(rec)
            if not ok:
                failed.append(idx)
                fail_details.append((idx, field, msg))
                self.records[idx] = old_rec
                self._refresh_main_row_validation_marks(idx)
        self.save_data()
        self.refresh_table()
        if failed:
            self.focus_record(failed[0], field)
            detail_lines = []
            for rec_idx, failed_field, msg in fail_details[:8]:
                task_name = str(self.records[rec_idx].get("task_name", "")).strip() or f"第{rec_idx + 1}行"
                field_name = HEADERS.get(failed_field, failed_field)
                detail_lines.append(f"- {task_name} | {field_name} | {msg}")
            more = ""
            if len(fail_details) > 8:
                more = f"\n- ... 其余 {len(fail_details) - 8} 条请查看表格红框"
            QMessageBox.warning(
                self,
                "批量操作",
                "已更新 "
                f"{len(selected_idx) - len(failed)} 条，失败 {len(failed)} 条。\n\n失败明细（最多显示8条）：\n"
                + "\n".join(detail_lines)
                + more,
            )
            return
        QMessageBox.information(self, "批量操作", f"已更新 {len(selected_idx)} 条记录")

    def focus_record(self, rec_idx: int, field: str):
        if rec_idx not in self.filtered_indices:
            self.search.setText("")
            self.refresh_table()
        if rec_idx not in self.filtered_indices:
            return
        row = self.filtered_indices.index(rec_idx)
        try:
            col = DISPLAY_FIELDS.index(field)
        except ValueError:
            col = 1
        if col <= 1:
            self.table_frozen.setCurrentCell(row, col)
            it0 = self.table_frozen.item(row, col)
            if it0:
                self.table_frozen.scrollToItem(it0)
            return
        sc = col - FROZEN_COLUMNS
        self.table.setCurrentCell(row, sc)
        w = self._url_plain_editor_at(row, sc)
        if isinstance(w, QPlainTextEdit):
            self.table.scrollTo(self.table.model().index(row, sc))
            w.setFocus()
            return
        it = self.table.item(row, sc)
        if it:
            self.table.scrollToItem(it)

    def on_item_changed(self, item: QTableWidgetItem):
        if self.updating_table:
            return
        tw = item.tableWidget()
        base = 0 if tw is self.table_frozen else FROZEN_COLUMNS
        row = item.row()
        col = item.column() + base
        if row < 0 or row >= len(self.filtered_indices):
            return
        rec_idx = self.filtered_indices[row]
        rec = self.records[rec_idx]
        if col == 0:
            rec["checked"] = item.checkState() == Qt.Checked
            self.save_data()
            self.update_header_check()
            return
        if col >= len(DISPLAY_FIELDS) - 1:
            return
        name = DISPLAY_FIELDS[col]
        if name in ("print_count", "last_printed_at", "created_at", "checked"):
            return
        field = name
        if field in ("type_code", "operator_code", "duration", "url", "allow_print_url"):
            return
        if field in ("province", "exclude_province", "city", "exclude_city"):
            raw = item.text()
            norm = self._normalize_geo_field_value(field, raw, rec)
            if norm is None:
                QMessageBox.warning(self, "提示", "输入错误，请重新输入或者双击选择")
                self.updating_table = True
                item.setText(self.display_val(rec, field))
                self.updating_table = False
                return
            self.apply_field_update(rec_idx, field, norm)
            return
        val = item.text().replace("\n", " ").strip()
        self.apply_field_update(rec_idx, field, val)

    def on_inline_combo_changed(self, rec_idx: int, field: str, value: str):
        if self.updating_table:
            return
        if rec_idx < 0 or rec_idx >= len(self.records):
            return
        if str(self.records[rec_idx].get(field, "")) == str(value or ""):
            return
        self.apply_field_update(rec_idx, field, value or "")

    def task_name_hint(self, text: str) -> str:
        value = str(text or "").strip()
        if not value:
            return ""
        if "-" not in value:
            return "任务名建议包含 '-' 分隔产品名"
        left = value.split("-", 1)[0]
        parsed = self.parse_left(left)
        if not parsed:
            return "任务名前半段未识别到完整码值段（运营商+类型+行业编码）"
        if not parsed.get("customer", "").strip():
            return "任务名建议包含客户名"
        if not parsed.get("region", "").strip():
            return "任务名建议包含地域（省/市/全国/几省/几市）"
        return ""

    def mark_error(self, rec_idx: int, field: str, msg: str = ""):
        self.field_errors[(rec_idx, field)] = True
        self.field_error_msgs[(rec_idx, field)] = msg or "字段校验失败"

    def clear_error(self, rec_idx: int, field: str):
        self.field_errors.pop((rec_idx, field), None)
        self.field_error_msgs.pop((rec_idx, field), None)

    def _validate_row_on_leave(self, rec_idx: int) -> bool:
        if rec_idx < 0 or rec_idx >= len(self.records):
            return False
        before = {
            f: self.field_error_msgs.get((rec_idx, f), "")
            for f in MAIN_ROW_VALIDATE_FIELDS
            if (rec_idx, f) in self.field_errors
        }
        self._refresh_main_row_validation_marks(rec_idx)
        after = {
            f: self.field_error_msgs.get((rec_idx, f), "")
            for f in MAIN_ROW_VALIDATE_FIELDS
            if (rec_idx, f) in self.field_errors
        }
        return before != after

    def _shift_field_errors_after_insert(self, insert_at: int):
        """在 insert_at 插入空行后，原索引 >= insert_at 的校验错误键整体 +1。"""
        new_err: dict[tuple[int, str], bool] = {}
        new_msg: dict[tuple[int, str], str] = {}
        for (ri, f), v in self.field_errors.items():
            new_err[(ri + 1 if ri >= insert_at else ri, f)] = v
        for (ri, f), v in self.field_error_msgs.items():
            new_msg[(ri + 1 if ri >= insert_at else ri, f)] = v
        self.field_errors = new_err
        self.field_error_msgs = new_msg

    def _shift_field_errors_after_delete(self, deleted_idx: int):
        """删除 deleted_idx 行后：去掉该行错误；原索引 > deleted_idx 的键整体 -1。"""
        new_err: dict[tuple[int, str], bool] = {}
        new_msg: dict[tuple[int, str], str] = {}
        for (ri, f), v in self.field_errors.items():
            if ri == deleted_idx:
                continue
            new_err[(ri - 1 if ri > deleted_idx else ri, f)] = v
        for (ri, f), v in self.field_error_msgs.items():
            if ri == deleted_idx:
                continue
            new_msg[(ri - 1 if ri > deleted_idx else ri, f)] = v
        self.field_errors = new_err
        self.field_error_msgs = new_msg

    def apply_field_update(self, rec_idx: int, field: str, value: str):
        if field in ("created_at", "print_count", "last_printed_at"):
            return
        rec = self.records[rec_idx]
        cleaned = value.replace("\n", " ").replace("\r", " ").strip() if field == "task_name" else value
        if field == "age_max":
            s = str(cleaned or "").strip()
            cleaned = s if s else "55"
        elif field == "pv":
            s = str(cleaned or "").strip()
            cleaned = s if s else "1"
        if field == "task_name":
            hint = self.task_name_hint(cleaned)
            if hint:
                self.statusBar().showMessage(hint, 5000)
        # 单元格提交不回滚；标红由 _refresh_main_row_validation_marks 按导出规则刷新（不弹窗）。
        rec[field] = cleaned
        self.clear_error(rec_idx, field)
        if field in ("age_min", "age_max"):
            self.clear_error(rec_idx, "age_min")
            self.clear_error(rec_idx, "age_max")
        changed_ok = self.after_field_change(rec, field, rec_idx)
        if field == "task_name" and changed_ok:
            self.clear_error(rec_idx, "task_name")
        soft_task_name_warn = ""
        if field == "task_name" and (not changed_ok) and (not getattr(self, "_task_name_parse_err_marked", False)):
            soft_task_name_warn = "任务名未按规范完全解析：已保留原文本，仅同步可识别片段"
        self._refresh_main_row_validation_marks(rec_idx)
        if soft_task_name_warn:
            self.mark_error(rec_idx, "task_name", soft_task_name_warn)
        self.save_data()
        self.refresh_table()

    def validate_record(self, rec):
        for k, title in [("quantity", "数量"), ("age_max", "年龄上限"), ("age_min", "年龄下限"), ("pv", "pv")]:
            v = str(rec.get(k, "")).strip()
            if v and not v.isdigit():
                return False, f"{title} 必须是数字"
        ok_url, url_msg = self.validate_urls(str(rec.get("url", "")))
        if not ok_url:
            return False, url_msg
        if rec.get("quantity", "").strip() == "0":
            return False, "数量必须大于0"
        amin, amax = rec.get("age_min", "").strip(), rec.get("age_max", "").strip()
        if amin and amax and int(amax) < int(amin):
            return False, "年龄上限必须大于等于年龄下限"
        return True, ""

    def validate_urls(self, raw: str):
        lines = [x.strip() for x in str(raw or "").replace("\r\n", "\n").replace("\r", "\n").replace("|", "\n").split("\n") if x.strip()]
        if not lines:
            return False, "URL 为必填项"
        return True, ""

    def _validate_urls_collect(self, raw: str) -> list[str]:
        """URL 字段仅做必填校验。"""
        lines = [x.strip() for x in str(raw or "").replace("\r\n", "\n").replace("\r", "\n").replace("|", "\n").split("\n") if x.strip()]
        return [] if lines else ["URL 为必填项"]

    def validate_export_record(self, rec):
        required_fields = [
            ("task_name", "任务名"),
            ("type_code", "类型"),
            ("operator_code", "运营商"),
            ("industry_code", "行业编码"),
            ("url", "URL"),
            ("quantity", "数量"),
            ("duration", "时长"),
        ]
        for field, title in required_fields:
            if not str(rec.get(field, "")).strip():
                return False, f"{title} 为必填项"
        tn = str(rec.get("task_name", "")).strip()
        if "-" not in tn:
            return False, "任务名格式错误：缺少 '-' 分隔"
        parsed, perrs = self.parse_left_with_errors(tn.split("-", 1)[0])
        if perrs:
            return False, perrs[0]
        if not parsed:
            return False, "任务名格式错误：无法解析运营商/类型/行业编码"
        if parsed["op"] != rec.get("operator_code", ""):
            return False, "任务名中的运营商码值与运营商字段不匹配"
        if parsed["tp"] != rec.get("type_code", ""):
            return False, "任务名中的类型码值与类型字段不匹配"
        if parsed["ind"] != rec.get("industry_code", ""):
            return False, "任务名中的行业编码与行业编码字段不匹配"
        region_from_task = parsed.get("region", "") or "全国"
        region_from_fields = self.region_part(rec)
        if region_from_task != region_from_fields:
            return False, f"任务名中的地域“{region_from_task}”与字段地域“{region_from_fields}”不匹配"
        return True, ""

    def collect_record_validation_issues(self, rec: dict) -> list[tuple[str, str]]:
        """收集主表一行全部校验问题（格式校验 + 导出必填/任务名关联），供标红与打印前汇总。"""
        issues: list[tuple[str, str]] = []
        for k, title in [("quantity", "数量"), ("age_max", "年龄上限"), ("age_min", "年龄下限"), ("pv", "pv")]:
            v = str(rec.get(k, "")).strip()
            if v and not v.isdigit():
                issues.append((k, f"{title} 必须是数字"))
        for msg in self._validate_urls_collect(str(rec.get("url", ""))):
            issues.append(("url", msg))
        if str(rec.get("quantity", "")).strip() == "0":
            issues.append(("quantity", "数量必须大于0"))
        amin, amax = str(rec.get("age_min", "")).strip(), str(rec.get("age_max", "")).strip()
        if amin and amax and amin.isdigit() and amax.isdigit() and int(amax) < int(amin):
            pair = "年龄上限必须大于等于年龄下限"
            issues.append(("age_min", pair))
            issues.append(("age_max", pair))
        required_fields = [
            ("task_name", "任务名"),
            ("type_code", "类型"),
            ("operator_code", "运营商"),
            ("industry_code", "行业编码"),
            ("url", "URL"),
            ("quantity", "数量"),
            ("duration", "时长"),
        ]
        for field, title in required_fields:
            if not str(rec.get(field, "")).strip():
                issues.append((field, f"{title} 为必填项"))
        tn = str(rec.get("task_name", "")).strip()
        if tn and "-" not in tn:
            issues.append(("task_name", "任务名格式错误：缺少 '-' 分隔"))
        elif tn and "-" in tn:
            left_part = tn.split("-", 1)[0]
            parsed, perrs = self.parse_left_with_errors(left_part)
            if perrs:
                for msg in perrs:
                    issues.append(("task_name", msg))
            elif not parsed:
                issues.append(("task_name", "任务名格式错误：无法解析运营商/类型/行业编码"))
            else:
                if parsed["op"] != rec.get("operator_code", ""):
                    m = "任务名中的运营商码值与运营商字段不匹配"
                    issues.append(("operator_code", m))
                    issues.append(("task_name", m))
                if parsed["tp"] != rec.get("type_code", ""):
                    m = "任务名中的类型码值与类型字段不匹配"
                    issues.append(("type_code", m))
                    issues.append(("task_name", m))
                if parsed["ind"] != rec.get("industry_code", ""):
                    m = "任务名中的行业编码与行业编码字段不匹配"
                    issues.append(("industry_code", m))
                    issues.append(("task_name", m))
                region_from_task = parsed.get("region", "") or "全国"
                region_from_fields = self.region_part(rec)
                if region_from_task != region_from_fields:
                    m = f"任务名中的地域“{region_from_task}”与字段地域“{region_from_fields}”不匹配"
                    issues.append(("task_name", m))
                    issues.append(("province", m))
                    issues.append(("city", m))
        return issues

    def _merge_validation_issues_by_field(self, issues: list[tuple[str, str]]) -> dict[str, str]:
        merged_lists: dict[str, list[str]] = {}
        for f, m in issues:
            merged_lists.setdefault(f, []).append(m)
        return {f: "\n".join(dict.fromkeys(msgs)) for f, msgs in merged_lists.items()}

    def _clear_main_row_validation_errors(self, rec_idx: int):
        for f in MAIN_ROW_VALIDATE_FIELDS:
            self.clear_error(rec_idx, f)

    def _mark_row_from_validation_issues(self, rec_idx: int, issues: list[tuple[str, str]]):
        merged = self._merge_validation_issues_by_field(issues)
        for f, msg in merged.items():
            self.mark_error(rec_idx, f, msg)

    def _place_dialog_below_rect(self, dlg: QDialog, anchor_global: QPoint, anchor_w: int) -> None:
        """将弹窗定位到锚点正下方，且尽量不越出当前屏幕。"""
        screen = QGuiApplication.screenAt(anchor_global) or QGuiApplication.primaryScreen()
        geo = screen.availableGeometry() if screen else QRect(0, 0, 1920, 1080)
        size = dlg.sizeHint()
        x = anchor_global.x()
        y = anchor_global.y() + 2
        if x + size.width() > geo.right():
            x = geo.right() - size.width()
        if x < geo.left():
            x = geo.left()
        # 锚点宽度较大时，尽量水平居中到控件下方
        if anchor_w > size.width():
            x = max(geo.left(), min(anchor_global.x() + (anchor_w - size.width()) // 2, geo.right() - size.width()))
        if y + size.height() > geo.bottom():
            y = max(geo.top(), anchor_global.y() - size.height() - 2)
        dlg.move(x, y)

    def _task_no_from_task_name(self, task_name: str) -> str:
        tn = str(task_name or "").strip()
        if not tn:
            return ""
        left = tn.split("-", 1)[0]
        reg = find_earliest_region_in_left(left)
        if not reg:
            return ""
        start, _end, _region = reg
        return left[:start].strip()

    def _build_merge_default_filename(self, merged_rows: list[list[Any]], merged_data_rows: int) -> str:
        prefix = "合并结果"
        if merged_rows:
            first = merged_rows[0]
            if len(first) >= 2:
                first_task_no = self._task_no_from_task_name(str(first[1] or ""))
                if first_task_no:
                    prefix = first_task_no
        return f"{sanitize_filename(prefix)}-{max(0, merged_data_rows)}.xlsx"

    def _open_url_picker_dialog(self, rec_idx: int, anchor_widget: QWidget):
        dlg = QDialog(self)
        dlg.setWindowTitle("选择配件URL")
        dlg.resize(620, 460)
        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(10, 10, 10, 10)
        lo.setSpacing(8)
        search = QLineEdit()
        search.setPlaceholderText("搜索节点：名称/描述/URL...")
        lo.addWidget(search)
        tree = QTreeWidget()
        tree.setHeaderLabels(["名称", "URL"])
        tree.setColumnWidth(0, 280)
        tree.setColumnWidth(1, 360)
        lo.addWidget(tree, 1)
        btns = QDialogButtonBox(QDialogButtonBox.Cancel | QDialogButtonBox.Ok)
        lo.addWidget(btns)
        picked: dict[str, Any] = {"urls": [], "descs": []}
        rec = self.records[rec_idx] if 0 <= rec_idx < len(self.records) else {}
        pre_values = set(self._selected_url_descs_for_record(rec))

        def on_item_double_clicked(item: QTreeWidgetItem, _col: int):
            node_id = str(item.data(0, Qt.ItemDataRole.UserRole) or "")
            node, _ = self._find_accessory_node(node_id)
            if not node or node.get("node_type") != "leaf":
                return
            nxt = Qt.CheckState.Unchecked if item.checkState(0) == Qt.CheckState.Checked else Qt.CheckState.Checked
            item.setCheckState(0, nxt)

        def render():
            kw = search.text().strip().lower()
            tree.clear()

            def matches_or_has_match(node: dict[str, Any]) -> bool:
                node_text = f"{node.get('name','')} {node.get('desc','')} {node.get('url','')}".lower()
                if not kw or kw in node_text:
                    return True
                for ch in node.get("children", []) or []:
                    if isinstance(ch, dict) and matches_or_has_match(ch):
                        return True
                return False

            def add_node(node: dict[str, Any], parent_item: QTreeWidgetItem | None):
                if not matches_or_has_match(node):
                    return
                item = QTreeWidgetItem(
                    [
                        str(node.get("name", "")),
                        str(node.get("url", "")),
                    ]
                )
                item.setData(0, Qt.ItemDataRole.UserRole, str(node.get("id", "")))
                if node.get("node_type") == "leaf":
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    desc_name = str(node.get("desc", "") or node.get("name", "") or "")
                    node_url = str(node.get("url", "") or "").strip()
                    checked = desc_name in pre_values or node_url in pre_values
                    item.setCheckState(0, Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)
                if parent_item is None:
                    tree.addTopLevelItem(item)
                else:
                    parent_item.addChild(item)
                for ch in node.get("children", []) or []:
                    if isinstance(ch, dict):
                        add_node(ch, item)

            add_node(self.accessories_root, None)
            # 首次选择（无已选且无搜索）默认全收缩；其余仅展开 URL 命中（或已勾选）路径。
            tree.collapseAll()
            if not kw and not pre_values:
                return

            def mark_expand_path(it: QTreeWidgetItem) -> bool:
                has_child_match = False
                for i in range(it.childCount()):
                    if mark_expand_path(it.child(i)):
                        has_child_match = True
                node_id = str(it.data(0, Qt.ItemDataRole.UserRole) or "")
                node, _ = self._find_accessory_node(node_id)
                self_match = False
                if node and node.get("node_type") == "leaf":
                    node_url = str(node.get("url", "") or "").strip().lower()
                    desc_name = str(node.get("desc", "") or node.get("name", "") or "").strip()
                    checked = it.checkState(0) == Qt.CheckState.Checked or desc_name in pre_values
                    if kw:
                        self_match = kw in node_url
                    else:
                        self_match = checked
                should_expand = has_child_match or self_match
                if should_expand and it.childCount() > 0:
                    it.setExpanded(True)
                return should_expand

            for i in range(tree.topLevelItemCount()):
                mark_expand_path(tree.topLevelItem(i))

        def on_accept():
            picked_urls: list[str] = []
            picked_descs: list[str] = []

            def walk(item: QTreeWidgetItem):
                node_id = str(item.data(0, Qt.ItemDataRole.UserRole) or "")
                node, _ = self._find_accessory_node(node_id)
                if node and node.get("node_type") == "leaf" and item.checkState(0) == Qt.CheckState.Checked:
                    url = str(node.get("url", "") or "").strip()
                    desc = str(node.get("desc", "") or node.get("name", "") or "").strip()
                    if url:
                        picked_urls.append(url)
                        picked_descs.append(desc or url)
                for i in range(item.childCount()):
                    walk(item.child(i))

            for i in range(tree.topLevelItemCount()):
                walk(tree.topLevelItem(i))

            # 去重且保序
            seen_u: set[str] = set()
            final_urls: list[str] = []
            for u in picked_urls:
                if u not in seen_u:
                    seen_u.add(u)
                    final_urls.append(u)
            seen_d: set[str] = set()
            final_descs: list[str] = []
            for d in picked_descs:
                if d not in seen_d:
                    seen_d.add(d)
                    final_descs.append(d)
            if not final_urls:
                QMessageBox.information(dlg, "提示", "请至少勾选一个描述")
                return
            picked["urls"] = final_urls
            picked["descs"] = final_descs
            dlg.accept()

        search.textChanged.connect(render)
        tree.itemDoubleClicked.connect(on_item_double_clicked)
        btns.accepted.connect(on_accept)
        btns.rejected.connect(dlg.reject)
        render()
        anchor = anchor_widget.property("_url_anchor_global")
        width = anchor_widget.property("_url_anchor_width")
        if isinstance(anchor, QPoint) and isinstance(width, int):
            self._place_dialog_below_rect(dlg, anchor, width)
        else:
            p = anchor_widget.mapToGlobal(anchor_widget.rect().bottomLeft())
            self._place_dialog_below_rect(dlg, p, anchor_widget.width())
        if dlg.exec() == QDialog.Accepted:
            self.apply_field_update(rec_idx, "url", "\n".join(picked["urls"]))

    def _write_merge_stats_sheet(self, wb: Workbook, merged_rows: list[list[Any]]) -> None:
        """统计页来源于 sum-template.xlsx 的统计页模板，再按规则填充数据。"""

        def _copy_sheet_template(src_ws, dst_ws):
            # 复制基础表级属性
            dst_ws.sheet_format = copy(src_ws.sheet_format)
            dst_ws.sheet_properties = copy(src_ws.sheet_properties)
            dst_ws.page_margins = copy(src_ws.page_margins)
            dst_ws.page_setup = copy(src_ws.page_setup)
            dst_ws.print_options = copy(src_ws.print_options)
            dst_ws.sheet_view = copy(src_ws.sheet_view)
            dst_ws.views = copy(src_ws.views)
            dst_ws.protection = copy(src_ws.protection)
            dst_ws.sheet_state = src_ws.sheet_state
            dst_ws.auto_filter.ref = src_ws.auto_filter.ref
            dst_ws.freeze_panes = src_ws.freeze_panes
            # 复制打印区域/打印标题（跨系统显示差异较大，尽量保真）
            try:
                dst_ws.print_title_rows = src_ws.print_title_rows
            except Exception:
                pass
            try:
                dst_ws.print_title_cols = src_ws.print_title_cols
            except Exception:
                pass
            try:
                dst_ws.print_area = src_ws.print_area
            except Exception:
                pass
            # 复制列宽/隐藏/分组
            for col_key, dim in src_ws.column_dimensions.items():
                dst = dst_ws.column_dimensions[col_key]
                dst.width = dim.width
                dst.hidden = dim.hidden
                dst.outlineLevel = dim.outlineLevel
                dst.bestFit = dim.bestFit
                dst.collapsed = dim.collapsed
                dst.style = dim.style
            # 复制行高/隐藏/分组
            for row_idx, dim in src_ws.row_dimensions.items():
                dst = dst_ws.row_dimensions[row_idx]
                dst.height = dim.height
                dst.hidden = dim.hidden
                dst.outlineLevel = dim.outlineLevel
                dst.collapsed = dim.collapsed
                dst.style = dim.style
            # 复制单元格值与样式（避免直接拷贝 _style 造成样式索引错乱）
            for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, min_col=1, max_col=src_ws.max_column):
                for cell in row:
                    dcell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    dcell.font = copy(cell.font)
                    dcell.fill = copy(cell.fill)
                    dcell.border = copy(cell.border)
                    dcell.protection = copy(cell.protection)
                    dcell.alignment = copy(cell.alignment)
                    dcell.number_format = cell.number_format
            # 复制合并单元格
            for rg in src_ws.merged_cells.ranges:
                dst_ws.merge_cells(str(rg))
            # 复制条件格式（Mac 端显示差异常见于此）
            try:
                for cf_range, rules in src_ws.conditional_formatting._cf_rules.items():
                    for rule in rules:
                        dst_ws.conditional_formatting.add(str(cf_range), copy(rule))
            except Exception:
                pass
            # 复制数据验证
            try:
                for dv in src_ws.data_validations.dataValidation:
                    dst_ws.add_data_validation(deepcopy(dv))
            except Exception:
                pass

        if "统计" in wb.sheetnames:
            del wb["统计"]

        sum_tpl = SUM_TEMPLATE_FILE
        if sum_tpl.exists():
            try:
                sum_wb = load_workbook(str(sum_tpl))
                try:
                    src_ws = sum_wb["统计"] if "统计" in sum_wb.sheetnames else sum_wb[sum_wb.sheetnames[0]]
                    ws = wb.create_sheet("统计")
                    _copy_sheet_template(src_ws, ws)
                finally:
                    sum_wb.close()
            except Exception:
                ws = wb.create_sheet("统计")
        else:
            ws = wb.create_sheet("统计")

        ws["C1"] = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        counts: Counter[str] = Counter()
        for vals in merged_rows:
            if len(vals) < 2:
                continue
            name = str(vals[1] or "").strip()
            if name:
                counts[name] += 1
        row = 4
        serial = 1
        for name, cnt in counts.most_common():
            ws.cell(row=row, column=2, value=serial)  # B列 序号
            ws.cell(row=row, column=3, value=name)    # C列 任务名
            ws.cell(row=row, column=4, value=cnt)     # D列 统计行数
            row += 1
            serial += 1
        # 默认打开定位到统计页 A1
        wb.active = wb.sheetnames.index(ws.title)
        try:
            ws.sheet_view.topLeftCell = "A1"
            if ws.sheet_view.selection:
                ws.sheet_view.selection[0].activeCell = "A1"
                ws.sheet_view.selection[0].sqref = "A1"
        except Exception:
            pass

    def _refresh_main_row_validation_marks(self, rec_idx: int):
        """按导出规则刷新主表一行标红；不弹窗（仅打印按钮失败时弹窗）。"""
        if rec_idx < 0 or rec_idx >= len(self.records):
            return
        rec = self.records[rec_idx]
        self._clear_main_row_validation_errors(rec_idx)
        issues = self.collect_record_validation_issues(rec)
        if issues:
            self._mark_row_from_validation_issues(rec_idx, issues)

    def _format_validation_issues_message(self, issues: list[tuple[str, str]], *, with_title: bool = True) -> str:
        if not issues:
            return ""
        merged = self._merge_validation_issues_by_field(issues)
        order = [f for f in FIELDS if f in merged]
        for f in merged:
            if f not in order:
                order.append(f)
        lines = [f"• {HEADERS.get(f, f)}：{merged[f]}" for f in order]
        body = "\n".join(lines)
        if with_title:
            return "请修正以下问题：\n\n" + body
        return body

    def region_part(self, rec):
        p, c = split_multi(rec.get("province", "")), split_multi(rec.get("city", ""))
        if p and not c:
            return p[0] if len(p) == 1 else f"{int_to_cn(len(p))}省"
        if c and not p:
            return c[0] if len(c) == 1 else f"{int_to_cn(len(c))}市"
        return "全国"

    def _try_parse_left_region_six(self, left: str) -> tuple[dict | None, list[str]]:
        """地域（最前中文地域词）+ 紧随 6 位字母数字码；失败时返回 (None, 错误文案列表)。"""
        reg = find_earliest_region_in_left(left)
        if not reg:
            return None, []
        start, end, region = reg
        tail = left[end:].lstrip()
        if len(tail) < 6 or not re.match(r"^[A-Za-z0-9]{6}$", tail[:6]):
            return None, []
        code6 = tail[:6]
        op, tp, ind = code6[0:2], code6[2:4], code6[4:6]
        errs: list[str] = []
        if op not in OP_MAP:
            errs.append(f"运营商编码「{op}」输入不正确")
        if tp not in TYPE_MAP:
            errs.append(f"类型编码「{tp}」输入不正确")
        if errs:
            return None, errs
        customer_raw = (left[:start] + tail[6:]).strip()
        if region_in_task_name_requires_customer_prefix(region) and not customer_raw:
            return None, ["请输入任务编号"]
        customer = customer_raw or "客户"
        return {"customer": customer, "region": region, "op": op, "tp": tp, "ind": ind}, []

    def parse_left_with_errors(self, left: str) -> tuple[dict | None, list[str]]:
        left = (left or "").strip()
        if not left:
            return None, []
        v2, errs = self._try_parse_left_region_six(left)
        if errs:
            return None, errs
        if v2:
            return v2, []
        m = re.search(r"(YD|LT|DX|YX)(DB|DJ|XC|DH|DY)([A-Za-z0-9]+)$", left)
        if not m:
            return None, []
        prefix = left[: m.start()]
        region, customer = "", prefix
        for token in sorted(PROVINCES + CITIES + ["全国"], key=len, reverse=True):
            if prefix.endswith(token):
                customer, region = prefix[: -len(token)], token
                break
        if not region:
            rm = re.search(r"([一二两三四五六七八九十\d]+[省市])$", prefix)
            if rm:
                region = rm.group(1)
                customer = prefix[: -len(region)]
        if region_in_task_name_requires_customer_prefix(region) and not str(customer or "").strip():
            return None, ["请输入任务编号"]
        return {"customer": customer or "客户", "region": region, "op": m.group(1), "tp": m.group(2), "ind": m.group(3)}, []

    def parse_left(self, left):
        d, _ = self.parse_left_with_errors(left)
        return d

    def build_task_name(self, rec):
        old = rec.get("task_name", "")
        product = old.split("-", 1)[1] if "-" in old else "产品"
        parsed = self.parse_left(old.split("-", 1)[0]) if "-" in old else None
        customer = parsed["customer"] if parsed else "客户"
        return (
            f"{customer}{self.region_part(rec)}{rec.get('operator_code') or '##'}"
            f"{rec.get('type_code') or '##'}{rec.get('industry_code') or '##'}-{product}"
        )

    def parse_task_name(self, rec, rec_idx: int | None = None):
        self._task_name_parse_err_marked = False
        tn = rec.get("task_name", "")
        if "-" not in tn:
            return False
        left = tn.split("-", 1)[0]
        snap = {k: rec.get(k, "") for k in ("operator_code", "type_code", "industry_code", "province", "city")}
        parsed, errs = self.parse_left_with_errors(left)
        if errs:
            for k, v in snap.items():
                rec[k] = v
            if rec_idx is not None:
                self._task_name_parse_err_marked = True
            return False
        if not parsed:
            return False
        rec["operator_code"], rec["type_code"], rec["industry_code"] = parsed["op"], parsed["tp"], parsed["ind"]
        region = parsed["region"]
        if not region or region == "全国" or re.fullmatch(r"[一二两三四五六七八九十\d]+[省市]", region):
            rec["province"], rec["city"] = "", ""
        elif region in PROVINCES:
            rec["province"], rec["city"] = region, ""
        elif region in CITIES:
            rec["city"], rec["province"] = region, ""
        return True

    def after_field_change(self, rec, field, rec_idx: int | None = None):
        if field == "province" and rec.get("province", "").strip():
            rec["city"] = ""
            rec["exclude_province"] = ""
            provs = split_multi(rec.get("province", ""))
            allow = set(cities_under_provinces(provs)) if provs else set()
            if allow:
                rec["exclude_city"] = "|".join(x for x in split_multi(rec.get("exclude_city", "")) if x in allow)
            else:
                rec["exclude_city"] = ""
        elif field == "city" and rec.get("city", "").strip():
            rec["province"] = ""
            rec["exclude_province"] = ""
            rec["exclude_city"] = ""
        if field == "exclude_province" and rec.get("exclude_province", "").strip():
            rec["province"] = ""
            rec["city"] = ""
        if field == "exclude_city" and rec.get("exclude_city", "").strip():
            rec["city"] = ""
        if field == "task_name":
            return self.parse_task_name(rec, rec_idx)
        elif field in ("province", "city", "type_code", "operator_code", "industry_code"):
            rec["task_name"] = self.build_task_name(rec)
        return True

    def export_customer_name(self, rec):
        task_name = str(rec.get("task_name", "")).strip()
        left = task_name.split("-", 1)[0] if "-" in task_name else task_name
        parsed = self.parse_left(left) if left else None
        if parsed and str(parsed.get("customer", "")).strip():
            return sanitize_filename(str(parsed["customer"])[:12])
        fallback = left[:12] if left else "客户"
        return sanitize_filename(fallback)

    def export_excel(self):
        selected_with_index = [(idx, r) for idx, r in enumerate(self.records) if r.get("checked")]
        if not selected_with_index:
            QMessageBox.information(self, "提示", "请先勾选至少一条记录")
            return
        confirm_box = QMessageBox(self)
        confirm_box.setIcon(QMessageBox.Icon.Question)
        confirm_box.setWindowTitle("确认打印")
        confirm_box.setText(f"确定要把选中的这 {len(selected_with_index)} 条数据打印到 Excel 文件？")
        btn_ok = confirm_box.addButton("OK", QMessageBox.ButtonRole.AcceptRole)
        confirm_box.addButton("取消", QMessageBox.ButtonRole.RejectRole)
        confirm_box.exec()
        if confirm_box.clickedButton() is not btn_ok:
            return
        include_url = self.chk_print_url.isChecked()
        selected = [x[1] for x in selected_with_index]
        # 导出前做一次全量校验；仅在此处对校验失败弹窗（输入过程中只标红不弹窗）。
        issues_by_idx: dict[int, list[tuple[str, str]]] = {}
        has_invalid = False
        for rec_idx, rec in selected_with_index:
            self._clear_main_row_validation_errors(rec_idx)
            issues = self.collect_record_validation_issues(rec)
            if not issues:
                continue
            has_invalid = True
            issues_by_idx[rec_idx] = issues
            self._mark_row_from_validation_issues(rec_idx, issues)
        if has_invalid:
            self.refresh_table()
            parts = ["勾选记录存在以下问题（已标红对应字段），请修正后再点击打印。\n"]
            for rec_idx, rec in selected_with_index:
                issues = issues_by_idx.get(rec_idx)
                if not issues:
                    continue
                label = (str(rec.get("task_name", "")).strip()[:40] or f"第 {rec_idx + 1} 条").replace("\n", " ")
                parts.append("\n————————————————")
                parts.append(f"【{label}】")
                parts.append(self._format_validation_issues_message(issues, with_title=False))
            full = "\n".join(parts)
            if len(full) > 12000:
                full = full[:12000] + "\n\n……（以下节选，请以表格标红为准）"
            QMessageBox.warning(self, "无法打印", full)
            return
        if not TEMPLATE_FILE.exists():
            QMessageBox.warning(self, "导出失败", f"模板不存在: {TEMPLATE_FILE.name}")
            return
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb[wb.sheetnames[0]]
        col_map = {"task_name": "A", "type_code": "B", "operator_code": "C", "industry_code": "D", "url": "E", "quantity": "F", "duration": "G", "age_max": "H", "age_min": "I", "pv": "J", "province": "K", "exclude_province": "L", "city": "M", "exclude_city": "N"}
        start_row = 2
        for r in range(start_row, ws.max_row + 1):
            for f, c in col_map.items():
                ws[f"{c}{r}"].value = None
        for idx, rec in enumerate(selected):
            rr = start_row + idx
            for f, c in col_map.items():
                if f == "url":
                    raw_descs = str(rec.get("url", ""))
                    row_allow = coerce_allow_print_url(rec.get("allow_print_url"))
                    if include_url and row_allow:
                        desc_map = self._accessory_desc_url_map()
                        raw_lines = [
                            x.strip()
                            for x in raw_descs.replace("\r\n", "\n").replace("\r", "\n").replace("|", "\n").split("\n")
                            if x.strip()
                        ]
                        lines: list[str] = []
                        for line in raw_lines:
                            # 兼容历史数据（保存的是描述）与新数据（保存的是 URL）。
                            lines.append(desc_map.get(line, line))
                        cell = ws[f"{c}{rr}"]
                        # URL 以多行写入单元格，便于在 Excel 中直接逐行查看。
                        cell.value = "\r\n".join(lines)
                        ali = copy(cell.alignment)
                        ali.wrap_text = True
                        cell.alignment = ali
                    else:
                        cell = ws[f"{c}{rr}"]
                        cell.value = ""
                        ali = copy(cell.alignment)
                        ali.wrap_text = True
                        cell.alignment = ali
                elif f == "type_code":
                    ws[f"{c}{rr}"] = TYPE_MAP.get(rec.get(f, ""), "")
                elif f == "operator_code":
                    ws[f"{c}{rr}"] = OP_MAP.get(rec.get(f, ""), "")
                else:
                    ws[f"{c}{rr}"] = rec.get(f, "")
        customer = self.export_customer_name(selected[0])
        filename = f"{customer}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        base_dir = self._default_output_dir_for("print")
        try:
            base_dir.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            QMessageBox.warning(self, "导出失败", f"创建目录失败：\n{base_dir}\n\n{e}")
            return
        default_out = base_dir / filename
        out_file, _ = QFileDialog.getSaveFileName(
            self,
            "选择保存位置并命名文件",
            str(default_out),
            "Excel (*.xlsx)",
        )
        if not out_file:
            return
        out_path = Path(out_file)
        if out_path.suffix.lower() != ".xlsx":
            out_path = out_path.with_suffix(".xlsx")
        try:
            out_path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            QMessageBox.warning(self, "导出失败", f"创建目录失败：\n{out_path.parent}\n\n{e}")
            return
        try:
            wb.save(str(out_path))
        except OSError as e:
            QMessageBox.warning(self, "导出失败", f"保存文件失败：\n{out_path}\n\n{e}")
            return
        out_file = str(out_path)
        stamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        for rec_idx, _r in selected_with_index:
            rr = self.records[rec_idx]
            try:
                c = int(rr.get("print_count", 0) or 0)
            except (TypeError, ValueError):
                c = 0
            rr["print_count"] = c + 1
            rr["last_printed_at"] = stamp
        self.save_data()
        self.refresh_table()
        self.print_records.insert(
            0,
            {
                "id": str(uuid.uuid4()),
                "path": out_file,
                "filename": os.path.basename(out_file),
                "source": "打印",
                "printed_at": stamp,
                "row_count": len(selected),
                "include_print_url": include_url,
            },
        )
        self.save_print_records()
        if self.content_stack.currentIndex() == 2:
            self.refresh_print_records_table()
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.Information)
        box.setWindowTitle("导出成功")
        box.setText(f"已导出: {out_file}")
        box.setInformativeText(
            "<span style='color:#2563eb;'>"
            "可点击“打开文件夹”快速定位到该导出文件。"
            "<br/>也可以到「历史打印记录」页面中打开该文件所在文件夹。"
            "</span>"
        )
        btn_open_dir = box.addButton("打开文件夹", QMessageBox.ButtonRole.ActionRole)
        box.addButton(QMessageBox.StandardButton.Ok)
        box.exec()
        if box.clickedButton() is btn_open_dir:
            self.open_print_record_folder(out_file)


def _show_license_expired_dialog(parent: QWidget | None = None) -> None:
    """无控制台（pythonw）时也必须能看见提示；运行中过期时 parent 为主窗口。"""
    box = QMessageBox(parent)
    box.setIcon(QMessageBox.Icon.Critical)
    box.setWindowTitle("授权已过期")
    box.setText(LICENSE_EXPIRED_MSG)
    box.setStandardButtons(QMessageBox.StandardButton.Ok)
    box.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
    box.exec()


def main():
    app = QApplication(sys.argv)
    if not is_license_valid():
        _show_license_expired_dialog()
        sys.exit(0)
    ui_font = QFont()
    ui_font.setFamilies(["Segoe UI Variable", "Segoe UI", "Microsoft YaHei UI", "PingFang SC"])
    ui_font.setPointSize(10)
    ui_font.setWeight(QFont.Weight.Medium)
    app.setFont(ui_font)
    try:
        win = BillApp()
        win.show()
    except Exception:
        import traceback

        err_path = Path(os.environ.get("TEMP", os.environ.get("TMP", "."))) / "TidanMgr_startup_error.log"
        try:
            err_path.write_text(traceback.format_exc(), encoding="utf-8")
        except OSError:
            pass
        QMessageBox.critical(
            None,
            "启动失败",
            f"程序启动异常，详情已尝试写入：\n{err_path}\n\n{traceback.format_exc()[-800:]}",
        )
        sys.exit(1)
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
